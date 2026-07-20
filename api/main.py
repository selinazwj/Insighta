from fastapi import FastAPI, APIRouter, Request, Form, Depends, HTTPException, Cookie, UploadFile, File, Query, BackgroundTasks
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse, PlainTextResponse, Response as HTTPResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.gzip import GZipMiddleware
from sqlalchemy.orm import Session
from passlib.context import CryptContext
from sqlalchemy import Column, Integer, String, Float, DateTime, ForeignKey, func, JSON, Boolean, case, inspect, text, or_
from pathlib import Path
from typing import Any, List, Optional
from datetime import datetime, timedelta, timezone, date, time
from urllib.parse import urlencode
from xml.sax.saxutils import escape as xml_escape
from io import BytesIO
import random
import re
import shutil
import uuid
import os
import secrets
import hashlib
import hmac
import html
import json
import stripe
import smtplib
import httpx
import asyncio
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from dotenv import load_dotenv

load_dotenv(Path(__file__).resolve().parent.parent / ".env")

from app.database import engine, get_db, SessionLocal
from app.models import Base, User, Survey, Response, Feedback, Notification, EmailVerificationCode, Question, Answer, ResponseQualityCheck, QualityBlacklist, SupportThread, SupportMessage, UserEvent
from app.quality_engine import (
    anthropic_api_key_configured,
    batch_anomaly_scores_for_features,
    compute_excel_row_quality,
    create_excel_quality_check,
    _duration_seconds_between,
    apply_auto_approve_checks,
    ensure_builtin_quality_checks,
    evaluate_builtin_response,
    resolve_excel_row_context,
    upsert_builtin_quality_check,
)
from app.payouts import (
    APPROVED,
    LEGACY_RELEASED,
    mark_response_under_review,
    reject_response_payout,
    release_response_payout,
    return_response_to_review,
)
from app.verification.routes import router as verification_router
from app.ai_growth.routes import router as ai_growth_router
from app.ai_growth.security import is_safe_internal_next
from app.ai_growth.jump import mark_latest_jump_completed_for_response
from app.ai_growth.prediction import recommend_surveys_for_user
from app.ai_growth.models import JumpEvent, RespondentPrediction, SurveySegmentStats, UserActivityEvent
from app.discovery.router import router as discovery_router
from app.discovery.discovery import discover as discover_channels
from app.discovery.models import Criteria as DiscoveryCriteria
from app.discovery.ranking import rank as rank_discovery_channels
from app.seo import (
    CATEGORY_CONTENT,
    INDEX_PUBLIC_STUDIES,
    SITE_LANGUAGE,
    category_content,
    category_image,
    category_label,
    content_page_seo,
    home_seo,
    participant_seo,
    plain_text,
    site_url as seo_site_url,
    studies_directory_seo,
    category_seo,
    study_seo,
)

app = FastAPI(title="Insighta")
app.add_middleware(GZipMiddleware, minimum_size=1000)
app.include_router(verification_router)
app.include_router(ai_growth_router)
app.include_router(discovery_router)

BASE_DIR = Path(__file__).resolve().parent
templates = Jinja2Templates(directory="app/templates")
app.mount("/static", StaticFiles(directory="app/static"), name="static")

SEO_PUBLIC_EXACT_PATHS = {
    "/",
    "/participant",
    "/studies",
    "/about",
    "/privacy",
    "/terms",
    "/robots.txt",
    "/sitemap.xml",
}

def _is_public_seo_path(path: str) -> bool:
    if path in SEO_PUBLIC_EXACT_PATHS:
        return True
    if path.startswith("/studies/"):
        return True
    if path.startswith("/r/") and not path.endswith("/qr.png"):
        return True
    return False

@app.middleware("http")
async def apply_seo_and_delivery_headers(request: Request, call_next):
    response = await call_next(request)
    path = request.url.path

    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    if response.headers.get("content-type", "").startswith("text/html"):
        response.headers.setdefault("Content-Language", SITE_LANGUAGE)

    if path.startswith("/static/"):
        # Assets are not fingerprinted, so use a moderate cache lifetime rather than immutable caching.
        response.headers.setdefault("Cache-Control", "public, max-age=86400, stale-while-revalidate=604800")
    elif response.status_code >= 400:
        response.headers.setdefault("X-Robots-Tag", "noindex, nofollow, noarchive")
    elif not _is_public_seo_path(path):
        # Defense in depth: the Jinja partial also emits noindex for non-public HTML pages.
        response.headers.setdefault("X-Robots-Tag", "noindex, nofollow, noarchive")

    return response

def no_store_response(response):
    response.headers["Cache-Control"] = "no-store, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
Base.metadata.create_all(bind=engine)

def ensure_user_profile_columns():
    columns = {col["name"] for col in inspect(engine).get_columns("users")}
    profile_columns = {
        "first_name": "VARCHAR",
        "last_name": "VARCHAR",
        "welcome_email_sent_at": "TIMESTAMP",
        "welcome_email_role": "VARCHAR",
        "phone_number": "VARCHAR",
        "birth_year": "VARCHAR",
        "birth_month": "VARCHAR",
        "profile_description": "VARCHAR",
        "current_country": "VARCHAR",
        "current_province": "VARCHAR",
        "current_city": "VARCHAR",
        "origin_country": "VARCHAR",
        "origin_province": "VARCHAR",
        "origin_city": "VARCHAR",
        "race": "VARCHAR",
        "income_level": "VARCHAR",
        "lifestyle_tags": "VARCHAR",
        "referral_code": "VARCHAR",
        "invited_by_user_id": "INTEGER",
        "student_status": "VARCHAR",
        "year_in_school": "VARCHAR",
        "international_domestic": "VARCHAR",
        "experience_tags": "VARCHAR",
        "participation_format": "VARCHAR",
        "device_type": "VARCHAR",
        "oauth_provider": "VARCHAR",
        "oauth_id": "VARCHAR",
        "stripe_account_id": "VARCHAR",
        "stripe_onboarding_complete": "VARCHAR",
        "pending_earnings": "FLOAT",
        "total_withdrawn": "FLOAT",
    }
    with engine.begin() as conn:
        for name, column_type in profile_columns.items():
            if name not in columns:
                conn.execute(text(f"ALTER TABLE users ADD COLUMN {name} {column_type}"))
        # Unique index for referral codes (ignore if already present)
        try:
            conn.execute(text(
                "CREATE UNIQUE INDEX IF NOT EXISTS ix_users_referral_code ON users (referral_code)"
            ))
        except Exception:
            pass

ensure_user_profile_columns()

def ensure_survey_listing_columns():
    columns = {col["name"] for col in inspect(engine).get_columns("surveys")}
    listing_columns = {
        "target_student_status": "VARCHAR",
        "target_year_in_school": "VARCHAR",
        "target_international_domestic": "VARCHAR",
        "target_experience_tags": "VARCHAR",
        "target_participation_format": "VARCHAR",
        "target_device": "VARCHAR",
        "urgency_level": "VARCHAR",
        "incentive_type": "VARCHAR",
        "task_type": "VARCHAR",
        "target_income_level": "VARCHAR",
        "target_lifestyle_tags": "VARCHAR",
        "target_niche_requirements": "VARCHAR",
        "participant_benefits": "VARCHAR",
        "raffle_prize_type": "VARCHAR",
        "availability_slots": "TEXT",
        "interview_location": "VARCHAR",
        "session_count": "INTEGER",
        "sessions_per_week": "INTEGER",
        "admin_display_reward_amount": "FLOAT",
        "share_slug": "VARCHAR",
        "total_budget": "FLOAT",
        "per_person_gross": "FLOAT",
        "commission_rate": "FLOAT",
        "payment_status": "VARCHAR",
        "stripe_payment_intent_id": "VARCHAR",
        "quality_auto_filter_enabled": "BOOLEAN DEFAULT false",
        "quality_auto_filter_min_score": "FLOAT DEFAULT 80.0",
    }
    with engine.begin() as conn:
        for name, column_type in listing_columns.items():
            if name not in columns:
                conn.execute(text(f"ALTER TABLE surveys ADD COLUMN {name} {column_type}"))

ensure_survey_listing_columns()

def ensure_response_tracking_columns():
    columns = {col["name"] for col in inspect(engine).get_columns("responses")}
    response_columns = {
        "client_ip": "VARCHAR",
        "user_agent": "VARCHAR",
        "device_fingerprint": "VARCHAR",
        "booking_slot": "TEXT",
        "start_followup_scheduled_at": "TIMESTAMP",
        "start_followup_sent_at": "TIMESTAMP",
        "payout_status": "VARCHAR",
        "payout_amount": "FLOAT",
        "stripe_transfer_id": "VARCHAR",
    }
    with engine.begin() as conn:
        for name, column_type in response_columns.items():
            if name not in columns:
                conn.execute(text(f"ALTER TABLE responses ADD COLUMN {name} {column_type}"))

ensure_response_tracking_columns()

def ensure_user_event_table():
    id_type = "SERIAL PRIMARY KEY" if engine.dialect.name == "postgresql" else "INTEGER PRIMARY KEY AUTOINCREMENT"
    metadata_type = "JSONB" if engine.dialect.name == "postgresql" else "JSON"
    with engine.begin() as conn:
        conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS user_events (
                id {id_type},
                user_id INTEGER NULL REFERENCES users(id) ON DELETE SET NULL,
                anonymous_id VARCHAR NULL,
                event_name VARCHAR NOT NULL,
                target_type VARCHAR NULL,
                target_id VARCHAR NULL,
                page_path TEXT NULL,
                metadata_json {metadata_type} NULL,
                user_agent TEXT NULL,
                client_ip VARCHAR NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """))

ensure_user_event_table()

stripe.api_key = os.environ.get("STRIPE_SECRET_KEY")
STRIPE_PUBLISHABLE_KEY = os.environ.get("STRIPE_PUBLISHABLE_KEY")
STRIPE_WEBHOOK_SECRET = os.environ.get("STRIPE_WEBHOOK_SECRET")

EMAIL_ADDRESS = os.environ.get("EMAIL_ADDRESS")
EMAIL_PASSWORD = os.environ.get("EMAIL_PASSWORD")
EMAIL_PROVIDER = os.environ.get("EMAIL_PROVIDER", "gmail").strip().lower()
RESEND_API_KEY = os.environ.get("RESEND_API_KEY")
EMAIL_FROM = os.environ.get("EMAIL_FROM") or EMAIL_ADDRESS
EMAIL_REPLY_TO = os.environ.get("EMAIL_REPLY_TO")
SUPPORT_ALERT_EMAIL = os.environ.get("SUPPORT_ALERT_EMAIL", "vfsa@bu.edu")
VERIFICATION_CODE_EXPIRE_MINUTES = 10
VERIFICATION_CODE_RESEND_COOLDOWN_SECONDS = 60
VERIFICATION_CODE_MAX_PER_HOUR = 5
SURVEY_START_FOLLOWUP_DELAY_MINUTES = int(os.environ.get("SURVEY_START_FOLLOWUP_DELAY_MINUTES", "5"))
SURVEY_START_FOLLOWUP_POLL_SECONDS = int(os.environ.get("SURVEY_START_FOLLOWUP_POLL_SECONDS", "60"))

# ---------------------------
# OAuth 2.0 configuration
# ---------------------------
BASE_URL = os.environ.get("BASE_URL", "https://insightaco.org")

GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID")
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET")
GOOGLE_AUTH_URL = "https://accounts.google.com/o/oauth2/v2/auth"
GOOGLE_TOKEN_URL = "https://oauth2.googleapis.com/token"
GOOGLE_USERINFO_URL = "https://www.googleapis.com/oauth2/v2/userinfo"

LINKEDIN_CLIENT_ID = os.environ.get("LINKEDIN_CLIENT_ID")
LINKEDIN_CLIENT_SECRET = os.environ.get("LINKEDIN_CLIENT_SECRET")
LINKEDIN_AUTH_URL = "https://www.linkedin.com/oauth/v2/authorization"
LINKEDIN_TOKEN_URL = "https://www.linkedin.com/oauth/v2/accessToken"
LINKEDIN_USERINFO_URL = "https://api.linkedin.com/v2/userinfo"

def _email_plain_text(body: str) -> str:
    text_body = re.sub(r"(?i)<br\s*/?>", "\n", body or "")
    text_body = re.sub(r"(?i)</p\s*>", "\n\n", text_body)
    text_body = re.sub(r"<[^>]+>", "", text_body)
    return html.unescape(text_body).strip()

def _email_brand_header() -> str:
    logo_url = f"{BASE_URL.rstrip('/')}/static/favicon.png"
    return f"""
    <div style="max-width:620px;margin:0 auto;padding:24px 22px 0;font-family:Arial,sans-serif;">
      <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;border-bottom:1px solid #eeece6;padding-bottom:16px;">
        <tr>
          <td width="48" style="padding:0 10px 16px 0;vertical-align:middle;">
            <img src="{html.escape(logo_url, quote=True)}" alt="Insighta" width="40" height="40" style="display:block;width:40px;height:40px;object-fit:contain;border:0;">
          </td>
          <td style="padding:0 0 16px 0;vertical-align:middle;font-size:22px;font-weight:700;letter-spacing:-0.3px;color:#184e77;line-height:1;font-family:Arial,sans-serif;">Insighta</td>
        </tr>
      </table>
    </div>
    """

def _with_email_brand_header(body: str) -> str:
    if "data-insighta-email-brand" in (body or ""):
        return body
    return f"""
    <div data-insighta-email-brand="1" style="margin:0;padding:0;background:#ffffff;">
      {_email_brand_header()}
      {body or ""}
    </div>
    """

def send_email(to: str, subject: str, body: str, text_body: Optional[str] = None) -> tuple[bool, Optional[str]]:
    plain_text = text_body or _email_plain_text(body)
    html_body = _with_email_brand_header(body)
    if EMAIL_PROVIDER == "resend":
        if not RESEND_API_KEY or not EMAIL_FROM:
            return False, "Resend email configuration is incomplete"
        payload = {
            "from": EMAIL_FROM,
            "to": [to],
            "subject": subject,
            "html": html_body,
            "text": plain_text,
        }
        if EMAIL_REPLY_TO:
            payload["reply_to"] = EMAIL_REPLY_TO
        try:
            response = httpx.post(
                "https://api.resend.com/emails",
                headers={
                    "Authorization": f"Bearer {RESEND_API_KEY}",
                    "Content-Type": "application/json",
                    "User-Agent": "Insighta/1.0",
                },
                json=payload,
                timeout=8.0,
            )
            if response.is_success:
                return True, None
            error = f"Resend returned {response.status_code}: {response.text[:500]}"
            print(f"Email error: {error}")
            return False, error
        except Exception as exc:
            error = f"Resend request failed: {exc}"
            print(f"Email error: {error}")
            return False, error

    if not EMAIL_ADDRESS or not EMAIL_PASSWORD:
        return False, "Gmail email configuration is incomplete"
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = EMAIL_ADDRESS
        msg["To"] = to
        if EMAIL_REPLY_TO:
            msg["Reply-To"] = EMAIL_REPLY_TO
        msg.attach(MIMEText(plain_text, "plain"))
        msg.attach(MIMEText(html_body, "html"))
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            server.sendmail(EMAIL_ADDRESS, to, msg.as_string())
        return True, None
    except Exception as exc:
        error = f"Gmail SMTP failed: {exc}"
        print(f"Email error: {error}")
        return False, error

def send_support_alert_email(user: User, thread: SupportThread, message: SupportMessage):
    if not SUPPORT_ALERT_EMAIL:
        return
    user_label = html.escape(user.email or user.username or f"User #{user.id}")
    user_name = html.escape(user.username or "")
    subject_user = (user.email or "a user").replace("\r", " ").replace("\n", " ")
    message_body = html.escape(message.body or "").replace("\n", "<br>")
    admin_url = f"{BASE_URL.rstrip('/')}/admin#support"
    send_email(
        to=SUPPORT_ALERT_EMAIL,
        subject=f"[Insighta Support] New message from {subject_user}",
        body=f"""
        <div style="font-family: sans-serif; max-width: 620px; margin: 0 auto; padding: 28px 24px; color: #1a1a18;">
          <h2 style="font-size: 22px; margin-bottom: 8px;">New support message</h2>
          <p style="color: #8a8a82; margin-bottom: 22px;">A user sent a message through Insighta support.</p>
          <div style="background: #f3f1ea; border-radius: 10px; padding: 18px 20px; margin-bottom: 22px;">
            <div style="font-size: 12px; color: #8a8a82; margin-bottom: 4px;">User</div>
            <div style="font-size: 15px; font-weight: 600;">{user_label}</div>
            {f'<div style="font-size: 13px; color: #8a8a82; margin-top: 4px;">{user_name}</div>' if user_name else ''}
            <div style="font-size: 12px; color: #8a8a82; margin-top: 14px; margin-bottom: 4px;">Thread</div>
            <div style="font-size: 15px;">#{thread.id}</div>
          </div>
          <div style="border: 1px solid #e0ddd3; border-radius: 10px; padding: 16px 18px; margin-bottom: 22px; line-height: 1.55;">
            {message_body}
          </div>
          <a href="{admin_url}" style="display: inline-block; padding: 11px 18px; background: #2d6a4f; color: white; border-radius: 8px; text-decoration: none; font-weight: 600; font-size: 14px;">
            Open support inbox
          </a>
        </div>
        """,
    )


def _welcome_email_role(role: Optional[str], return_to: Optional[str], participant_app: Optional[str]) -> str:
    normalized = (role or "").strip().lower()
    if normalized in {"participant", "researcher"}:
        return normalized
    safe_return = return_to or ""
    if participant_app == "1" or safe_return.startswith(("/dashboard", "/r/", "/surveys/")):
        return "participant"
    return "researcher"


def _send_welcome_followup_email(user: User, role: str) -> tuple[bool, Optional[str]]:
    first_name = (getattr(user, "first_name", None) or getattr(user, "username", None) or "there").strip()
    escaped_first = html.escape(first_name)
    role = "researcher" if role == "researcher" else "participant"

    if role == "researcher":
        subject = "Welcome to Insighta - let's get your study staffed"
        text_body = f"""Hi {first_name},

Welcome to Insighta, and thanks for signing up.

We help researchers reach the right participants, especially for studies targeting niche or hard-to-reach populations, where general panels tend to fall short. You focus on the research; we handle finding and matching qualified participants through channels they already trust.

Here's how to get started:
1. Tell us about your study: your target population, criteria, sample size, and timeline.
2. We'll map where those participants are and bring you a first batch to review.
3. You only move forward with the ones who fit.

The fastest way to begin is a quick 15-minute call so we can understand your study and show you exactly how we'd reach your participants. You can grab a time here: https://calendly.com/yuhan-wei-2001-oh-h/30min
Or just reply to this email with a few details about your study, and we'll take it from there.

Looking forward to helping you recruit,
The Insighta Team"""
        body = f"""
        <div style="font-family: Arial, sans-serif; max-width: 620px; margin: 0 auto; padding: 28px 22px; color: #1a1a18; line-height: 1.65;">
          <h2 style="margin:0 0 18px;color:#184e77;">Welcome to Insighta</h2>
          <p>Hi {escaped_first},</p>
          <p>Welcome to Insighta, and thanks for signing up.</p>
          <p>We help researchers reach the right participants, especially for studies targeting niche or hard-to-reach populations, where general panels tend to fall short. You focus on the research; we handle finding and matching qualified participants through channels they already trust.</p>
          <p><strong>Here's how to get started:</strong></p>
          <ol>
            <li>Tell us about your study: your target population, criteria, sample size, and timeline.</li>
            <li>We'll map where those participants are and bring you a first batch to review.</li>
            <li>You only move forward with the ones who fit.</li>
          </ol>
          <p>The fastest way to begin is a quick 15-minute call so we can understand your study and show you exactly how we'd reach your participants. You can grab a time here: <a href="https://calendly.com/yuhan-wei-2001-oh-h/30min">https://calendly.com/yuhan-wei-2001-oh-h/30min</a></p>
          <p>Or just reply to this email with a few details about your study, and we'll take it from there.</p>
          <p>Looking forward to helping you recruit,<br>The Insighta Team</p>
        </div>
        """
        return send_email(user.email, subject, body, text_body=text_body)

    subject = "Welcome to Insighta! Here's what happens next"
    text_body = f"""Hi {first_name},

Thanks for joining Insighta. You're all set.

Here's what happens now: when a study matches your profile, we'll email you with the details: what it's about, what's involved, the time commitment, and any compensation. You're never automatically signed up for anything. You decide, study by study, whether it's a fit for you.

A few things we want you to know upfront:
* We only work with studies that have proper ethical (IRB) approval.
* We'll never share your information without your consent.

Here's some research that you can start off with:
BU CARD STOP study for Social Anxiety: https://insightaco.org/r/Ta29WANH7J
Paid research study on time and memory: https://insightaco.org/r/940DVmCuNO

If you ever have questions, feel free to use our chat box and connect to our team member directly.
Now, please explore around!

Warmly,
The Insighta Team"""
    body = f"""
    <div style="font-family: Arial, sans-serif; max-width: 620px; margin: 0 auto; padding: 28px 22px; color: #1a1a18; line-height: 1.65;">
      <h2 style="margin:0 0 18px;color:#168aad;">Welcome to Insighta!</h2>
      <p>Hi {escaped_first},</p>
      <p>Thanks for joining Insighta. You're all set.</p>
      <p>Here's what happens now: when a study matches your profile, we'll email you with the details: what it's about, what's involved, the time commitment, and any compensation. You're never automatically signed up for anything. You decide, study by study, whether it's a fit for you.</p>
      <p><strong>A few things we want you to know upfront:</strong></p>
      <ul>
        <li>We only work with studies that have proper ethical (IRB) approval.</li>
        <li>We'll never share your information without your consent.</li>
      </ul>
      <p>If you ever have questions, feel free to use our chat box and connect to our team member directly.</p>
      <p>Now, please explore around!</p>
      <p>Warmly,<br>The Insighta Team</p>
    </div>
    """
    return send_email(user.email, subject, body, text_body=text_body)


def _send_reward_setup_reminder_email(user: User) -> tuple[bool, Optional[str]]:
    if not user.email:
        return False, "Missing user email"
    first_name = (user.first_name or user.username or "there").strip()
    escaped_first = html.escape(first_name)
    profile_url = f"{BASE_URL.rstrip('/')}/profile"
    stripe_url = f"{BASE_URL.rstrip('/')}/connect/onboard"
    subject = "Complete your Insighta profile and payout setup"
    text_body = f"""Hi {first_name},

Quick reminder from Insighta: to make sure rewards can be matched, reviewed, and paid correctly, please complete your Insighta profile.

If you participate in paid studies, you will also need to connect Stripe before rewards can be paid out. This helps us send approved rewards securely.

What to do next:
1. Complete or update your profile: {profile_url}
2. If you are participating in paid studies, connect Stripe payouts: {stripe_url}
3. Keep your contact information current so we can reach you about study matches, bookings, and rewards.

Rewards are not issued automatically. Insighta reviews study participation and eligibility before releasing approved rewards.

Warmly,
The Insighta Team"""
    body = f"""
    <div style="font-family: Arial, sans-serif; max-width: 620px; margin: 0 auto; padding: 28px 22px; color: #1a1a18; line-height: 1.65;">
      <h2 style="margin:0 0 18px;color:#168aad;">Complete your profile and payout setup</h2>
      <p>Hi {escaped_first},</p>
      <p>Quick reminder from Insighta: to make sure rewards can be matched, reviewed, and paid correctly, please complete your Insighta profile.</p>
      <p>If you participate in paid studies, you will also need to connect Stripe before rewards can be paid out. This helps us send approved rewards securely.</p>
      <div style="background:#f6fbf8;border:1px solid rgba(24,78,119,0.12);border-radius:10px;padding:16px 18px;margin:18px 0;">
        <p style="margin:0 0 8px;font-weight:700;color:#184e77;">What to do next</p>
        <ol style="margin:0;padding-left:20px;">
          <li>Complete or update your profile.</li>
          <li>If you are participating in paid studies, connect Stripe payouts.</li>
          <li>Keep your contact information current so we can reach you about study matches, bookings, and rewards.</li>
        </ol>
      </div>
      <p>
        <a href="{profile_url}" style="display:inline-block;background:#168aad;color:#fff;text-decoration:none;border-radius:8px;padding:11px 16px;font-weight:700;margin-right:8px;">Complete profile</a>
        <a href="{stripe_url}" style="display:inline-block;background:#184e77;color:#fff;text-decoration:none;border-radius:8px;padding:11px 16px;font-weight:700;">Connect Stripe</a>
      </p>
      <p style="font-size:13px;color:#5a8297;">Rewards are not issued automatically. Insighta reviews study participation and eligibility before releasing approved rewards.</p>
      <p>Warmly,<br>The Insighta Team</p>
    </div>
    """
    return send_email(user.email, subject, body, text_body=text_body)


def _send_survey_start_followup_email(db: Session, response_id: int, dashboard_url: str) -> None:
    response = db.query(Response).filter(Response.id == response_id).first()
    if not response or getattr(response, "start_followup_sent_at", None):
        return

    participant = db.query(User).filter(User.id == response.participant_id).first()
    survey = db.query(Survey).filter(Survey.id == response.survey_id).first()
    if not participant or not participant.email or not survey:
        return

    first_name = (getattr(participant, "first_name", None) or getattr(participant, "username", None) or "there").strip()
    study_name = (survey.title or "the study").strip()
    escaped_first = html.escape(first_name)
    escaped_study = html.escape(study_name)
    escaped_dashboard = html.escape(dashboard_url, quote=True)
    subject = "How did the survey go?"
    text_body = f"""Hi {first_name},

We saw you started the {study_name} survey, just checking in: did you manage to finish it?

If you already submitted it, wonderful. Thank you for taking the time! If you stopped partway through, no worries at all. You can pick up right where you left off from your dashboard whenever you're ready: {dashboard_url}

If something wasn't clear, or you have any questions about the study, just reply to this email and a real person will help. No rush, and no pressure.

Warmly,
The Insighta Team"""
    body = f"""
    <div style="font-family: Arial, sans-serif; max-width: 620px; margin: 0 auto; padding: 28px 22px; color: #1a1a18; line-height: 1.65;">
      <h2 style="margin:0 0 18px;color:#168aad;">How did the survey go?</h2>
      <p>Hi {escaped_first},</p>
      <p>We saw you started the <strong>{escaped_study}</strong> survey, just checking in: did you manage to finish it?</p>
      <p>If you already submitted it, wonderful. Thank you for taking the time! If you stopped partway through, no worries at all. You can pick up right where you left off from your dashboard whenever you're ready: <a href="{escaped_dashboard}">dashboard</a>.</p>
      <p>If something wasn't clear, or you have any questions about the study, just reply to this email and a real person will help. No rush, and no pressure.</p>
      <p>Warmly,<br>The Insighta Team</p>
    </div>
    """
    sent, error = send_email(participant.email, subject, body, text_body=text_body)
    if sent:
        response.start_followup_sent_at = datetime.utcnow()
        db.commit()
    elif error:
        print(f"Survey start follow-up email error for response {response_id}: {error}")


def _send_booking_confirmation_emails(db: Session, survey: Survey, participant: User, booking_slot: str) -> None:
    if not booking_slot or not _uses_booking_flow(getattr(survey, "task_type", None)):
        return
    booking_label = _booking_slots_label(booking_slot)
    publisher = db.query(User).filter(User.id == survey.publisher_id).first()
    study_url = f"{BASE_URL.rstrip('/')}/publisher/study/{survey.id}"
    dashboard_url = f"{BASE_URL.rstrip('/')}/dashboard"
    location = getattr(survey, "interview_location", None) or "The researcher will share the exact location."
    participant_name = participant.first_name or participant.username or participant.email
    if participant.email:
        send_email(
            participant.email,
            f"[Insighta] Your time is booked for {survey.title}",
            f"""
            <div style="font-family:sans-serif;max-width:560px;margin:0 auto;padding:28px 22px;color:#1a1a18;line-height:1.6;">
              <h2 style="margin:0 0 10px;font-size:22px;">Your study time is booked</h2>
              <p>Hi {html.escape(participant_name)},</p>
              <p>You booked a time for <strong>{html.escape(survey.title)}</strong>.</p>
              <div style="background:#f5f3ee;border-radius:10px;padding:16px 18px;margin:18px 0;">
                <div style="font-size:12px;color:#8a8a82;text-transform:uppercase;font-weight:700;">Time</div>
                <div style="font-size:17px;font-weight:700;margin-bottom:10px;">{html.escape(booking_label)}</div>
                <div style="font-size:12px;color:#8a8a82;text-transform:uppercase;font-weight:700;">Location</div>
                <div style="font-size:15px;">{html.escape(location)}</div>
              </div>
              <p>If something changes, reply to this email and a real person will help.</p>
              <p><a href="{dashboard_url}" style="color:#3b7c4f;font-weight:700;">Back to your dashboard</a></p>
            </div>
            """
        )
    if publisher and publisher.email:
        send_email(
            publisher.email,
            f"[Insighta] New in-person booking: {survey.title}",
            f"""
            <div style="font-family:sans-serif;max-width:560px;margin:0 auto;padding:28px 22px;color:#1a1a18;line-height:1.6;">
              <h2 style="margin:0 0 10px;font-size:22px;">New participant booking</h2>
              <p><strong>{html.escape(participant.email)}</strong> booked a time for <strong>{html.escape(survey.title)}</strong>.</p>
              <div style="background:#f5f3ee;border-radius:10px;padding:16px 18px;margin:18px 0;">
                <div style="font-size:12px;color:#8a8a82;text-transform:uppercase;font-weight:700;">Time</div>
                <div style="font-size:17px;font-weight:700;margin-bottom:10px;">{html.escape(booking_label)}</div>
                <div style="font-size:12px;color:#8a8a82;text-transform:uppercase;font-weight:700;">Participant</div>
                <div style="font-size:15px;">{html.escape(participant.email)}</div>
              </div>
              <p><a href="{study_url}" style="color:#3b7c4f;font-weight:700;">Open schedule in Insighta</a></p>
            </div>
            """
        )


def _send_booking_confirmation_emails_for_response(response_id: int, booking_slot: str) -> None:
    db = SessionLocal()
    try:
        response = db.query(Response).filter(Response.id == response_id).first()
        if not response:
            return
        survey = db.query(Survey).filter(Survey.id == response.survey_id).first()
        participant = db.query(User).filter(User.id == response.participant_id).first()
        if survey and participant:
            _send_booking_confirmation_emails(db, survey, participant, booking_slot)
    finally:
        db.close()


async def _send_survey_start_followup_after_delay(response_id: int, dashboard_url: str) -> None:
    await asyncio.sleep(SURVEY_START_FOLLOWUP_DELAY_MINUTES * 60)
    db = SessionLocal()
    try:
        _send_survey_start_followup_email(db, response_id, dashboard_url)
    finally:
        db.close()


def _send_due_survey_start_followups(limit: int = 25) -> int:
    db = SessionLocal()
    sent_count = 0
    try:
        due_before = datetime.utcnow() - timedelta(minutes=SURVEY_START_FOLLOWUP_DELAY_MINUTES)
        due_responses = db.query(Response).filter(
            Response.start_followup_scheduled_at.isnot(None),
            Response.start_followup_sent_at.is_(None),
            Response.start_followup_scheduled_at <= due_before,
        ).order_by(Response.start_followup_scheduled_at.asc()).limit(limit).all()
        dashboard_url = f"{BASE_URL.rstrip('/')}/dashboard"
        for response in due_responses:
            before_sent = response.start_followup_sent_at
            _send_survey_start_followup_email(db, response.id, dashboard_url)
            db.refresh(response)
            if not before_sent and response.start_followup_sent_at:
                sent_count += 1
        return sent_count
    finally:
        db.close()


async def _survey_start_followup_worker() -> None:
    await asyncio.sleep(10)
    while True:
        try:
            sent_count = _send_due_survey_start_followups()
            if sent_count:
                print(f"Survey start follow-up worker sent {sent_count} email(s)")
        except Exception as exc:
            print(f"Survey start follow-up worker error: {exc}")
        await asyncio.sleep(SURVEY_START_FOLLOWUP_POLL_SECONDS)


@app.on_event("startup")
async def start_survey_followup_worker() -> None:
    asyncio.create_task(_survey_start_followup_worker())


def _client_ip(request: Request) -> Optional[str]:
    forwarded = request.headers.get("x-forwarded-for")
    if forwarded:
        return forwarded.split(",")[0].strip()[:120]
    if request.client:
        return request.client.host[:120]
    return None


def _safe_event_metadata(metadata: Any) -> dict:
    if not isinstance(metadata, dict):
        return {}
    allowed = {}
    for key, value in metadata.items():
        if key in {"password", "token", "verification_code", "answers", "content"}:
            continue
        safe_key = str(key)[:80]
        if isinstance(value, (str, int, float, bool)) or value is None:
            allowed[safe_key] = value if not isinstance(value, str) else value[:500]
        elif isinstance(value, list):
            allowed[safe_key] = [str(item)[:120] for item in value[:20]]
        elif isinstance(value, dict):
            allowed[safe_key] = {str(k)[:80]: str(v)[:200] for k, v in list(value.items())[:20]}
        else:
            allowed[safe_key] = str(value)[:200]
    return allowed


RESEARCHER_EVENT_NAMES = frozenset({
    "study_created", "study_published", "participant_approved",
    "participant_rejected", "study_shared",
})
VIEW_EVENT_NAMES = frozenset({
    "listing_view", "listing_viewed", "study_impression", "study_card_viewed",
    "page_view", "page_viewed",
})
START_EVENT_NAMES = frozenset({"study_start", "survey_started"})
CLICK_EVENT_NAMES = frozenset({"study_click"})
COMPLETE_EVENT_NAMES = frozenset({"study_complete", "survey_completed"})


def _event_source_from_request(request: Request, fallback: str = "") -> str:
    referer = (request.headers.get("referer") or "").lower()
    if "my-studies" in referer:
        return "My Studies"
    if "/dashboard" in referer:
        return "Dashboard"
    if "/r/" in referer:
        return "Study Page"
    return fallback or "—"


def _admin_event_role(event: UserEvent, user: Optional[User], metadata: dict) -> str:
    role = (metadata.get("user_role") or metadata.get("role") or "").strip().lower()
    if role == "participant":
        return "Participant"
    if role == "researcher":
        return "Researcher"
    if user and getattr(user, "welcome_email_role", None):
        welcome_role = (user.welcome_email_role or "").strip().lower()
        if welcome_role == "participant":
            return "Participant"
        if welcome_role == "researcher":
            return "Researcher"
    if event.event_name in RESEARCHER_EVENT_NAMES:
        return "Researcher"
    if event.user_id or event.anonymous_id:
        return "Participant"
    return "—"


def _admin_event_source(metadata: dict) -> str:
    source = (metadata.get("source") or metadata.get("surface") or "").strip()
    if not source:
        return "—"
    return source.replace("_", " ").title()


def _admin_event_details(metadata: dict) -> str:
    if not metadata:
        return "—"
    pieces = []
    if metadata.get("match_score") is not None:
        pieces.append(f"Match score: {metadata['match_score']}")
    if metadata.get("position") is not None:
        pieces.append(f"Position: {metadata['position']}")
    if metadata.get("reward_amount") is not None:
        try:
            pieces.append(f"Reward: ${float(metadata['reward_amount']):.2f}")
        except (TypeError, ValueError):
            pieces.append(f"Reward: {metadata['reward_amount']}")
    if metadata.get("button"):
        pieces.append(f"Button: {metadata['button']}")
    if not pieces:
        extras = []
        for key in ("referrer", "href"):
            if metadata.get(key):
                extras.append(str(metadata[key])[:120])
        if extras:
            return " · ".join(extras[:3])
        return "—"
    return ", ".join(pieces[:4])


def _record_user_event(
    db: Session,
    request: Request,
    event_name: str,
    user: Optional[User] = None,
    user_id: Optional[int] = None,
    anonymous_id: Optional[str] = None,
    target_type: Optional[str] = None,
    target_id: Optional[Any] = None,
    page_path: Optional[str] = None,
    metadata: Optional[dict] = None,
) -> Optional[UserEvent]:
    clean_name = (event_name or "").strip().lower()[:80]
    if not clean_name:
        return None
    resolved_user_id = user.id if user else user_id
    path = (page_path or request.url.path)[:500]
    if request.url.query and not page_path:
        path = f"{path}?{request.url.query}"[:500]
    event = UserEvent(
        user_id=resolved_user_id,
        anonymous_id=(anonymous_id or "")[:120] or None,
        event_name=clean_name,
        target_type=(target_type or "")[:80] or None,
        target_id=str(target_id)[:120] if target_id is not None else None,
        page_path=path,
        metadata_json=_safe_event_metadata(metadata or {}),
        user_agent=(request.headers.get("user-agent") or "")[:800] or None,
        client_ip=_client_ip(request),
    )
    db.add(event)
    return event


router = APIRouter()


# ---------------------------
# Commission helper
# ---------------------------

def calculate_commission(per_person_gross: float):
    if per_person_gross >= 25:
        rate = 0.25
    elif per_person_gross >= 15:
        rate = 0.20
    else:
        rate = 0.15
    reward = round(per_person_gross * (1 - rate), 2)
    return rate, reward

def timeline_multiplier(urgency_level: Optional[str]) -> float:
    urgency = (urgency_level or "flexible").strip().lower()
    if urgency == "within_1_week":
        return 1.10
    if urgency == "within_1_month":
        return 1.05
    return 1.0

def volunteer_platform_fee(target_responses: int) -> float:
    return round(((max(int(target_responses or 0), 1) + 9) // 10) * 1.0, 2)


def _extract_client_meta(request: Request, current_user: User) -> dict:
    forwarded = request.headers.get("x-forwarded-for", "")
    client_ip = forwarded.split(",")[0].strip() if forwarded else None
    if not client_ip and request.client:
        client_ip = request.client.host
    user_agent = request.headers.get("user-agent")
    fingerprint = request.headers.get("x-device-fingerprint")
    if not fingerprint:
        raw = f"{user_agent or ''}|{getattr(current_user, 'device_type', '') or ''}|{current_user.id}"
        fingerprint = hashlib.sha256(raw.encode()).hexdigest()[:32]
    return {
        "client_ip": client_ip,
        "user_agent": user_agent,
        "device_fingerprint": fingerprint,
    }


def _get_quality_check_for_publisher(db: Session, check_id: int, publisher_id: int):
    row = db.query(ResponseQualityCheck).filter(ResponseQualityCheck.id == check_id).first()
    if not row:
        raise HTTPException(404, "Quality check not found")
    survey = db.query(Survey).filter(
        Survey.id == row.survey_id,
        Survey.publisher_id == publisher_id,
    ).first()
    if not survey:
        raise HTTPException(403, "Not allowed")
    return row, survey


# ---------------------------
# Matching helpers
# ---------------------------

EDUCATION_RANK = {
    "Below High School": 0,
    "High School": 1,
    "Undergraduate": 2,
    "Graduate": 3,
    "PhD": 4,
    "Postdoc": 5,
}

def _education_rank(level: Optional[str], fallback: int) -> int:
    if not level:
        return fallback
    return EDUCATION_RANK.get(level, fallback)

def _age_range_from_birth_date(birth_year: Optional[str], birth_month: Optional[str]) -> Optional[str]:
    if not birth_year:
        return None
    try:
        year = int(birth_year)
        month = int(birth_month or "1")
    except ValueError:
        return None
    today = datetime.now(timezone.utc)
    age = today.year - year - (1 if today.month < month else 0)
    if age < 18:
        return None
    if age <= 24:
        return "18-24"
    if age <= 34:
        return "25-34"
    if age <= 44:
        return "35-44"
    return "45+"

def _value_with_other(value: Optional[str], other_value: Optional[str]) -> Optional[str]:
    value = (value or "").strip()
    other_value = (other_value or "").strip()
    if value in {"Other", "Prefer to self-describe"} and other_value:
        return f"{value}: {other_value}"
    return value or None

def _list_with_other(values: list[str], other_value: Optional[str]) -> list[str]:
    cleaned = [v.strip() for v in values if v and v.strip()]
    other_value = (other_value or "").strip()
    if "Other" in cleaned and other_value:
        cleaned.append(f"Other: {other_value}")
    return cleaned

def _is_empty(val: Optional[str]) -> bool:
    return val is None or val.strip() == "" or val.strip().lower() == "all"

def _field_matches(target: Optional[str], user_val: Optional[str]) -> bool:
    if _is_empty(target):
        return True
    if not user_val:
        return False
    return target.strip().lower() == user_val.strip().lower()

FIELD_RELEVANCE_ALIASES = {
    "computer science & engineering": ["computer science", "engineering", "software", "developer", "coding", "programming", "technology", "tech"],
    "data science / ai": ["data science", "data", "ai", "artificial intelligence", "machine learning", "ml", "analytics", "algorithm"],
    "business & economics": ["business", "economics", "finance", "entrepreneurship", "startup", "management", "consumer", "market"],
    "marketing / design": ["marketing", "design", "brand", "ux", "user experience", "product", "advertising", "creative"],
    "medicine & healthcare": ["medicine", "healthcare", "health", "clinical", "patient", "care", "medical", "public health"],
    "psychology / neuroscience": ["psychology", "neuroscience", "mental health", "behavior", "cognition", "brain", "wellness"],
    "natural sciences": ["biology", "chemistry", "physics", "environment", "science", "lab", "ecology"],
    "social sciences": ["social science", "sociology", "anthropology", "political science", "community", "society", "policy"],
    "education": ["education", "teaching", "learning", "school", "student", "classroom", "curriculum"],
    "law / public policy": ["law", "legal", "policy", "government", "civic", "regulation", "justice"],
    "arts & humanities": ["arts", "humanities", "history", "literature", "philosophy", "culture", "music", "theater"],
    "communications / media": ["communications", "media", "journalism", "social media", "content", "creator", "news"],
    "architecture / design": ["architecture", "urban", "built environment", "space", "housing", "design"],
}

def _text_relevance_score(needles: list[str], haystack: str) -> float:
    if not needles or not haystack:
        return 0.0
    haystack = haystack.lower()
    hits = sum(1 for needle in needles if needle and needle.lower() in haystack)
    if hits <= 0:
        return 0.0
    return min(0.75, 0.25 + hits * 0.15)

def _field_relevance_score(survey: Survey, user: User) -> float:
    user_field = (getattr(user, "field", None) or "").strip()
    if not user_field:
        return 0.0
    survey_target = (getattr(survey, "target_field", None) or "").strip()
    user_field_clean = user_field.lower()
    survey_target_clean = survey_target.lower()
    if survey_target_clean and survey_target_clean != "all":
        if survey_target_clean == user_field_clean:
            return 1.0
        aliases = FIELD_RELEVANCE_ALIASES.get(user_field_clean, [])
        if survey_target_clean in aliases or any(term in survey_target_clean for term in aliases):
            return 0.85
        return 0.0

    aliases = FIELD_RELEVANCE_ALIASES.get(user_field_clean, [])
    user_terms = [user_field_clean] + aliases
    survey_text = " ".join([
        getattr(survey, "title", None) or "",
        getattr(survey, "description", None) or "",
        getattr(survey, "category", None) or "",
        getattr(survey, "target_niche_requirements", None) or "",
        getattr(survey, "participant_benefits", None) or "",
    ])
    score = _text_relevance_score(user_terms, survey_text)

    interest_text = (getattr(user, "profile_description", None) or "").strip()
    if interest_text:
        interest_terms = [
            token for token in re.findall(r"[A-Za-z][A-Za-z+/#-]{3,}", interest_text.lower())
            if token not in {"with", "that", "this", "from", "about", "study", "research", "interested"}
        ][:12]
        score = max(score, min(0.6, _text_relevance_score(interest_terms, survey_text)))
    return round(score, 4)

def _recommendation_sort_score(survey: Survey, user: User, recommendation: dict) -> float:
    completion_probability = float((recommendation or {}).get("completion_probability") or 0.0)
    field_fit = _field_relevance_score(survey, user)
    return round((completion_probability * 0.75) + (field_fit * 0.25), 4)

def _age_matches(target: Optional[str], user: User) -> bool:
    if _is_empty(target):
        return True
    target_clean = target.strip().lower()
    try:
        birth_year = int(getattr(user, "birth_year", None) or 0)
        birth_month = int(getattr(user, "birth_month", None) or 1)
        today = date.today()
        user_age = today.year - birth_year - ((today.month, today.day) < (birth_month, 1))
    except (TypeError, ValueError):
        user_age = None
    numbers = [int(n) for n in re.findall(r"\d+", target_clean)]
    if user_age is not None and numbers:
        if len(numbers) >= 2:
            return numbers[0] <= user_age <= numbers[1]
        if "+" in target_clean or "older" in target_clean:
            return user_age >= numbers[0]
        if "under" in target_clean or "below" in target_clean:
            return user_age <= numbers[0]
    return _field_matches(target, getattr(user, "age_range", None))

def _location_matches(target: Optional[str], user: User) -> bool:
    if _is_empty(target):
        return True
    target_clean = target.strip().lower()
    user_values = [
        getattr(user, "state", None),
        getattr(user, "current_province", None),
    ]
    return any((value or "").strip().lower() == target_clean for value in user_values)

def _language_matches(target: Optional[str], user_languages: Optional[str]) -> bool:
    if _is_empty(target):
        return True
    if not user_languages:
        return False
    user_list = [lang.strip().lower() for lang in user_languages.split(",") if lang.strip()]
    return target.strip().lower() in user_list

def _tags_match(target_tags: Optional[str], user_tags: Optional[str]) -> bool:
    if _is_empty(target_tags):
        return True
    if not user_tags:
        return False
    target_set = {t.strip().lower() for t in target_tags.split(",") if t.strip()}
    user_set = {t.strip().lower() for t in user_tags.split(",") if t.strip()}
    return bool(target_set & user_set)

def _participation_format_matches(target: Optional[str], user_val: Optional[str]) -> bool:
    target_clean = (target or "").strip().lower()
    user_clean = (user_val or "").strip().lower()
    if _is_empty(target) or target_clean in {"both", "any format", "hybrid"}:
        return True
    if not user_clean:
        return True
    if user_clean in {"both", "any format", "hybrid"}:
        return True
    return target_clean == user_clean

def _device_matches(target: Optional[str], user_val: Optional[str]) -> bool:
    if _is_empty(target) or (target and target.strip().lower() == "any"):
        return True
    if not user_val:
        return False
    if user_val.strip().lower() == "any":
        return True
    return target.strip().lower() == user_val.strip().lower()

def _parse_booking_slots(value: Optional[str]) -> list[str]:
    raw = (value or "").strip()
    if not raw:
        return []
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, list):
            return [str(item).strip() for item in parsed if str(item).strip()]
        if isinstance(parsed, str):
            return [parsed.strip()] if parsed.strip() else []
    except Exception:
        pass
    return [raw]

def _serialize_booking_slots(slots: list[str]) -> Optional[str]:
    cleaned = []
    seen = set()
    for slot in slots:
        label = str(slot or "").strip()
        if label and label not in seen:
            cleaned.append(label)
            seen.add(label)
    if not cleaned:
        return None
    return cleaned[0] if len(cleaned) == 1 else json.dumps(cleaned)

def _booking_slots_label(value: Optional[str]) -> str:
    return "; ".join(_parse_booking_slots(value))


# ---------------------------
# Other helpers
# ---------------------------

def _parse_optional_int(v) -> Optional[int]:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return int(v)
    except (ValueError, TypeError):
        return None

def _parse_optional_float(v) -> Optional[float]:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None

def _normalize_task_type(value: Optional[str]) -> str:
    normalized = (value or "").strip().lower()
    if normalized in {"interview", "online_interview", "online-interview", "remote_interview"}:
        return "interview"
    if normalized in {"in_person", "in-person", "in_person_study"}:
        return "in_person"
    return "survey"

def _task_type_label(value: Optional[str]) -> str:
    raw = (value or "").strip().lower()
    if raw in {"online_interview", "online-interview", "remote_interview"}:
        return "Online interview"
    if raw in {"in_person", "in-person", "in_person_study"}:
        return "In-person study"
    task_type = _normalize_task_type(value)
    if task_type == "interview":
        return "Interview"
    if task_type == "in_person":
        return "In-person study"
    return "Survey"

def _uses_booking_flow(value: Optional[str]) -> bool:
    return _normalize_task_type(value) in {"interview", "in_person"}

RESEARCH_PARTICIPATION_DEMO_TITLE = "Understand the Motivations and Barriers to Participating in Online Surveys and Research Studies"
RESEARCH_PARTICIPATION_DEMO_CALENDLY_URL = "https://calendly.com/vfsa-bu/understanding-research-participation"

def _compact_text(value: Optional[str]) -> str:
    return "".join((value or "").split()).lower()

def _is_research_participation_demo_survey(survey: Survey) -> bool:
    compact_title = _compact_text(getattr(survey, "title", None))
    return (
        compact_title == _compact_text(RESEARCH_PARTICIPATION_DEMO_TITLE)
        or (
            "motivation" in compact_title
            and "barrier" in compact_title
            and "onlinesurveys" in compact_title
            and "researchstudies" in compact_title
        )
    )

def _survey_external_start_url(survey: Survey) -> str:
    if _is_research_participation_demo_survey(survey):
        return RESEARCH_PARTICIPATION_DEMO_CALENDLY_URL
    form_url = (getattr(survey, "form_url", None) or "").strip()
    return form_url if form_url and form_url != "__builtin__" else ""

def _survey_has_external_start_link(survey: Survey) -> bool:
    return bool(_survey_external_start_url(survey))

def _is_online_interview_survey(survey: Survey) -> bool:
    task_type = _normalize_task_type(getattr(survey, "task_type", None))
    if task_type != "interview":
        return False
    if _is_research_participation_demo_survey(survey):
        return True
    participation_format = (getattr(survey, "target_participation_format", None) or "").strip().lower()
    form_url = _survey_external_start_url(survey).lower()
    return (
        participation_format in {"video interview", "online interview", "remote live session"}
        or "calendly.com" in form_url
        or "zoom.us" in form_url
        or "meet.google.com" in form_url
    )

def _participant_study_type_label(survey: Survey) -> Optional[str]:
    task_type = _normalize_task_type(getattr(survey, "task_type", None))
    if task_type == "interview":
        if _is_online_interview_survey(survey):
            return "Online interview"
        participation_format = (getattr(survey, "target_participation_format", None) or "").strip().lower()
        if participation_format in {"in-person study", "in person", "in-person"}:
            return "In-person study"
        return "Interview"
    if task_type == "in_person":
        return "In-person study"
    return None

def _participant_study_action_label(survey: Survey) -> str:
    task_type = _normalize_task_type(getattr(survey, "task_type", None))
    if task_type == "interview":
        return "Schedule interview" if _is_online_interview_survey(survey) else "Book time"
    return "Take study"

def _clean_target(val: Optional[str]) -> str:
    return '' if not val or val.strip().lower() == 'all' else val

def _join_form_list_with_other(values: list[str], other_value: Optional[str]) -> Optional[str]:
    cleaned = []
    seen = set()
    for value in values:
        item = str(value or "").strip()
        if item and item.lower() != "all" and item not in seen:
            cleaned.append(item)
            seen.add(item)
    other = (other_value or "").strip()
    if other and other not in seen:
        cleaned.append(other)
    return "; ".join(cleaned) if cleaned else None


def _normalize_excel_header(value: Optional[str]) -> str:
    return (value or "").strip().lower()


def _value_as_float(value) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _json_safe_value(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, dict):
        return {k: _json_safe_value(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_json_safe_value(v) for v in value]
    return value


def _latest_excel_quality_rows(rows: list) -> list:
    excel_rows = [row for row in rows if row.source_type == "excel"]
    other_rows = [row for row in rows if row.source_type != "excel"]
    if not excel_rows:
        return rows

    def _upload_key(row):
        ref = row.source_ref or ""
        return ref.rsplit(":row_", 1)[0] if ":row_" in ref else ref

    latest_row = max(excel_rows, key=lambda row: row.created_at or datetime.min.replace(tzinfo=timezone.utc))
    latest_key = _upload_key(latest_row)
    latest_excel = [row for row in excel_rows if _upload_key(row) == latest_key]
    latest_excel.sort(key=lambda row: row.created_at or datetime.min.replace(tzinfo=timezone.utc), reverse=True)
    return latest_excel + other_rows


def _quality_row_label(index: int = 0) -> str:
    return str(index + 1)


# ---------------------------
# Auth helpers
# ---------------------------

def _normalize_email(value: Optional[str]) -> str:
    return (value or "").strip().lower()

MOBILE_USER_AGENT_TOKENS = ("mobile", "android", "iphone", "ipad")

def _is_mobile_request(request: Request) -> bool:
    user_agent = request.headers.get("user-agent", "").lower()
    return any(token in user_agent for token in MOBILE_USER_AGENT_TOKENS)

def _participant_dashboard_url(request: Request) -> str:
    return "/dashboard/mobile" if _is_mobile_request(request) else "/dashboard"

def _should_use_participant_app(request: Request, participant_app: Optional[str] = None) -> bool:
    return _is_mobile_request(request) or participant_app == "1"

def _auth_uses_participant_app(request: Request, role: Optional[str] = None, participant_app: Optional[str] = None) -> bool:
    normalized_role = (role or "").strip().lower()
    if normalized_role == "participant":
        return True
    if normalized_role == "researcher":
        return False
    return _should_use_participant_app(request, participant_app)

def _post_auth_url(request: Request, participant_app: Optional[str] = None, welcome: bool = False) -> str:
    if _should_use_participant_app(request, participant_app):
        base_url = _participant_dashboard_url(request)
    else:
        base_url = "/publisher"
    return f"{base_url}?welcome=1" if welcome else base_url

def _cookie_policy(request: Optional[Request]) -> dict:
    is_https = False
    if request is not None:
        forwarded_proto = (request.headers.get("x-forwarded-proto") or "").split(",")[0].strip().lower()
        is_https = (request.url.scheme == "https") or (forwarded_proto == "https")
    if is_https:
        return {"httponly": True, "samesite": "none", "secure": True}
    return {"httponly": True, "samesite": "lax", "secure": False}

def _post_auth_url_with_next(
    request: Request,
    participant_app: Optional[str] = None,
    next_url: Optional[str] = None,
    role: Optional[str] = None,
    welcome: bool = False
) -> str:
    if is_safe_internal_next(next_url):
        return next_url
    normalized_role = (role or "").strip().lower()
    if normalized_role == "participant":
        return _participant_dashboard_url(request)
    if normalized_role == "researcher":
        return "/publisher"
    return _post_auth_url(request, participant_app, welcome=welcome)


def _needs_identity_onboarding(user: User) -> bool:
    return not all([
        (getattr(user, "first_name", None) or "").strip(),
        (getattr(user, "last_name", None) or "").strip(),
        (getattr(user, "username", None) or "").strip(),
        (getattr(user, "birth_year", None) or "").strip(),
        (getattr(user, "birth_month", None) or "").strip(),
        (getattr(user, "education_level", None) or "").strip(),
        (getattr(user, "field", None) or "").strip(),
        (getattr(user, "status", None) or "").strip(),
        (getattr(user, "current_province", None) or getattr(user, "state", None) or "").strip(),
        (getattr(user, "language", None) or "").strip(),
        (getattr(user, "participation_format", None) or "").strip(),
    ])


def _identity_onboarding_url(return_to: Optional[str] = None, role: Optional[str] = None) -> str:
    safe_return = return_to if is_safe_internal_next(return_to) and return_to != "/complete-profile" else ""
    normalized_role = role if role in {"participant", "researcher"} else ""
    query = {}
    if safe_return:
        query["next"] = safe_return
    if normalized_role:
        query["role"] = normalized_role
    qs = urlencode(query) if query else ""
    return f"/complete-profile?{qs}" if qs else "/complete-profile"


def _post_auth_or_onboarding_url(
    user: User,
    final_url: str,
    role: Optional[str] = None,
    participant_app: Optional[str] = None,
) -> str:
    if _needs_identity_onboarding(user):
        return _identity_onboarding_url(
            final_url,
            _welcome_email_role(role, final_url, participant_app),
        )
    return final_url


AUTH_RETURN_COOKIE = "auth_return_to"
AUTH_COOKIE_MAX_AGE = 60 * 60 * 24 * 30

def _set_user_cookie(response, request: Request, user: User) -> None:
    response.set_cookie(
        "user_id",
        str(user.id),
        max_age=AUTH_COOKIE_MAX_AGE,
        **_cookie_policy(request),
    )


def _safe_auth_return(*candidates: Optional[str]) -> str:
    for candidate in candidates:
        if is_safe_internal_next(candidate):
            return candidate
    return ""


def _remember_auth_return(response: Response, request: Request, next_url: Optional[str]) -> None:
    safe_next = _safe_auth_return(next_url)
    if safe_next:
        response.set_cookie(
            AUTH_RETURN_COOKIE,
            safe_next,
            max_age=30 * 60,
            **_cookie_policy(request),
        )


def _clear_auth_return(response: Response, request: Request) -> None:
    response.delete_cookie(AUTH_RETURN_COOKIE, **_cookie_policy(request))

def _absolute_url(request: Request, path: str) -> str:
    base = str(request.base_url).rstrip("/")
    if not path.startswith("/"):
        path = f"/{path}"
    return f"{base}{path}"

def _new_share_slug() -> str:
    return secrets.token_urlsafe(8).replace("-", "").replace("_", "")[:10]

def _ensure_survey_share_slug(db: Session, survey: Survey, commit: bool = False) -> str:
    if getattr(survey, "share_slug", None):
        return survey.share_slug
    for _ in range(12):
        slug = _new_share_slug()
        if not db.query(Survey).filter(Survey.share_slug == slug).first():
            survey.share_slug = slug
            if commit:
                db.commit()
                db.refresh(survey)
            return slug
    raise HTTPException(500, "Could not create share link")

def _survey_share_path(db: Session, survey: Survey, commit: bool = False) -> str:
    return f"/r/{_ensure_survey_share_slug(db, survey, commit=commit)}"

def _survey_share_url(request: Request, db: Session, survey: Survey, commit: bool = False) -> str:
    return _absolute_url(request, _survey_share_path(db, survey, commit=commit))

def _mark_participant_app(response: RedirectResponse, request: Optional[Request] = None):
    response.set_cookie(
        "participant_app",
        "1",
        max_age=60 * 60 * 24 * 30,
        **_cookie_policy(request),
    )

def _clear_participant_app(response, request: Optional[Request] = None):
    policy = _cookie_policy(request)
    response.delete_cookie("participant_app", samesite=policy["samesite"], secure=policy["secure"])

def _mask_email(email: str) -> str:
    if "@" not in email:
        return email
    name, domain = email.split("@", 1)
    if len(name) <= 2:
        masked_name = name[0] + "*" * max(len(name) - 1, 0)
    else:
        masked_name = name[:2] + "*" * max(len(name) - 2, 0)
    return f"{masked_name}@{domain}"

def _generate_verification_code() -> str:
    return f"{random.randint(0, 999999):06d}"

PASSWORD_POLICY_MESSAGE = (
    "Password must be at least 8 characters and include an uppercase letter, "
    "a lowercase letter, a number, and a special character."
)

def _validate_registration_password(password: str) -> Optional[str]:
    value = password or ""
    if len(value) < 8: return PASSWORD_POLICY_MESSAGE
    if not re.search(r"[A-Z]", value): return PASSWORD_POLICY_MESSAGE
    if not re.search(r"[a-z]", value): return PASSWORD_POLICY_MESSAGE
    if not re.search(r"\d", value): return PASSWORD_POLICY_MESSAGE
    if not re.search(r"[^A-Za-z0-9]", value): return PASSWORD_POLICY_MESSAGE
    return None

def _mark_previous_codes_used(db: Session, email: str, purpose: str):
    pending_codes = db.query(EmailVerificationCode).filter(
        EmailVerificationCode.email == email,
        EmailVerificationCode.purpose == purpose,
        EmailVerificationCode.used_at.is_(None)
    ).all()
    now = datetime.utcnow()
    for item in pending_codes:
        item.used_at = now

def _latest_active_verification_code(db: Session, email: str, purpose: str) -> Optional[EmailVerificationCode]:
    now = datetime.utcnow()
    return db.query(EmailVerificationCode).filter(
        EmailVerificationCode.email == email,
        EmailVerificationCode.purpose == purpose,
        EmailVerificationCode.used_at.is_(None),
        EmailVerificationCode.expires_at > now,
    ).order_by(EmailVerificationCode.created_at.desc()).first()

def _issue_verification_code(db: Session, email: str, purpose: str) -> tuple[str, bool]:
    existing = _latest_active_verification_code(db, email, purpose)
    if existing:
        return existing.code, False
    code = _generate_verification_code()
    db.add(EmailVerificationCode(
        email=email,
        purpose=purpose,
        code=code,
        expires_at=datetime.utcnow() + timedelta(minutes=VERIFICATION_CODE_EXPIRE_MINUTES)
    ))
    db.commit()
    return code, True

def _consume_verification_code(db: Session, email: str, purpose: str, code: str) -> bool:
    normalized_email = _normalize_email(email)
    normalized_code = (code or "").strip()
    record = db.query(EmailVerificationCode).filter(
        EmailVerificationCode.email == normalized_email,
        EmailVerificationCode.purpose == purpose,
        EmailVerificationCode.code == normalized_code,
        EmailVerificationCode.used_at.is_(None)
    ).order_by(EmailVerificationCode.created_at.desc()).first()
    if not record:
        return False
    if record.expires_at < datetime.utcnow():
        return False
    _mark_previous_codes_used(db, normalized_email, purpose)
    db.commit()
    return True


PENDING_REFERRAL_COOKIE = "pending_referral"
_REFERRAL_CODE_ALPHABET = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"


def _normalize_referral_code(value: Optional[str]) -> str:
    return re.sub(r"[^A-Za-z0-9]", "", (value or "").strip()).upper()


def _generate_referral_code(db: Session) -> str:
    for _ in range(40):
        code = "".join(secrets.choice(_REFERRAL_CODE_ALPHABET) for _ in range(8))
        exists = db.query(User.id).filter(User.referral_code == code).first()
        if not exists:
            return code
    return secrets.token_hex(5).upper()


def _ensure_user_referral_code(db: Session, user: User, *, commit: bool = False) -> str:
    if user.referral_code:
        return user.referral_code
    user.referral_code = _generate_referral_code(db)
    db.add(user)
    if commit:
        db.commit()
        db.refresh(user)
    return user.referral_code


def _find_referrer_by_code(db: Session, code: Optional[str]) -> Optional[User]:
    normalized = _normalize_referral_code(code)
    if not normalized:
        return None
    return db.query(User).filter(User.referral_code == normalized).first()


def _apply_referral_code(db: Session, user: User, code: Optional[str]) -> Optional[str]:
    """Attach invited_by if code is valid. Returns error message or None."""
    normalized = _normalize_referral_code(code)
    if not normalized:
        return None
    referrer = _find_referrer_by_code(db, normalized)
    if not referrer:
        return "Referral code not found."
    if user.id and referrer.id == user.id:
        return "You cannot use your own referral code."
    if not user.invited_by_user_id:
        user.invited_by_user_id = referrer.id
    return None


def _referral_invite_url(code: str, role: Optional[str] = None) -> str:
    params = {"ref": code}
    if role in {"participant", "researcher"}:
        params["role"] = role
    return f"{BASE_URL.rstrip('/')}/register?{urlencode(params)}"


def _remember_pending_referral(response, request: Request, code: Optional[str]) -> None:
    normalized = _normalize_referral_code(code)
    if not normalized:
        return
    response.set_cookie(
        PENDING_REFERRAL_COOKIE,
        normalized,
        max_age=60 * 60 * 24 * 14,
        **_cookie_policy(request),
    )


def _clear_pending_referral(response, request: Optional[Request] = None) -> None:
    response.delete_cookie(PENDING_REFERRAL_COOKIE, **_cookie_policy(request))


def _send_verification_email(email: str, purpose: str, code: str) -> tuple[bool, Optional[str]]:
    if purpose == "register":
        subject = "Insighta registration verification code"
        title = "Verify your email"
        body_text = "Use the following verification code to finish creating your Insighta account."
    else:
        subject = "Insighta password reset verification code"
        title = "Reset your password"
        body_text = "Use the following verification code to reset your Insighta password."

    body = f'''
    <div style="font-family: DM Sans, Arial, sans-serif; max-width: 560px; margin: 0 auto; padding: 32px 24px; color: #1a1a18;">
      <div style="margin-bottom: 20px;">
        <div style="font-family: Georgia, serif; font-size: 28px; color: #2d6a4f; margin-bottom: 8px;">Insighta</div>
        <div style="font-size: 13px; color: #8a8a82;">Secure verification</div>
      </div>
      <div style="background: #f3f1ea; border: 1px solid #e0ddd3; border-radius: 16px; padding: 24px;">
        <h2 style="font-family: Georgia, serif; font-size: 26px; margin: 0 0 10px; color: #1a1a18;">{title}</h2>
        <p style="font-size: 15px; line-height: 1.7; color: #4a4a44; margin: 0 0 20px;">{body_text}</p>
        <div style="font-size: 13px; color: #8a8a82; margin-bottom: 8px;">Verification code</div>
        <div style="font-size: 36px; font-weight: 700; letter-spacing: 8px; color: #2d6a4f; background: white; border: 1px solid #d6d0c1; border-radius: 12px; padding: 18px 20px; text-align: center;">
          {code}
        </div>
        <p style="font-size: 13px; line-height: 1.7; color: #8a8a82; margin: 20px 0 0;">
          This code expires in {VERIFICATION_CODE_EXPIRE_MINUTES} minutes. If you did not request this, you can safely ignore it.
        </p>
      </div>
      <p style="font-size: 12px; color: #8a8a82; margin-top: 20px;">Sent securely by Insighta.</p>
    </div>
    '''
    text_body = (
        f"Insighta - {title}\n\n{body_text}\n\n"
        f"Verification code: {code}\n\n"
        f"This code expires in {VERIFICATION_CODE_EXPIRE_MINUTES} minutes. "
        "If you did not request this, you can safely ignore it."
    )
    return send_email(email, subject, body, text_body=text_body)


# ---------------------------
# Send verification code API
# ---------------------------

@app.get("/auth/check-email")
async def check_email_exists(email: str, db: Session = Depends(get_db)):
    normalized = _normalize_email(email)
    if not normalized:
        return JSONResponse({"exists": False})
    exists = db.query(User).filter(User.email == normalized).first() is not None
    return JSONResponse({"exists": exists})


@app.post("/auth/send-code")
async def send_auth_code(
    email: str = Form(...),
    purpose: str = Form(...),
    db: Session = Depends(get_db)
):
    normalized_email = _normalize_email(email)
    normalized_purpose = (purpose or "").strip().lower()

    if not normalized_email:
        return JSONResponse({"ok": False, "message": "Please enter your email address first."}, status_code=400)
    if normalized_purpose not in {"register", "reset_password"}:
        return JSONResponse({"ok": False, "message": "Unsupported verification purpose."}, status_code=400)

    existing_user = db.query(User).filter(User.email == normalized_email).first()

    if normalized_purpose == "register" and existing_user:
        return JSONResponse({"ok": False, "message": "This email is already registered. Please sign in instead."}, status_code=400)
    if normalized_purpose == "reset_password" and not existing_user:
        return JSONResponse({"ok": False, "message": "No account found with this email."}, status_code=400)

    now = datetime.utcnow()
    active_code = _latest_active_verification_code(db, normalized_email, normalized_purpose)
    if active_code and active_code.created_at and active_code.created_at > now - timedelta(seconds=VERIFICATION_CODE_RESEND_COOLDOWN_SECONDS):
        wait_seconds = max(1, VERIFICATION_CODE_RESEND_COOLDOWN_SECONDS - int((now - active_code.created_at).total_seconds()))
        return JSONResponse({
            "ok": False,
            "message": f"Please wait {wait_seconds} seconds before requesting another code."
        }, status_code=429)

    recent_code_count = db.query(EmailVerificationCode).filter(
        EmailVerificationCode.email == normalized_email,
        EmailVerificationCode.purpose == normalized_purpose,
        EmailVerificationCode.created_at >= now - timedelta(hours=1),
    ).count()
    if recent_code_count >= VERIFICATION_CODE_MAX_PER_HOUR and not active_code:
        return JSONResponse({
            "ok": False,
            "message": "Too many verification requests. Please try again later."
        }, status_code=429)

    code, created_new = _issue_verification_code(db, normalized_email, normalized_purpose)
    sent, send_error = _send_verification_email(normalized_email, normalized_purpose, code)
    if not sent:
        if created_new:
            db.query(EmailVerificationCode).filter(
                EmailVerificationCode.email == normalized_email,
                EmailVerificationCode.purpose == normalized_purpose,
                EmailVerificationCode.code == code,
                EmailVerificationCode.used_at.is_(None),
            ).delete(synchronize_session=False)
            db.commit()
        print(f"Verification email failed for {_mask_email(normalized_email)}: {send_error}")
        return JSONResponse({
            "ok": False,
            "message": "We could not send the verification email. Please try again in a moment."
        }, status_code=502)

    return JSONResponse({
        "ok": True,
        "message": f"Verification code sent to {_mask_email(normalized_email)}."
    })


# ---------------------------
# Google OAuth
# ---------------------------

@app.get("/auth/google")
def google_login(request: Request):
    if not GOOGLE_CLIENT_ID:
        raise HTTPException(503, "Google login is not configured.")
    state = secrets.token_urlsafe(16)
    params = urlencode({
        "client_id": GOOGLE_CLIENT_ID,
        "redirect_uri": f"{BASE_URL}/auth/google/callback",
        "response_type": "code",
        "scope": "openid email profile",
        "state": state,
        "access_type": "offline",
        "prompt": "select_account",
    })
    response = RedirectResponse(f"{GOOGLE_AUTH_URL}?{params}", status_code=302)
    response.set_cookie("oauth_state", state, max_age=300, **_cookie_policy(request))
    return response


@app.get("/auth/google/callback")
async def google_callback(
    request: Request,
    code: Optional[str] = None,
    state: Optional[str] = None,
    error: Optional[str] = None,
    db: Session = Depends(get_db),
):
    if error:
        return RedirectResponse("/login?oauth_error=Google+login+was+cancelled.", status_code=302)

    stored_state = request.cookies.get("oauth_state")
    if not state or state != stored_state:
        return RedirectResponse("/login?oauth_error=Invalid+state.+Please+try+again.", status_code=302)

    async with httpx.AsyncClient() as client:
        token_resp = await client.post(GOOGLE_TOKEN_URL, data={
            "code": code,
            "client_id": GOOGLE_CLIENT_ID,
            "client_secret": GOOGLE_CLIENT_SECRET,
            "redirect_uri": f"{BASE_URL}/auth/google/callback",
            "grant_type": "authorization_code",
        })
        token_data = token_resp.json()
        access_token = token_data.get("access_token")
        if not access_token:
            return RedirectResponse("/login?oauth_error=Google+token+exchange+failed.", status_code=302)

        userinfo_resp = await client.get(
            GOOGLE_USERINFO_URL,
            headers={"Authorization": f"Bearer {access_token}"},
        )
        userinfo = userinfo_resp.json()

    email = _normalize_email(userinfo.get("email", ""))
    google_id = str(userinfo.get("id", ""))
    name = userinfo.get("name") or userinfo.get("given_name")

    if not email:
        return RedirectResponse("/login?oauth_error=Could+not+retrieve+email+from+Google.", status_code=302)

    user = db.query(User).filter(User.email == email).first()
    is_new = False
    if not user:
        user = User(
            email=email,
            password=pwd_context.hash(secrets.token_urlsafe(32)),
            username=name,
            oauth_provider="google",
            oauth_id=google_id,
            referral_code=_generate_referral_code(db),
        )
        pending_ref = request.cookies.get(PENDING_REFERRAL_COOKIE)
        _apply_referral_code(db, user, pending_ref)
        db.add(user)
        db.commit()
        db.refresh(user)
        is_new = True
    else:
        if not user.oauth_provider:
            user.oauth_provider = "google"
            user.oauth_id = google_id
            db.commit()
        _ensure_user_referral_code(db, user, commit=True)

    return_to = _safe_auth_return(request.cookies.get(AUTH_RETURN_COOKIE))
    final_url = _post_auth_url_with_next(
        request,
        request.cookies.get("participant_app"),
        return_to,
        welcome=is_new,
    )
    redirect_url = _post_auth_or_onboarding_url(
        user,
        final_url,
        participant_app=request.cookies.get("participant_app"),
    )
    resp = RedirectResponse(redirect_url, status_code=303)
    policy = _cookie_policy(request)
    _set_user_cookie(resp, request, user)
    resp.delete_cookie("oauth_state", samesite=policy["samesite"], secure=policy["secure"])
    _clear_auth_return(resp, request)
    if is_new:
        _clear_pending_referral(resp, request)
    return resp


# ---------------------------
# LinkedIn OAuth
# ---------------------------

@app.get("/auth/linkedin")
def linkedin_login(request: Request):
    if not LINKEDIN_CLIENT_ID:
        raise HTTPException(503, "LinkedIn login is not configured.")
    state = secrets.token_urlsafe(16)
    params = urlencode({
        "response_type": "code",
        "client_id": LINKEDIN_CLIENT_ID,
        "redirect_uri": f"{BASE_URL}/auth/linkedin/callback",
        "state": state,
        "scope": "openid profile email",
    })
    response = RedirectResponse(f"{LINKEDIN_AUTH_URL}?{params}", status_code=302)
    response.set_cookie("oauth_state", state, max_age=300, **_cookie_policy(request))
    return response


@app.get("/auth/linkedin/callback")
async def linkedin_callback(
    request: Request,
    code: Optional[str] = None,
    state: Optional[str] = None,
    error: Optional[str] = None,
    db: Session = Depends(get_db),
):
    if error:
        return RedirectResponse("/login?oauth_error=LinkedIn+login+was+cancelled.", status_code=302)

    stored_state = request.cookies.get("oauth_state")
    if not state or state != stored_state:
        return RedirectResponse("/login?oauth_error=Invalid+state.+Please+try+again.", status_code=302)

    async with httpx.AsyncClient() as client:
        token_resp = await client.post(
            LINKEDIN_TOKEN_URL,
            data={
                "grant_type": "authorization_code",
                "code": code,
                "redirect_uri": f"{BASE_URL}/auth/linkedin/callback",
                "client_id": LINKEDIN_CLIENT_ID,
                "client_secret": LINKEDIN_CLIENT_SECRET,
            },
            headers={"Content-Type": "application/x-www-form-urlencoded"},
        )
        token_data = token_resp.json()
        access_token = token_data.get("access_token")
        if not access_token:
            return RedirectResponse("/login?oauth_error=LinkedIn+token+exchange+failed.", status_code=302)

        userinfo_resp = await client.get(
            LINKEDIN_USERINFO_URL,
            headers={"Authorization": f"Bearer {access_token}"},
        )
        userinfo = userinfo_resp.json()

    email = _normalize_email(userinfo.get("email", ""))
    linkedin_id = str(userinfo.get("sub", ""))
    name = userinfo.get("name") or userinfo.get("given_name")

    if not email:
        return RedirectResponse("/login?oauth_error=Could+not+retrieve+email+from+LinkedIn.", status_code=302)

    user = db.query(User).filter(User.email == email).first()
    is_new = False
    if not user:
        user = User(
            email=email,
            password=pwd_context.hash(secrets.token_urlsafe(32)),
            username=name,
            oauth_provider="linkedin",
            oauth_id=linkedin_id,
            referral_code=_generate_referral_code(db),
        )
        pending_ref = request.cookies.get(PENDING_REFERRAL_COOKIE)
        _apply_referral_code(db, user, pending_ref)
        db.add(user)
        db.commit()
        db.refresh(user)
        is_new = True
    else:
        if not user.oauth_provider:
            user.oauth_provider = "linkedin"
            user.oauth_id = linkedin_id
            db.commit()
        _ensure_user_referral_code(db, user, commit=True)

    return_to = _safe_auth_return(request.cookies.get(AUTH_RETURN_COOKIE))
    final_url = _post_auth_url_with_next(
        request,
        request.cookies.get("participant_app"),
        return_to,
        welcome=is_new,
    )
    redirect_url = _post_auth_or_onboarding_url(
        user,
        final_url,
        participant_app=request.cookies.get("participant_app"),
    )
    resp = RedirectResponse(redirect_url, status_code=303)
    policy = _cookie_policy(request)
    _set_user_cookie(resp, request, user)
    resp.delete_cookie("oauth_state", samesite=policy["samesite"], secure=policy["secure"])
    _clear_auth_return(resp, request)
    if is_new:
        _clear_pending_referral(resp, request)
    return resp


# ---------------------------
# Public discovery and SEO infrastructure
# ---------------------------

def _public_study_cards(db: Session, category_slug: Optional[str] = None) -> list[dict]:
    if not INDEX_PUBLIC_STUDIES:
        return []
    query = db.query(Survey).filter(Survey.status == "published")
    if category_slug:
        query = query.filter(func.lower(Survey.category) == category_slug.lower())
    surveys = query.order_by(Survey.published_at.desc(), Survey.created_at.desc()).all()

    slugs_changed = False
    for survey in surveys:
        if not getattr(survey, "share_slug", None):
            _ensure_survey_share_slug(db, survey)
            slugs_changed = True
    if slugs_changed:
        db.commit()

    cards: list[dict] = []
    for survey in surveys:
        display_reward = getattr(survey, "admin_display_reward_amount", None)
        if display_reward is None:
            display_reward = survey.reward_amount
        raw_category = (survey.category or "other").strip().lower()
        normalized_category = raw_category if raw_category in CATEGORY_CONTENT else "other"
        cards.append({
            "id": survey.id,
            "title": survey.title,
            "summary": plain_text(survey.description, 180),
            "description": survey.description,
            "category": normalized_category,
            "category_label": category_label(normalized_category),
            "category_path": f"/studies/{normalized_category}",
            "share_path": f"/r/{survey.share_slug}",
            "image": survey.image_url or category_image(normalized_category),
            "estimated_time": survey.estimated_time,
            "display_reward": float(display_reward or 0),
            "task_type_label": _participant_study_type_label(survey) or _task_type_label(getattr(survey, "task_type", None)),
            "published_at": survey.published_at or survey.created_at,
        })
    return cards


def _category_navigation(cards: list[dict]) -> list[dict]:
    counts = {slug: 0 for slug in CATEGORY_CONTENT}
    for card in cards:
        slug = card.get("category") or "other"
        if slug not in counts:
            slug = "other"
        counts[slug] += 1
    return [
        {
            "slug": slug,
            "label": content["label"],
            "path": f"/studies/{slug}",
            "count": counts.get(slug, 0),
        }
        for slug, content in CATEGORY_CONTENT.items()
    ]


def _sitemap_lastmod(value: Optional[datetime]) -> Optional[str]:
    if not value:
        return None
    return value.date().isoformat()


@app.get("/robots.txt", response_class=PlainTextResponse, include_in_schema=False)
def robots_txt():
    # Private HTML pages remain crawlable long enough for their noindex directives to be seen.
    body = "\n".join([
        "User-agent: *",
        "Allow: /",
        "Disallow: /api/",
        "Disallow: /admin/",
        "Disallow: /auth/",
        "Disallow: /connect/",
        "Disallow: /webhook/",
        "Disallow: /static/uploads/",
        "",
        f"Sitemap: {seo_site_url('/sitemap.xml')}",
        "",
    ])
    return PlainTextResponse(
        body,
        media_type="text/plain; charset=utf-8",
        headers={"Cache-Control": "public, max-age=3600"},
    )


@app.get("/sitemap.xml", include_in_schema=False)
def sitemap_xml(db: Session = Depends(get_db)):
    entries: list[tuple[str, Optional[str]]] = [
        (seo_site_url("/"), None),
        (seo_site_url("/participant"), None),
        (seo_site_url("/studies"), None),
        (seo_site_url("/about"), None),
        (seo_site_url("/privacy"), None),
        (seo_site_url("/terms"), None),
    ]
    cards = _public_study_cards(db)
    category_lastmods: dict[str, Optional[datetime]] = {}
    for card in cards:
        slug = card["category"] if card["category"] in CATEGORY_CONTENT else "other"
        published_at = card.get("published_at")
        current = category_lastmods.get(slug)
        if published_at and (current is None or published_at > current):
            category_lastmods[slug] = published_at
        entries.append((seo_site_url(card["share_path"]), _sitemap_lastmod(published_at)))
    for slug, lastmod in category_lastmods.items():
        entries.append((seo_site_url(f"/studies/{slug}"), _sitemap_lastmod(lastmod)))

    lines = ['<?xml version="1.0" encoding="UTF-8"?>', '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">']
    for loc, lastmod in entries[:50000]:
        lines.append("  <url>")
        lines.append(f"    <loc>{xml_escape(loc)}</loc>")
        if lastmod:
            lines.append(f"    <lastmod>{lastmod}</lastmod>")
        lines.append("  </url>")
    lines.append("</urlset>")
    return HTTPResponse(
        content="\n".join(lines),
        media_type="application/xml; charset=utf-8",
        headers={"Cache-Control": "public, max-age=300, stale-while-revalidate=3600"},
    )


@app.get("/studies", response_class=HTMLResponse)
def public_studies(request: Request, db: Session = Depends(get_db)):
    cards = _public_study_cards(db)
    return templates.TemplateResponse("studies.html", {
        "request": request,
        "surveys": cards,
        "categories": _category_navigation(cards),
        "total_studies": len(cards),
        "active_category": None,
        "page_heading": "Open research studies",
        "page_intro": "Browse surveys, interviews, and other research opportunities. Review the audience, format, estimated time, and reward before deciding whether to participate.",
        "seo": studies_directory_seo(cards),
    })


@app.get("/studies/{category_slug}", response_class=HTMLResponse)
def public_studies_by_category(category_slug: str, request: Request, db: Session = Depends(get_db)):
    normalized = category_slug.strip().lower()
    if normalized not in CATEGORY_CONTENT:
        raise HTTPException(404, "Study category not found")
    all_cards = _public_study_cards(db)
    cards = [card for card in all_cards if card.get("category") == normalized]
    content = category_content(normalized)
    response = templates.TemplateResponse("studies.html", {
        "request": request,
        "surveys": cards,
        "categories": _category_navigation(all_cards),
        "total_studies": len(all_cards),
        "active_category": normalized,
        "page_heading": content["heading"],
        "page_intro": content["intro"],
        "seo": category_seo(normalized, cards),
    })
    if not cards:
        response.headers["X-Robots-Tag"] = "noindex, follow"
    return response


@app.get("/about", response_class=HTMLResponse)
def about_page(request: Request):
    seo = content_page_seo(
        title="About Insighta | Research Recruitment Platform",
        description="Learn how Insighta helps researchers recruit qualified participants and helps people discover surveys, interviews, and studies that fit their background.",
        path="/about",
        page_type="AboutPage",
        breadcrumb_label="About",
    )
    return templates.TemplateResponse("about.html", {"request": request, "seo": seo})


@app.get("/privacy", response_class=HTMLResponse)
def privacy_page(request: Request):
    seo = content_page_seo(
        title="Privacy | Insighta",
        description="Read how Insighta handles account, profile, study, payment, and technical information used to operate the research participation platform.",
        path="/privacy",
        breadcrumb_label="Privacy",
    )
    return templates.TemplateResponse("privacy.html", {"request": request, "seo": seo})


@app.get("/terms", response_class=HTMLResponse)
def terms_page(request: Request):
    seo = content_page_seo(
        title="Terms of Use | Insighta",
        description="Review the core rules for researchers and participants using Insighta to publish, discover, complete, and manage research studies.",
        path="/terms",
        breadcrumb_label="Terms of use",
    )
    return templates.TemplateResponse("terms.html", {"request": request, "seo": seo})


# ---------------------------
# Index
# ---------------------------

@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    # Serve one canonical, responsive homepage to desktop and mobile crawlers alike.
    return templates.TemplateResponse(
        "index.html",
        {"request": request, "show": None, "error": None, "seo": home_seo()}
    )


@app.get("/participant")
def participant_app_entry(request: Request, user_id: str = Cookie(None)):
    if user_id:
        response = RedirectResponse(_participant_dashboard_url(request), status_code=302)
    else:
        response = no_store_response(
            templates.TemplateResponse(
                "participant_landing.html",
                {"request": request, "seo": participant_seo()}
            )
        )
    _mark_participant_app(response, request)
    return response


@app.get("/participant/login")
def participant_login_entry(request: Request):
    response = RedirectResponse("/login?role=participant", status_code=302)
    _mark_participant_app(response, request)
    return response


# ---------------------------
# Login
# ---------------------------

@app.get("/login", response_class=HTMLResponse)
def login_page(
    request: Request,
    success: Optional[str] = None,
    reset_success: Optional[str] = None,
    email: Optional[str] = None,
    oauth_error: Optional[str] = None,
    next: Optional[str] = None,
    role: Optional[str] = None,
    participant_app: Optional[str] = Cookie(None),
):
    normalized_role = role if role in {"participant", "researcher"} else ""
    response = no_store_response(templates.TemplateResponse("login.html", {
        "request": request,
        "error": oauth_error or None,
        "success": success,
        "reset_error": None,
        "reset_success": reset_success,
        "reset_open": False,
        "login_email": _normalize_email(email or ""),
        "reset_email": _normalize_email(email or ""),
        "login_next": next if is_safe_internal_next(next) else "",
        "login_role": normalized_role,
        "participant_app": _auth_uses_participant_app(request, normalized_role, participant_app),
    }))
    if normalized_role == "participant":
        _mark_participant_app(response, request)
    elif normalized_role == "researcher":
        _clear_participant_app(response, request)
    _remember_auth_return(response, request, next)
    return response

@app.post("/login")
def login(
    request: Request,
    email: Optional[str] = Form(None),
    password: str = Form(...),
    next: Optional[str] = Form(None),
    role: Optional[str] = Form(None),
    participant_app: Optional[str] = Cookie(None),
    auth_return_to: Optional[str] = Cookie(None),
    db: Session = Depends(get_db)
):
    next_url = _safe_auth_return(next, auth_return_to)
    normalized_email = _normalize_email(email or "")
    try:
        user = db.query(User).filter(User.email == normalized_email).first()
    except Exception as e:
        normalized_role = role if role in {"participant", "researcher"} else ""
        return no_store_response(templates.TemplateResponse("login.html", {
            "request": request,
            "error": f"Database error: {e}",
            "success": None, "reset_error": None, "reset_success": None,
            "reset_open": False, "login_email": normalized_email, "reset_email": "",
            "login_next": next_url,
            "login_role": normalized_role,
            "participant_app": _auth_uses_participant_app(request, normalized_role, participant_app),
        }))

    if not user or not pwd_context.verify(password, user.password):
        normalized_role = role if role in {"participant", "researcher"} else ""
        return no_store_response(templates.TemplateResponse("login.html", {
            "request": request,
            "error": "Invalid email or password",
            "success": None, "reset_error": None, "reset_success": None,
            "reset_open": False, "login_email": normalized_email, "reset_email": "",
            "login_next": next_url,
            "login_role": normalized_role,
            "participant_app": _auth_uses_participant_app(request, normalized_role, participant_app),
        }))

    final_url = _post_auth_url_with_next(request, participant_app, next_url, role)
    response = RedirectResponse(
        _post_auth_or_onboarding_url(user, final_url, role=role, participant_app=participant_app),
        status_code=303,
    )
    _record_user_event(db, request, "login", user=user, metadata={"role": role or "", "user_role": role or "", "next": next_url or "", "source": "Login"})
    db.commit()
    _set_user_cookie(response, request, user)
    _clear_auth_return(response, request)
    if (role or "").strip().lower() == "participant":
        _mark_participant_app(response, request)
    elif (role or "").strip().lower() == "researcher":
        _clear_participant_app(response, request)
    return response


# ---------------------------
# Password reset
# ---------------------------

@app.post("/password-reset")
def password_reset(
    request: Request,
    email: str = Form(...),
    verification_code: str = Form(...),
    new_password: str = Form(...),
    confirm_password: str = Form(...),
    db: Session = Depends(get_db)
):
    normalized_email = _normalize_email(email)
    user = db.query(User).filter(User.email == normalized_email).first()

    def reset_error(msg):
        return no_store_response(templates.TemplateResponse("login.html", {
            "request": request,
            "error": None, "success": None,
            "reset_error": msg, "reset_success": None,
            "reset_open": True,
            "login_email": "", "reset_email": normalized_email,
        }))

    if not user:
        return reset_error("No account found with this email.")
    if not new_password or not confirm_password:
        return reset_error("Please enter and confirm your new password.")
    if new_password != confirm_password:
        return reset_error("The two passwords do not match.")
    if len(new_password) < 6:
        return reset_error("New password must be at least 6 characters.")
    if not _consume_verification_code(db, normalized_email, "reset_password", verification_code):
        return reset_error("Invalid or expired verification code.")

    user.password = pwd_context.hash(new_password)
    db.commit()

    return no_store_response(templates.TemplateResponse("login.html", {
        "request": request,
        "error": None, "success": None,
        "reset_error": None,
        "reset_success": "Password updated successfully. Please sign in with your new password.",
        "reset_open": False,
        "login_email": normalized_email, "reset_email": "",
    }))


# ---------------------------
# Register
# ---------------------------

@app.get("/register", response_class=HTMLResponse)
def show_register(
    request: Request,
    role: Optional[str] = None,
    next: Optional[str] = None,
    ref: Optional[str] = None,
    participant_app: Optional[str] = Cookie(None),
    pending_referral: Optional[str] = Cookie(None),
):
    normalized_role = role if role in {"participant", "researcher"} else ""
    register_next = next if is_safe_internal_next(next) else ""
    referral_code = _normalize_referral_code(ref) or _normalize_referral_code(pending_referral)
    response = no_store_response(templates.TemplateResponse(
        "register.html",
        {
            "request": request, "error": None, "register_email": "", "register_phone": "", "register_code": "",
            "register_step": 1,
            "register_role": normalized_role,
            "register_next": register_next,
            "register_referral": referral_code,
            "register_values": {},
            "participant_app": _auth_uses_participant_app(request, normalized_role, participant_app),
        }
    ))
    if normalized_role == "participant":
        _mark_participant_app(response, request)
    elif normalized_role == "researcher":
        _clear_participant_app(response, request)
    _remember_auth_return(response, request, register_next)
    if referral_code:
        _remember_pending_referral(response, request, referral_code)
    return response

@app.post("/register", response_class=HTMLResponse)
async def do_register(
    request: Request,
    participant_app: Optional[str] = Cookie(None),
    auth_return_to: Optional[str] = Cookie(None),
    pending_referral: Optional[str] = Cookie(None),
    db: Session = Depends(get_db)
):
    form = await request.form()
    email = _normalize_email(form.get("email") or "")
    phone_number = (form.get("phone_number") or "").strip()
    password = form.get("password") or ""
    confirm = form.get("confirm") or ""
    verification_code = form.get("verification_code") or ""
    referral_input = _normalize_referral_code(form.get("referral_code") or pending_referral)
    next_url = _safe_auth_return(form.get("next"), auth_return_to)
    role = (form.get("role") or "").strip().lower()
    role = role if role in {"participant", "researcher"} else None
    register_values = {
        "first_name": (form.get("first_name") or "").strip(),
        "last_name": (form.get("last_name") or "").strip(),
        "username": (form.get("username") or "").strip(),
        "birth_year": (form.get("birth_year") or "").strip(),
        "birth_month": (form.get("birth_month") or "").strip(),
        "education_level": (form.get("education_level") or "").strip(),
        "field": (form.get("field") or "").strip(),
        "status": (form.get("status") or "").strip(),
        "current_province": (form.get("current_province") or "").strip(),
        "language": (form.get("language") or "").strip(),
        "participation_format": (form.get("participation_format") or "").strip(),
    }

    def reg_error(msg, step: int = 1):
        response = no_store_response(templates.TemplateResponse("register.html", {
            "request": request, "error": msg, "register_email": email,
            "register_phone": phone_number,
            "register_code": verification_code,
            "register_step": step,
            "register_role": role or "",
            "register_next": next_url if is_safe_internal_next(next_url) else "",
            "register_referral": referral_input,
            "register_values": register_values,
            "participant_app": _auth_uses_participant_app(request, role, participant_app),
        }))
        if role == "researcher":
            _clear_participant_app(response, request)
        if referral_input:
            _remember_pending_referral(response, request, referral_input)
        return response

    if not email: return reg_error("Email is required.")
    if password != confirm: return reg_error("Passwords do not match.", step=2)
    pw_error = _validate_registration_password(password)
    if pw_error: return reg_error(pw_error, step=2)
    required_demo_fields = {
        "first_name": "First name is required.",
        "last_name": "Last name is required.",
        "username": "Username is required.",
        "birth_year": "Birth year is required.",
        "birth_month": "Birth month is required.",
        "education_level": "Education level is required.",
        "field": "Field or major is required.",
        "status": "Current status is required.",
        "current_province": "Current state is required.",
        "language": "Primary language is required.",
        "participation_format": "Participation preference is required.",
    }
    for field_name, message in required_demo_fields.items():
        if not register_values.get(field_name):
            return reg_error(message, step=2)
    derived_age_range = _age_range_from_birth_date(register_values["birth_year"], register_values["birth_month"])
    if not derived_age_range:
        return reg_error("Please enter a valid birth year and month. Participants must be 18 or older.", step=2)
    if db.query(User).filter(User.email == email).first():
        return reg_error("Email already exists.")
    if referral_input and not _find_referrer_by_code(db, referral_input):
        return reg_error("Referral code not found.")
    if not _consume_verification_code(db, email, "register", verification_code):
        return reg_error("Invalid or expired verification code.")

    user = User(
        email=email,
        password=pwd_context.hash(password),
        phone_number=phone_number or None,
        first_name=register_values["first_name"],
        last_name=register_values["last_name"],
        username=register_values["username"],
        birth_year=register_values["birth_year"],
        birth_month=register_values["birth_month"],
        age_range=derived_age_range,
        education_level=register_values["education_level"],
        field=register_values["field"],
        status=register_values["status"],
        current_country="United States",
        current_province=register_values["current_province"],
        state=register_values["current_province"],
        language=register_values["language"],
        participation_format=register_values["participation_format"],
        referral_code=_generate_referral_code(db),
    )
    referral_error = _apply_referral_code(db, user, referral_input)
    if referral_error:
        return reg_error(referral_error)
    db.add(user)
    db.commit()
    db.refresh(user)
    _record_user_event(
        db,
        request,
        "sign_up",
        user=user,
        metadata={
            "role": role or "",
            "user_role": role or "",
            "next": next_url or "",
            "source": "Sign up",
            "referral_code": referral_input or "",
            "invited_by_user_id": user.invited_by_user_id or "",
        },
    )
    db.commit()

    final_url = _post_auth_url_with_next(request, participant_app, next_url, role=role, welcome=True)
    response = RedirectResponse(
        _post_auth_or_onboarding_url(user, final_url, role=role, participant_app=participant_app),
        status_code=303,
    )
    _set_user_cookie(response, request, user)
    _clear_auth_return(response, request)
    _clear_pending_referral(response, request)
    if role == "participant":
        _mark_participant_app(response, request)
    elif role == "researcher":
        _clear_participant_app(response, request)
    return response


# ---------------------------
# Current user
# ---------------------------

def get_current_user(
    user_id: str = Cookie(None),
    db: Session = Depends(get_db)
):
    if not user_id:
        raise HTTPException(401, "Not logged in")
    user = db.query(User).filter(User.id == int(user_id)).first()
    if not user:
        raise HTTPException(401, "User not found")
    return user


@app.post("/api/events")
async def record_client_event(
    request: Request,
    user_id: str = Cookie(None),
    db: Session = Depends(get_db),
):
    raw = await request.body()
    try:
        body = json.loads(raw.decode("utf-8")) if raw else {}
    except Exception:
        body = {}

    current_user = None
    if user_id:
        try:
            current_user = db.query(User).filter(User.id == int(user_id)).first()
        except Exception:
            current_user = None

    event_name = (body.get("event_name") or "").strip().lower()
    if not event_name:
        raise HTTPException(400, "event_name is required")
    anonymous_id = (body.get("anonymous_id") or "").strip()
    if current_user and anonymous_id:
        db.query(UserEvent).filter(
            UserEvent.anonymous_id == anonymous_id[:120],
            UserEvent.user_id.is_(None),
        ).update({"user_id": current_user.id}, synchronize_session=False)
    _record_user_event(
        db,
        request,
        event_name,
        user=current_user,
        anonymous_id=anonymous_id,
        target_type=body.get("target_type"),
        target_id=body.get("target_id"),
        page_path=body.get("page_path"),
        metadata=body.get("metadata") if isinstance(body.get("metadata"), dict) else {},
    )
    db.commit()
    return JSONResponse({"success": True})


@app.get("/complete-profile", response_class=HTMLResponse)
def complete_profile_get(
    request: Request,
    next: Optional[str] = None,
    role: Optional[str] = None,
    participant_app: Optional[str] = Cookie(None),
    current_user: User = Depends(get_current_user),
):
    return_to = next if is_safe_internal_next(next) else ""
    if not _needs_identity_onboarding(current_user):
        return RedirectResponse(return_to or _post_auth_url(request), status_code=303)
    return no_store_response(templates.TemplateResponse("complete_profile.html", {
        "request": request,
        "current_user": current_user,
        "next": return_to,
        "role": _welcome_email_role(role, return_to, participant_app),
        "error": None,
        "participant_app": _should_use_participant_app(request, participant_app),
    }))


@app.post("/complete-profile", response_class=HTMLResponse)
async def complete_profile_post(
    request: Request,
    participant_app: Optional[str] = Cookie(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    form = await request.form()
    first_name = (form.get("first_name") or "").strip()
    last_name = (form.get("last_name") or "").strip()
    username = (form.get("username") or "").strip()
    birth_year = (form.get("birth_year") or "").strip()
    birth_month = (form.get("birth_month") or "").strip()
    education_level = (form.get("education_level") or "").strip()
    field = (form.get("field") or "").strip()
    status = (form.get("status") or "").strip()
    current_province = (form.get("current_province") or "").strip()
    language = (form.get("language") or "").strip()
    participation_format = (form.get("participation_format") or "").strip()
    return_to = form.get("next") if is_safe_internal_next(form.get("next")) else ""
    welcome_role = _welcome_email_role(form.get("role"), return_to, participant_app)

    required_values = [
        first_name, last_name, username, birth_year, birth_month, education_level,
        field, status, current_province, language, participation_format
    ]
    if not all(required_values):
        return no_store_response(templates.TemplateResponse("complete_profile.html", {
            "request": request,
            "current_user": current_user,
            "next": return_to,
            "role": welcome_role,
            "error": "Please fill in all required profile basics.",
            "participant_app": _should_use_participant_app(request, participant_app),
        }))
    derived_age_range = _age_range_from_birth_date(birth_year, birth_month)
    if not derived_age_range:
        return no_store_response(templates.TemplateResponse("complete_profile.html", {
            "request": request,
            "current_user": current_user,
            "next": return_to,
            "role": welcome_role,
            "error": "Please enter a valid birth year and month. Participants must be 18 or older.",
            "participant_app": _should_use_participant_app(request, participant_app),
        }))

    current_user.first_name = first_name
    current_user.last_name = last_name
    current_user.username = username
    current_user.birth_year = birth_year
    current_user.birth_month = birth_month
    current_user.age_range = derived_age_range
    current_user.education_level = education_level
    current_user.field = field
    current_user.status = status
    current_user.current_country = "United States"
    current_user.current_province = current_province
    current_user.state = current_province
    current_user.language = language
    current_user.participation_format = participation_format
    if not getattr(current_user, "welcome_email_sent_at", None):
        sent, error = _send_welcome_followup_email(current_user, welcome_role)
        if sent:
            current_user.welcome_email_sent_at = datetime.utcnow()
            current_user.welcome_email_role = welcome_role
        elif error:
            print(f"Welcome email error for user {current_user.id}: {error}")
    _record_user_event(
        db,
        request,
        "profile_update",
        user=current_user,
        metadata={"source": "Complete profile", "user_role": "participant"},
    )
    db.commit()
    return RedirectResponse(return_to or _post_auth_url(request, participant_app), status_code=303)


@app.get("/r/{share_slug}", response_class=HTMLResponse)
def recruitment_share_page(
    share_slug: str,
    request: Request,
    from_: Optional[str] = Query(None, alias="from"),
    user_id: str = Cookie(None),
    db: Session = Depends(get_db),
):
    survey = db.query(Survey).filter(Survey.share_slug == share_slug).first()
    if not survey:
        raise HTTPException(404, "Listing not found")
    if survey.status not in {"published", "closed"}:
        raise HTTPException(404, "Listing is not published")

    current_user = None
    user_response = None
    if user_id:
        try:
            current_user = db.query(User).filter(User.id == int(user_id)).first()
        except Exception:
            current_user = None
    if current_user:
        user_response = db.query(Response).filter(
            Response.survey_id == survey.id,
            Response.participant_id == current_user.id,
        ).first()

    availability_slots = []
    if getattr(survey, "availability_slots", None):
        try:
            parsed_slots = json.loads(survey.availability_slots)
            if isinstance(parsed_slots, list):
                availability_slots = parsed_slots
        except Exception:
            availability_slots = []

    display_reward = getattr(survey, "admin_display_reward_amount", None)
    if display_reward is None:
        display_reward = survey.reward_amount
    booked_slots = []
    for row in db.query(Response.booking_slot).filter(
            Response.survey_id == survey.id,
            Response.booking_slot.isnot(None),
            Response.participant_id != (current_user.id if current_user else 0),
        ).all():
        booked_slots.extend(_parse_booking_slots(row[0]))

    next_path = f"/r/{share_slug}"
    dashboard_path = _participant_dashboard_url(request)
    from_participant_app = from_ == "participant"
    brand_action = "back" if from_participant_app else ("dashboard" if current_user else "home")
    brand_href = dashboard_path if brand_action in {"back", "dashboard"} else "/"
    brand_label = "Back to dashboard" if from_participant_app else ("Dashboard" if current_user else "Insighta")
    _record_user_event(
        db,
        request,
        "listing_view",
        user=current_user,
        target_type="survey",
        target_id=survey.id,
        page_path=next_path,
        metadata={"study_title": survey.title, "source": from_ or "Share link", "surface": "recruitment_share"},
    )
    db.commit()
    normalized_category = (survey.category or "other").strip().lower()
    if normalized_category not in CATEGORY_CONTENT:
        normalized_category = "other"
    seo = study_seo(survey, next_path, indexable=survey.status == "published")
    response = templates.TemplateResponse("recruitment_share.html", {
        "request": request,
        "survey": survey,
        "seo": seo,
        "category_label": category_label(normalized_category),
        "category_path": f"/studies/{normalized_category}",
        "study_image": survey.image_url or category_image(normalized_category),
        "current_user": current_user,
        "user_response": user_response,
        "availability_slots": availability_slots,
        "booked_slots": booked_slots,
        "display_reward": display_reward,
        "share_url": seo_site_url(next_path),
        "login_url": f"/login?{urlencode({'role': 'participant', 'next': next_path})}",
        "register_url": f"/register?{urlencode({'role': 'participant', 'next': next_path})}",
        "dashboard_url": dashboard_path,
        "brand_action": brand_action,
        "brand_href": brand_href,
        "brand_label": brand_label,
        "is_builtin": survey.form_url == "__builtin__",
        "is_interview": _uses_booking_flow(getattr(survey, "task_type", None)),
        "is_online_interview": _is_online_interview_survey(survey),
        "has_external_start_link": _survey_has_external_start_link(survey),
        "external_start_url": _survey_external_start_url(survey),
        "external_start_redirect_url": (
            f"/surveys/{survey.id}/start-redirect?{urlencode({'next': _survey_external_start_url(survey)})}"
            if _survey_external_start_url(survey) else ""
        ),
        "task_type_label": _participant_study_type_label(survey) or _task_type_label(getattr(survey, "task_type", None)),
    })
    if survey.status != "published":
        response.headers["X-Robots-Tag"] = "noindex, follow, noarchive"
    return response

@app.get("/r/{share_slug}/qr.png")
def recruitment_share_qr(
    share_slug: str,
    request: Request,
    download: bool = Query(False),
    db: Session = Depends(get_db),
):
    survey = db.query(Survey).filter(Survey.share_slug == share_slug).first()
    if not survey or survey.status not in {"published", "closed"}:
        raise HTTPException(404, "Listing not found")
    try:
        import qrcode
    except ImportError as exc:
        raise HTTPException(503, "QR generation is not installed") from exc

    share_url = seo_site_url(f"/r/{share_slug}")
    qr = qrcode.QRCode(version=None, box_size=10, border=3)
    qr.add_data(share_url)
    qr.make(fit=True)
    image = qr.make_image(fill_color="#184e77", back_color="white")
    output = BytesIO()
    image.save(output, format="PNG")
    output.seek(0)
    disposition = "attachment" if download else "inline"
    filename = f"insighta-{share_slug}-qr.png"
    return StreamingResponse(
        output,
        media_type="image/png",
        headers={"Content-Disposition": f'{disposition}; filename="{filename}"'},
    )


# ---------------------------
# Choice page
# ---------------------------

@app.get("/choice", response_class=HTMLResponse)
def choice(request: Request, user_id: str = Cookie(None), db: Session = Depends(get_db)):
    if _is_mobile_request(request):
        target_url = _participant_dashboard_url(request) if user_id else "/participant"
    else:
        target_url = "/publisher" if user_id else "/login?role=researcher"
    return no_store_response(RedirectResponse(target_url, status_code=303))

# ---------------------------
# Guide page
# ---------------------------

@app.get("/guide", response_class=HTMLResponse)
def guide_page(request: Request, current_user: User = Depends(get_current_user)):
    return templates.TemplateResponse("guide.html", {"request": request, "current_user": current_user})

# ---------------------------
# Publisher Dashboard
# ---------------------------

@app.get("/publisher", response_class=HTMLResponse)
def publisher_dashboard(
    request: Request,
    user_id: str = Cookie(None),
    db: Session = Depends(get_db)
):
    if _is_mobile_request(request):
        return RedirectResponse(_participant_dashboard_url(request) if user_id else "/participant", status_code=302)

    if not user_id:
        return RedirectResponse("/login", status_code=303)
    try:
        current_user = db.query(User).filter(User.id == int(user_id)).first()
    except:
        return RedirectResponse("/login", status_code=303)
    if not current_user:
        return RedirectResponse("/login", status_code=303)
    all_items = db.query(Survey).filter(Survey.publisher_id == current_user.id).all()
    slugs_changed = False
    for item in all_items:
        if not getattr(item, "share_slug", None):
            _ensure_survey_share_slug(db, item)
            slugs_changed = True
    if slugs_changed:
        db.commit()
    survey_ids = [s.id for s in all_items]

    if not survey_ids:
        completed_map = {}
    else:
        completed_map = dict(
            db.query(
                Response.survey_id,
                func.sum(case((Response.status == "completed", 1), else_=0)).label("completed_cnt"),
            )
            .filter(Response.survey_id.in_(survey_ids))
            .group_by(Response.survey_id)
            .all()
        )

    survey_items = [s for s in all_items if _normalize_task_type(getattr(s, "task_type", None)) == "survey"]
    interview_items = [s for s in all_items if _uses_booking_flow(getattr(s, "task_type", None))]

    return templates.TemplateResponse(
        "publish.html",
        {
            "request": request,
            "surveys": survey_items,
            "interviews": interview_items,
            "completed_map": completed_map,
            "current_user": current_user
        }
    )


@app.get("/publisher/schedule", response_class=HTMLResponse)
def publisher_schedule(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    interviews = db.query(Survey).filter(
        Survey.publisher_id == current_user.id,
        Survey.task_type == "interview",
    ).order_by(Survey.created_at.desc()).all()
    schedule_items = []
    for survey in interviews:
        availability_slots = []
        if getattr(survey, "availability_slots", None):
            try:
                parsed_slots = json.loads(survey.availability_slots)
                if isinstance(parsed_slots, list):
                    availability_slots = parsed_slots
            except Exception:
                availability_slots = []
        rows = db.query(Response, User).join(User, User.id == Response.participant_id).filter(
            Response.survey_id == survey.id,
            Response.booking_slot.isnot(None),
        ).order_by(Response.started_at.desc()).all()
        bookings = [{
            "participant": user.email,
            "participant_name": user.username or user.email,
            "slot": _booking_slots_label(response.booking_slot),
            "slots": _parse_booking_slots(response.booking_slot),
            "status": response.status,
            "started_at": response.started_at,
        } for response, user in rows]
        booked_set = set()
        for item in bookings:
            booked_set.update(item.get("slots") or [])
        schedule_items.append({
            "survey": survey,
            "availability_slots": availability_slots,
            "bookings": bookings,
            "booked_set": booked_set,
            "open_count": max(len(availability_slots) - len(booked_set), 0),
        })
    return templates.TemplateResponse("publisher_schedule.html", {
        "request": request,
        "current_user": current_user,
        "schedule_items": schedule_items,
    })


@app.get("/publisher/study/{survey_id}", response_class=HTMLResponse)
def publisher_study(
    survey_id: int, request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    survey = db.query(Survey).filter(
        Survey.id == survey_id,
        Survey.publisher_id == current_user.id
    ).first()
    if not survey:
        raise HTTPException(404, "Not found")
    completed_count = db.query(Response).filter(
        Response.survey_id == survey_id,
        Response.status == "completed"
    ).count()
    availability_slots = []
    if getattr(survey, "availability_slots", None):
        try:
            parsed_slots = json.loads(survey.availability_slots)
            if isinstance(parsed_slots, list):
                availability_slots = parsed_slots
        except Exception:
            availability_slots = []
    booking_rows = []
    if _uses_booking_flow(getattr(survey, "task_type", None)):
        rows = db.query(Response, User).join(User, User.id == Response.participant_id).filter(
            Response.survey_id == survey_id,
            Response.booking_slot.isnot(None),
        ).order_by(Response.started_at.desc()).all()
        booking_rows = [
            {
                "participant": user.email,
                "slot": _booking_slots_label(response.booking_slot),
                "slots": _parse_booking_slots(response.booking_slot),
                "status": response.status,
                "started_at": response.started_at,
            }
            for response, user in rows
        ]
    return templates.TemplateResponse("publisher_study.html", {
        "request": request,
        "survey": survey,
        "completed_count": completed_count,
        "current_user": current_user,
        "availability_slots": availability_slots,
        "booking_rows": booking_rows,
        "share_url": _survey_share_url(request, db, survey, commit=True),
    })

# ---------------------------
# Delete survey
# ---------------------------
def _delete_survey_tree(db: Session, survey: Survey) -> None:
    survey_id = survey.id
    # Delete dependent rows first. Postgres/Supabase enforces these foreign keys,
    # so deleting the survey directly can fail once quality/AI tracking exists.
    question_ids = [q.id for q in db.query(Question).filter(Question.survey_id == survey_id).all()]
    response_ids = [row[0] for row in db.query(Response.id).filter(Response.survey_id == survey_id).all()]
    if question_ids:
        db.query(Answer).filter(Answer.question_id.in_(question_ids)).delete(synchronize_session=False)
    if response_ids:
        db.query(Answer).filter(Answer.response_id.in_(response_ids)).delete(synchronize_session=False)
        db.query(JumpEvent).filter(JumpEvent.response_id.in_(response_ids)).delete(synchronize_session=False)
        db.query(ResponseQualityCheck).filter(ResponseQualityCheck.response_id.in_(response_ids)).delete(synchronize_session=False)
    db.query(JumpEvent).filter(JumpEvent.survey_id == survey_id).delete(synchronize_session=False)
    db.query(RespondentPrediction).filter(RespondentPrediction.survey_id == survey_id).delete(synchronize_session=False)
    db.query(SurveySegmentStats).filter(SurveySegmentStats.survey_id == survey_id).delete(synchronize_session=False)
    db.query(UserActivityEvent).filter(UserActivityEvent.survey_id == survey_id).delete(synchronize_session=False)
    db.query(ResponseQualityCheck).filter(ResponseQualityCheck.survey_id == survey_id).delete(synchronize_session=False)
    db.query(Question).filter(Question.survey_id == survey_id).delete(synchronize_session=False)
    db.query(Notification).filter(Notification.survey_id == survey_id).delete(synchronize_session=False)
    db.query(Response).filter(Response.survey_id == survey_id).delete(synchronize_session=False)
    db.delete(survey)
    db.commit()


@app.post("/publisher/delete/{survey_id}")
def delete_survey(
    survey_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    survey = db.query(Survey).filter(
        Survey.id == survey_id,
        Survey.publisher_id == current_user.id
    ).first()
    if not survey:
        raise HTTPException(404, "Survey not found")

    _delete_survey_tree(db, survey)
    return RedirectResponse("/publisher", status_code=303)

# ---------------------------
# Dashboard (participant view)
# ---------------------------

URGENCY_RANK = {
    "within_1_week":  3,
    "within_1_month": 2,
    "flexible":       1,
}

def _participant_survey_payload(s: Survey, db: Session, current_user: User, user_response: Optional[Response] = None) -> dict:
    completed_cnt = db.query(Response).filter(
        Response.survey_id == s.id, Response.status == "completed"
    ).count()
    if user_response is None:
        user_response = db.query(Response).filter(
            Response.survey_id == s.id, Response.participant_id == current_user.id
        ).first()

    category_images = {
        "research": "/static/psych.jpg", "life": "/static/campus_life.jpg",
        "clubs": "/static/fb.jpg", "market": "/static/habit.png",
        "academic": "/static/r2.jpg", "other": "/static/food.jpeg"
    }
    form_link = _survey_external_start_url(s)
    response_status = user_response.status if user_response else None
    display_reward = getattr(s, "admin_display_reward_amount", None)
    if display_reward is None:
        display_reward = s.reward_amount
    availability_slots = []
    if getattr(s, "availability_slots", None):
        try:
            parsed_slots = json.loads(s.availability_slots)
            if isinstance(parsed_slots, list):
                availability_slots = parsed_slots
        except Exception:
            availability_slots = []

    return {
        "id": s.id,
        "title": s.title,
        "desc": s.description,
        "niche": getattr(s, "target_niche_requirements", None),
        "benefits": getattr(s, "participant_benefits", None),
        "link": form_link,
        "share_path": _survey_share_path(db, s, commit=True),
        "type": _normalize_task_type(getattr(s, "task_type", None)),
        "format": getattr(s, "target_participation_format", None),
        "type_label": _participant_study_type_label(s),
        "action_label": _participant_study_action_label(s),
        "category": s.category,
        "time": f"{s.estimated_time} min",
        "reward": f"${display_reward:.2f}",
        "responses": f"{completed_cnt}/{s.target_responses}",
        "img": s.image_url if s.image_url else category_images.get(s.category, "/static/psych.jpg"),
        "is_started": response_status == "started",
        "is_completed": response_status == "completed",
        "is_skipped": response_status == "skipped",
        "status": response_status,
        "booking_slot": getattr(user_response, "booking_slot", None) if user_response else None,
        "booking_slots": _parse_booking_slots(getattr(user_response, "booking_slot", None)) if user_response else [],
        "availability_slots": availability_slots,
        "session_count": getattr(s, "session_count", None),
        "sessions_per_week": getattr(s, "sessions_per_week", None),
        "urgency": getattr(s, 'urgency_level', None) or 'flexible',
        "incentive_type": getattr(s, 'incentive_type', None),
    }

@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(
    request: Request,
    timezone_offset: int = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if _is_mobile_request(request):
        tab = request.query_params.get("tab")
        mobile_url = f"/dashboard/mobile?tab={tab}" if tab in {"home", "earnings", "profile"} else "/dashboard/mobile"
        return RedirectResponse(mobile_url, status_code=302)
    all_published = db.query(Survey).filter(Survey.status == "published").all()

    user_edu_min = _education_rank(current_user.education_level, fallback=0)
    user_edu_max = _education_rank(current_user.education_level, fallback=999)

    def survey_matches(s: Survey) -> bool:
        if getattr(s, "payment_status", None) == "admin_demo":
            return True
        if not _age_matches(s.target_age_range, current_user): return False
        if s.target_education_min is not None:
            if user_edu_min < s.target_education_min: return False
        if s.target_education_max is not None:
            if user_edu_max > s.target_education_max: return False
        if not _field_matches(s.target_field, current_user.field): return False
        if not _field_matches(s.target_status, current_user.status): return False
        if not _location_matches(s.target_state, current_user): return False
        if not _language_matches(s.target_language, current_user.language): return False
        if not _field_matches(s.target_ethnicity, current_user.ethnicity): return False
        if not _field_matches(s.target_sexual_orientation, current_user.sexual_orientation): return False
        if not _field_matches(s.target_mental_health_diagnosis, current_user.mental_health_diagnosis): return False
        if not _field_matches(s.target_physical_health_diagnosis, current_user.physical_health_diagnosis): return False
        if not _field_matches(s.target_sport_type, current_user.sport_type): return False
        if not _field_matches(s.target_sport_frequency, current_user.sport_frequency): return False
        if not _field_matches(s.target_smoking, current_user.smoking): return False
        if not _field_matches(s.target_cannabis_use, current_user.cannabis_use): return False
        if not _field_matches(getattr(s, 'target_student_status', None), getattr(current_user, 'student_status', None)): return False
        if not _field_matches(getattr(s, 'target_international_domestic', None), getattr(current_user, 'international_domestic', None)): return False
        if not _tags_match(getattr(s, 'target_experience_tags', None), getattr(current_user, 'experience_tags', None)): return False
        if not _participation_format_matches(getattr(s, 'target_participation_format', None), getattr(current_user, 'participation_format', None)): return False
        if not _device_matches(getattr(s, 'target_device', None), getattr(current_user, 'device_type', None)): return False
        if not _field_matches(getattr(s, 'target_income_level', None), getattr(current_user, 'income_level', None)): return False
        if not _tags_match(getattr(s, 'target_lifestyle_tags', None), getattr(current_user, 'lifestyle_tags', None)): return False
        return True

    participant_response_survey_ids = {
        row[0] for row in db.query(Response.survey_id).filter(
            Response.participant_id == current_user.id
        ).all()
    }
    matched = [
        s for s in all_published
        if s.id not in participant_response_survey_ids and survey_matches(s)
    ]

    # Blend Claude's completion estimate with field/major relevance from the participant profile.
    recommendation_map = recommend_surveys_for_user(db, matched, current_user, use_cache=True)
    matched.sort(
        key=lambda s: (
            _recommendation_sort_score(s, current_user, recommendation_map.get(s.id, {})),
            s.published_at.timestamp() if s.published_at else 0,
        ),
        reverse=True,
    )

    surveys_data = []
    for s in matched:
        llm_rec = recommendation_map.get(s.id, {})
        user_response = db.query(Response).filter(
            Response.survey_id == s.id, Response.participant_id == current_user.id
        ).first()
        payload = _participant_survey_payload(s, db, current_user, user_response)
        payload.update({
            "completion_probability": llm_rec.get("completion_probability"),
            "field_fit_score": _field_relevance_score(s, current_user),
            "recommendation_score": _recommendation_sort_score(s, current_user, llm_rec),
            "ai_confidence": llm_rec.get("confidence"),
            "why_recommended": (llm_rec.get("top_reasons") or [])[:3],
            "risk_reasons": (llm_rec.get("risk_reasons") or [])[:3],
            "recommended_action": llm_rec.get("recommended_action"),
            "ranking_note": llm_rec.get("ranking_note"),
            "model_version": llm_rec.get("model_version"),
        })
        surveys_data.append(payload)

    if timezone_offset is not None:
        user_tz = timezone(timedelta(minutes=-timezone_offset))
        now_user = datetime.now(user_tz)
        today_start_user = now_user.replace(hour=0, minute=0, second=0, microsecond=0)
        today_start_utc = today_start_user.astimezone(timezone.utc)
    else:
        now_utc = datetime.now(timezone.utc)
        today_start_utc = now_utc.replace(hour=0, minute=0, second=0, microsecond=0)

    completed_today = db.query(Response).filter(
        Response.participant_id == current_user.id,
        Response.status == "completed",
        Response.completed_at.isnot(None),
        Response.completed_at >= today_start_utc
    ).count()

    pending_earnings = getattr(current_user, 'pending_earnings', 0.0) or 0.0
    total_withdrawn = getattr(current_user, 'total_withdrawn', 0.0) or 0.0
    db.commit()

    return templates.TemplateResponse(
        "dashboard.html",
        {
            "request": request, "surveys": surveys_data,
            "completed_today": completed_today, "total_earned": pending_earnings,
            "total_withdrawn": total_withdrawn, "pending_earnings": pending_earnings,
            "available_surveys": len(surveys_data), "current_user": current_user,
            "stripe_onboarding_complete": getattr(current_user, 'stripe_onboarding_complete', 'false'),
        }
    )


@app.get("/my-studies", response_class=HTMLResponse)
def participant_my_studies(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    responses = db.query(Response).filter(
        Response.participant_id == current_user.id
    ).order_by(Response.started_at.desc()).all()

    surveys_data = []
    for response in responses:
        survey = db.query(Survey).filter(Survey.id == response.survey_id).first()
        if survey:
            surveys_data.append(_participant_survey_payload(survey, db, current_user, response))

    return templates.TemplateResponse("participant_my_studies.html", {
        "request": request,
        "current_user": current_user,
        "surveys": surveys_data,
        "pending_earnings": getattr(current_user, 'pending_earnings', 0.0) or 0.0,
    })


@app.get("/earnings", response_class=HTMLResponse)
def participant_earnings(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    responses = db.query(Response).filter(
        Response.participant_id == current_user.id,
        Response.status == "completed"
    ).order_by(Response.completed_at.desc()).all()

    surveys_data = []
    for response in responses:
        survey = db.query(Survey).filter(Survey.id == response.survey_id).first()
        if survey:
            surveys_data.append(_participant_survey_payload(survey, db, current_user, response))

    return templates.TemplateResponse("participant_earnings.html", {
        "request": request,
        "current_user": current_user,
        "surveys": surveys_data,
        "pending_earnings": getattr(current_user, 'pending_earnings', 0.0) or 0.0,
        "total_withdrawn": getattr(current_user, 'total_withdrawn', 0.0) or 0.0,
        "stripe_onboarding_complete": getattr(current_user, 'stripe_onboarding_complete', 'false'),
    })
# ---------------------------
# mobile dashboard (simplified for mobile users)
# ---------------------------
@app.get("/dashboard/mobile", response_class=HTMLResponse)
def dashboard_mobile(
    request: Request,
    timezone_offset: int = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    all_published = db.query(Survey).filter(Survey.status == "published").all()

    user_edu_min = _education_rank(current_user.education_level, fallback=0)
    user_edu_max = _education_rank(current_user.education_level, fallback=999)

    def survey_matches(s: Survey) -> bool:
        if getattr(s, "payment_status", None) == "admin_demo":
            return True
        if not _age_matches(s.target_age_range, current_user): return False
        if s.target_education_min is not None:
            if user_edu_min < s.target_education_min: return False
        if s.target_education_max is not None:
            if user_edu_max > s.target_education_max: return False
        if not _field_matches(s.target_field, current_user.field): return False
        if not _field_matches(s.target_status, current_user.status): return False
        if not _location_matches(s.target_state, current_user): return False
        if not _language_matches(s.target_language, current_user.language): return False
        if not _field_matches(s.target_ethnicity, current_user.ethnicity): return False
        if not _field_matches(s.target_sexual_orientation, current_user.sexual_orientation): return False
        if not _field_matches(s.target_mental_health_diagnosis, current_user.mental_health_diagnosis): return False
        if not _field_matches(s.target_physical_health_diagnosis, current_user.physical_health_diagnosis): return False
        if not _field_matches(s.target_sport_type, current_user.sport_type): return False
        if not _field_matches(s.target_sport_frequency, current_user.sport_frequency): return False
        if not _field_matches(s.target_smoking, current_user.smoking): return False
        if not _field_matches(s.target_cannabis_use, current_user.cannabis_use): return False
        if not _field_matches(getattr(s, 'target_student_status', None), getattr(current_user, 'student_status', None)): return False
        if not _field_matches(getattr(s, 'target_international_domestic', None), getattr(current_user, 'international_domestic', None)): return False
        if not _tags_match(getattr(s, 'target_experience_tags', None), getattr(current_user, 'experience_tags', None)): return False
        if not _participation_format_matches(getattr(s, 'target_participation_format', None), getattr(current_user, 'participation_format', None)): return False
        if not _device_matches(getattr(s, 'target_device', None), getattr(current_user, 'device_type', None)): return False
        if not _field_matches(getattr(s, 'target_income_level', None), getattr(current_user, 'income_level', None)): return False
        if not _tags_match(getattr(s, 'target_lifestyle_tags', None), getattr(current_user, 'lifestyle_tags', None)): return False
        return True

    matched = [s for s in all_published if survey_matches(s)]

    # Blend Claude's completion estimate with field/major relevance from the participant profile.
    recommendation_map = recommend_surveys_for_user(db, matched, current_user, use_cache=True)
    matched.sort(
        key=lambda s: (
            _recommendation_sort_score(s, current_user, recommendation_map.get(s.id, {})),
            s.published_at.timestamp() if s.published_at else 0,
        ),
        reverse=True,
    )

    category_images = {
        "research": "/static/psych.jpg", "life": "/static/campus_life.jpg",
        "clubs": "/static/fb.jpg", "market": "/static/habit.png",
        "academic": "/static/r2.jpg", "other": "/static/food.jpeg"
    }

    surveys_data = []
    for s in matched:
        llm_rec = recommendation_map.get(s.id, {})
        completed_cnt = db.query(Response).filter(
            Response.survey_id == s.id, Response.status == "completed"
        ).count()
        user_response = db.query(Response).filter(
            Response.survey_id == s.id, Response.participant_id == current_user.id
        ).first()
        is_completed = user_response and user_response.status == "completed"
        form_link = _survey_external_start_url(s)
        display_reward = getattr(s, "admin_display_reward_amount", None)
        if display_reward is None:
            display_reward = s.reward_amount
        availability_slots = []
        if getattr(s, "availability_slots", None):
            try:
                parsed_slots = json.loads(s.availability_slots)
                if isinstance(parsed_slots, list):
                    availability_slots = parsed_slots
            except Exception:
                availability_slots = []
        surveys_data.append({
            "id": s.id, "title": s.title, "desc": s.description,
            "niche": getattr(s, "target_niche_requirements", None),
            "benefits": getattr(s, "participant_benefits", None),
            "link": form_link,
            "share_path": _survey_share_path(db, s, commit=True),
            "type": _normalize_task_type(getattr(s, "task_type", None)),
            "format": getattr(s, "target_participation_format", None),
            "type_label": _participant_study_type_label(s),
            "action_label": _participant_study_action_label(s),
            "category": s.category, "time": f"{s.estimated_time} min",
            "reward": f"${display_reward:.2f}",
            "responses": f"{completed_cnt}/{s.target_responses}",
            "img": s.image_url if s.image_url else category_images.get(s.category, "/static/psych.jpg"),
            "is_completed": is_completed,
            "booking_slot": getattr(user_response, "booking_slot", None) if user_response else None,
            "booking_slots": _parse_booking_slots(getattr(user_response, "booking_slot", None)) if user_response else [],
            "availability_slots": availability_slots,
            "session_count": getattr(s, "session_count", None),
            "sessions_per_week": getattr(s, "sessions_per_week", None),
            "urgency": getattr(s, 'urgency_level', None) or 'flexible',
            "incentive_type": getattr(s, 'incentive_type', None),
            "completion_probability": llm_rec.get("completion_probability"),
            "field_fit_score": _field_relevance_score(s, current_user),
            "recommendation_score": _recommendation_sort_score(s, current_user, llm_rec),
            "ai_confidence": llm_rec.get("confidence"),
            "why_recommended": (llm_rec.get("top_reasons") or [])[:3],
            "risk_reasons": (llm_rec.get("risk_reasons") or [])[:3],
            "recommended_action": llm_rec.get("recommended_action"),
            "ranking_note": llm_rec.get("ranking_note"),
            "model_version": llm_rec.get("model_version"),
        })

    pending_earnings = getattr(current_user, 'pending_earnings', 0.0) or 0.0
    total_withdrawn = getattr(current_user, 'total_withdrawn', 0.0) or 0.0

    completed_today = db.query(Response).filter(
        Response.participant_id == current_user.id,
        Response.status == "completed",
    ).count()

    return templates.TemplateResponse("dashboard_mobile.html", {
        "request": request,
        "surveys": surveys_data,
        "completed_today": completed_today,
        "pending_earnings": pending_earnings,
        "total_withdrawn": total_withdrawn,
        "available_surveys": len(surveys_data),
        "current_user": current_user,
        "stripe_onboarding_complete": getattr(current_user, 'stripe_onboarding_complete', 'false'),
        "researcher_desktop_url": BASE_URL,
        "researcher_desktop_host": BASE_URL.replace("https://", "").replace("http://", ""),
    })
# ---------------------------
# Dashboard stats API
# ---------------------------

@app.get("/api/dashboard/stats")
def get_dashboard_stats(
    timezone_offset: int = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if timezone_offset is not None:
        user_tz = timezone(timedelta(minutes=-timezone_offset))
        now_user = datetime.now(user_tz)
        today_start_user = now_user.replace(hour=0, minute=0, second=0, microsecond=0)
        today_start_utc = today_start_user.astimezone(timezone.utc)
    else:
        now_utc = datetime.now(timezone.utc)
        today_start_utc = now_utc.replace(hour=0, minute=0, second=0, microsecond=0)

    completed_today = db.query(Response).filter(
        Response.participant_id == current_user.id,
        Response.status == "completed",
        Response.completed_at.isnot(None),
        Response.completed_at >= today_start_utc
    ).count()

    completed_responses = db.query(Response).filter(
        Response.participant_id == current_user.id,
        Response.status == "completed",
        Response.completed_at.isnot(None)
    ).all()

    total_earned = 0.0
    for resp in completed_responses:
        survey = db.query(Survey).filter(Survey.id == resp.survey_id).first()
        if survey:
            total_earned += survey.reward_amount

    return JSONResponse({"completed_today": completed_today, "total_earned": total_earned})


# ---------------------------
# Survey start / complete / modify
# ---------------------------

@app.post("/surveys/{survey_id}/start")
async def start_survey(
    survey_id: int,
    request: Request,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    survey = db.query(Survey).filter(Survey.id == survey_id).first()
    if not survey: raise HTTPException(404, "Survey not found")
    if survey.status != "published": raise HTTPException(400, "Survey not published")

    existing = db.query(Response).filter(
        Response.survey_id == survey_id, Response.participant_id == current_user.id
    ).first()
    booking_slots = []
    try:
        body = await request.json()
        if isinstance(body, dict):
            if isinstance(body.get("booking_slots"), list):
                booking_slots = [str(item).strip() for item in body.get("booking_slots") if str(item).strip()]
            else:
                booking_slots = _parse_booking_slots((body.get("booking_slot") or "").strip())
    except Exception:
        booking_slots = []
    booking_slot = _serialize_booking_slots(booking_slots)

    should_schedule_followup = False
    should_send_booking_email = False
    response = existing

    if booking_slots and _uses_booking_flow(getattr(survey, "task_type", None)):
        existing_bookings = db.query(Response).filter(
            Response.survey_id == survey_id,
            Response.booking_slot.isnot(None),
            Response.participant_id != current_user.id,
        ).all()
        taken_slots = set()
        for existing_booking in existing_bookings:
            taken_slots.update(_parse_booking_slots(existing_booking.booking_slot))
        if any(slot in taken_slots for slot in booking_slots):
            raise HTTPException(409, "That time was just booked. Please choose another available time.")

    if not response:
        response = Response(
            survey_id=survey_id,
            participant_id=current_user.id,
            status="started",
            booking_slot=booking_slot,
            start_followup_scheduled_at=datetime.utcnow(),
        )
        db.add(response)
        db.commit()
        db.refresh(response)
        should_schedule_followup = True
        should_send_booking_email = bool(booking_slot)
    elif response.status != "completed":
        response.status = "started"
        response.completed_at = None
        response.started_at = datetime.now(timezone.utc)
        if booking_slot:
            should_send_booking_email = response.booking_slot != booking_slot
            response.booking_slot = booking_slot
        db.commit()
    elif booking_slot:
        should_send_booking_email = response.booking_slot != booking_slot
        response.booking_slot = booking_slot
        db.commit()

    if response and not response.start_followup_sent_at and not response.start_followup_scheduled_at:
        response.start_followup_scheduled_at = datetime.utcnow()
        db.commit()
        should_schedule_followup = True

    if should_schedule_followup:
        background_tasks.add_task(
            _send_survey_start_followup_after_delay,
            response.id,
            _absolute_url(request, "/dashboard"),
        )
    _record_user_event(
        db,
        request,
        "study_start",
        user=current_user,
        target_type="survey",
        target_id=survey.id,
        page_path=f"/surveys/{survey.id}/start",
        metadata={
            "study_title": survey.title,
            "source": _event_source_from_request(request, "Study Page"),
            "booking_slot": _booking_slots_label(booking_slot) if booking_slot else "",
            "booking_slots": booking_slots,
            "task_type": _normalize_task_type(getattr(survey, "task_type", None)),
            "user_role": "participant",
        },
    )
    db.commit()

    if should_send_booking_email and booking_slot:
        background_tasks.add_task(_send_booking_confirmation_emails_for_response, response.id, booking_slot)

    return {"message": "Survey started successfully"}

@app.get("/surveys/{survey_id}/start-redirect")
def start_survey_and_redirect(
    survey_id: int,
    request: Request,
    next_url: str = Query("", alias="next"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    survey = db.query(Survey).filter(Survey.id == survey_id).first()
    if not survey:
        raise HTTPException(404, "Survey not found")
    if survey.status != "published":
        raise HTTPException(400, "Survey not published")

    target_url = (next_url or "").strip()
    if not (target_url.startswith("https://") or target_url.startswith("http://")):
        target_url = _participant_dashboard_url(request)

    response = db.query(Response).filter(
        Response.survey_id == survey_id,
        Response.participant_id == current_user.id,
    ).first()
    if not response:
        response = Response(
            survey_id=survey_id,
            participant_id=current_user.id,
            status="started",
            start_followup_scheduled_at=datetime.utcnow(),
        )
        db.add(response)
        db.commit()
        db.refresh(response)
    elif response.status != "completed":
        response.status = "started"
        response.completed_at = None
        response.started_at = datetime.now(timezone.utc)
        if not response.start_followup_sent_at and not response.start_followup_scheduled_at:
            response.start_followup_scheduled_at = datetime.utcnow()
        db.commit()

    _record_user_event(
        db,
        request,
        "study_start",
        user=current_user,
        target_type="survey",
        target_id=survey.id,
        page_path=f"/surveys/{survey.id}/start-redirect",
        metadata={
            "study_title": survey.title,
            "source": _event_source_from_request(request, "Study Page"),
            "task_type": _normalize_task_type(getattr(survey, "task_type", None)),
            "user_role": "participant",
            "redirect_target": target_url,
        },
    )
    db.commit()

    return RedirectResponse(target_url, status_code=303)


@app.post("/surveys/{survey_id}/complete")
def complete_survey(
    survey_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    survey = db.query(Survey).filter(Survey.id == survey_id).first()
    if not survey: raise HTTPException(404, "Survey not found")
    if survey.status != "published": raise HTTPException(400, "Survey not published")

    r = db.query(Response).filter(
        Response.survey_id == survey_id, Response.participant_id == current_user.id
    ).first()

    if not r:
        r = Response(survey_id=survey_id, participant_id=current_user.id, status="started")
        db.add(r)

    if r.status != "completed":
        r.status = "completed"
        r.completed_at = datetime.now(timezone.utc)
        r.payout_amount = survey.reward_amount
        mark_response_under_review(r)

        publisher = db.query(User).filter(User.id == survey.publisher_id).first()
        if publisher and publisher.email:
            send_email(
                to=publisher.email,
                subject=f"[Insighta] New response ready for review: {survey.title}",
                body=f"""
                <div style="font-family: sans-serif; max-width: 560px; margin: 0 auto; padding: 32px 24px; color: #1a1a18;">
                  <h2 style="font-size: 22px; margin-bottom: 8px;">📋 Response Ready for Review</h2>
                  <p style="color: #8a8a82; margin-bottom: 24px;">A participant has completed your survey and is awaiting your approval.</p>
                  <div style="background: #f3f1ea; border-radius: 10px; padding: 20px 24px; margin-bottom: 24px;">
                    <div style="font-size: 13px; color: #8a8a82; margin-bottom: 4px;">Survey</div>
                    <div style="font-size: 17px; font-weight: 600; margin-bottom: 12px;">{survey.title}</div>
                    <div style="font-size: 13px; color: #8a8a82; margin-bottom: 4px;">Participant</div>
                    <div style="font-size: 15px;">{current_user.email}</div>
                    <div style="font-size: 13px; color: #8a8a82; margin-top: 12px; margin-bottom: 4px;">Reward</div>
                    <div style="font-size: 15px; font-weight: 600; color: #2d6a4f;">${survey.reward_amount:.2f}</div>
                  </div>
                  <a href="https://insightaco.org/publisher" style="display: inline-block; padding: 12px 24px; background: #2d6a4f; color: white; border-radius: 8px; text-decoration: none; font-weight: 600; font-size: 14px;">
                    Review & Approve →
                  </a>
                </div>
                """
            )
        _record_user_event(
            db,
            request,
            "study_complete",
            user=current_user,
            target_type="survey",
            target_id=survey.id,
            page_path=f"/surveys/{survey.id}/take",
            metadata={
                "study_title": survey.title,
                "task_type": _normalize_task_type(getattr(survey, "task_type", None)),
                "reward_amount": survey.reward_amount,
                "source": "Survey",
                "user_role": "participant",
            },
        )

    db.commit()
    return RedirectResponse(url="/dashboard", status_code=302)


@app.post("/surveys/{survey_id}/modify")
def modify_completed_survey(
    survey_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    r = db.query(Response).filter(
        Response.survey_id == survey_id, Response.participant_id == current_user.id
    ).first()

    if r and r.status == "completed":
        survey = db.query(Survey).filter(Survey.id == survey_id).first()
        r.status = "started"; r.completed_at = None
        return_response_to_review(db, r)
        db.commit()
        return {"message": "Response modified"}

    return JSONResponse({"detail": "Response not found or not completed"}, status_code=404)


# ---------------------------
# Notifications API
# ---------------------------

@app.get("/api/notifications")
def get_notifications(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    notifs = db.query(Notification).filter(
        Notification.publisher_id == current_user.id
    ).order_by(Notification.created_at.desc()).all()
    return JSONResponse([{
        "id": n.id,
        "participant_email": n.participant_email,
        "survey_title": n.survey_title,
        "task_type": _normalize_task_type(n.task_type),
        "task_type_label": _task_type_label(n.task_type),
        "status": n.status,
        "created_at": n.created_at.strftime("%b %d, %H:%M") if n.created_at else "",
    } for n in notifs])


@app.post("/api/notifications/{notif_id}/accept")
def accept_notification(
    notif_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    notif = db.query(Notification).filter(
        Notification.id == notif_id,
        Notification.publisher_id == current_user.id,
    ).first()
    if not notif:
        raise HTTPException(404, "Notification not found")
    notif.status = "accepted"

    response = db.query(Response).filter(
        Response.survey_id == notif.survey_id,
        Response.participant_id == notif.participant_id,
    ).first()
    if response:
        release_response_payout(db, response)
        survey = db.query(Survey).filter(Survey.id == notif.survey_id).first()
        participant = db.query(User).filter(User.id == notif.participant_id).first()
        if survey:
            _record_user_event(
                db,
                request,
                "participant_approved",
                user=current_user,
                target_type="survey",
                target_id=survey.id,
                metadata={
                    "study_title": survey.title,
                    "participant_id": notif.participant_id,
                    "participant_email": notif.participant_email,
                    "source": "Notification",
                    "user_role": "researcher",
                },
            )
        if participant and survey:
            send_email(
                to=participant.email,
                subject="[Insighta] Your response was approved",
                body=f"""
                <div style="font-family: sans-serif; max-width: 560px; margin: 0 auto; padding: 32px 24px; color: #1a1a18;">
                  <h2 style="font-size: 22px; margin-bottom: 8px;">Response Approved</h2>
                  <p style="color: #8a8a82; margin-bottom: 24px;">Your response has been verified and approved.</p>
                  <div style="background: #f3f1ea; border-radius: 10px; padding: 20px 24px; margin-bottom: 24px;">
                    <div style="font-size: 13px; color: #8a8a82; margin-bottom: 4px;">Survey</div>
                    <div style="font-size: 17px; font-weight: 600; margin-bottom: 12px;">{survey.title}</div>
                    <div style="font-size: 13px; color: #8a8a82; margin-bottom: 4px;">Reward</div>
                    <div style="font-size: 22px; font-weight: 700; color: #2d6a4f;">${survey.reward_amount:.2f}</div>
                  </div>
                </div>
                """,
            )
    db.commit()
    return {"message": "accepted"}


@app.post("/api/notifications/{notif_id}/reject")
def reject_notification(
    notif_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    notif = db.query(Notification).filter(
        Notification.id == notif_id,
        Notification.publisher_id == current_user.id,
    ).first()
    if not notif:
        raise HTTPException(404, "Notification not found")
    notif.status = "rejected"

    response = db.query(Response).filter(
        Response.survey_id == notif.survey_id,
        Response.participant_id == notif.participant_id,
    ).first()
    if response:
        reject_response_payout(db, response)
    survey = db.query(Survey).filter(Survey.id == notif.survey_id).first()
    if survey:
        _record_user_event(
            db,
            request,
            "participant_rejected",
            user=current_user,
            target_type="survey",
            target_id=survey.id,
            metadata={
                "study_title": survey.title,
                "participant_id": notif.participant_id,
                "participant_email": notif.participant_email,
                "source": "Notification",
                "user_role": "researcher",
            },
        )
    db.commit()
    return {"message": "rejected"}


# ---------------------------
# Publisher get pending responses
# ---------------------------

@app.get("/api/publisher/pending-responses")
def get_pending_responses(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    surveys = db.query(Survey).filter(Survey.publisher_id == current_user.id).all()
    survey_ids = [s.id for s in surveys]
    survey_map = {s.id: s for s in surveys}
    if not survey_ids:
        return JSONResponse([])
    pending = db.query(Response).filter(
        Response.survey_id.in_(survey_ids), Response.status == "started"
    ).all()
    result = []
    for r in pending:
        participant = db.query(User).filter(User.id == r.participant_id).first()
        survey = survey_map.get(r.survey_id)
        quality = db.query(ResponseQualityCheck).filter(
            ResponseQualityCheck.response_id == r.id
        ).order_by(ResponseQualityCheck.created_at.desc()).first()
        result.append({
            "response_id": r.id, "survey_id": r.survey_id,
            "survey_title": survey.title if survey else "Unknown",
            "participant_email": participant.email if participant else "Unknown",
            "participant_name": participant.username or participant.email if participant else "Unknown",
            "reward": survey.reward_amount if survey else 0,
            "started_at": str(r.started_at),
            "quality_score": quality.quality_score if quality else None,
            "quality_label": quality.quality_label if quality else None,
            "fraud_risk": quality.fraud_risk if quality else None,
            "quality_reasons": quality.reasons if quality else [],
        })
    return JSONResponse(result)


# ---------------------------
# Publish survey page
# ---------------------------
@app.get("/publish", response_class=HTMLResponse)
def publish_page(
    request: Request,
    builtin: int = 0,
    survey_id: Optional[int] = None,
    title: Optional[str] = None,
    desc: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    existing_survey = None
    if survey_id:
        existing_survey = db.query(Survey).filter(
            Survey.id == survey_id,
            Survey.publisher_id == current_user.id,
        ).first()
    is_builtin = bool(builtin) or (
        existing_survey is not None and existing_survey.form_url == "__builtin__"
    )
    return templates.TemplateResponse("publish_external.html", {
        "request": request,
        "stripe_publishable_key": STRIPE_PUBLISHABLE_KEY,
        "builtin": is_builtin,
        "existing_survey_id": survey_id or 0,
        "prefill_title": title or (existing_survey.title if existing_survey else ""),
        "prefill_desc": desc or (existing_survey.description if existing_survey else ""),
        "existing_survey": existing_survey,
        "current_user": current_user
    })

# ---------------------------
# Calculate price API
# ---------------------------

@app.post("/api/calculate-price")
async def calculate_price(request: Request, current_user: User = Depends(get_current_user)):
    body = await request.json()
    per_person = body.get("per_person")
    total_budget = body.get("total_budget")
    target_responses = body.get("target_responses", 1)
    if per_person:
        ppg = float(per_person)
    elif total_budget and target_responses:
        ppg = float(total_budget) / int(target_responses)
    else:
        raise HTTPException(400, "Provide per_person or total_budget + target_responses")
    rate, reward = calculate_commission(ppg)
    total = round(ppg * int(target_responses), 2)
    return JSONResponse({
        "per_person_gross": round(ppg, 2), "commission_rate": rate,
        "commission_pct": int(rate * 100), "reward_amount": reward, "total_budget": total,
    })


# ---------------------------
# Publish survey → Stripe Checkout
# ---------------------------



# ---------------------------
# Publish in-person study
# ---------------------------

@app.get("/publish_interview", response_class=HTMLResponse)
def publish_interview_page(
    request: Request,
    mode: str = Query("in_person"),
    current_user: User = Depends(get_current_user),
):
    is_online = (mode or "").strip().lower() in {"online", "remote", "video"}
    return templates.TemplateResponse("publish_interview.html", {
        "request": request,
        "current_user": current_user,
        "is_online_interview": is_online,
        "interview_mode": "online" if is_online else "in_person",
    })

@app.post("/publish_interview")
async def publish_interview(
    request: Request, title: str = Form(...), description: str = Form(...),
    category: str = Form(...), estimated_time: int = Form(...), target_responses: int = Form(...),
    interview_format: str = Form("video"), scheduling_link: Optional[str] = Form(None),
    availability_notes: Optional[str] = Form(None), interview_location: Optional[str] = Form(None),
    urgency_level: Optional[str] = Form(None), deadline_date: Optional[str] = Form(None),
    incentive_type: Optional[str] = Form(None), per_person_gross: Optional[str] = Form(None),
    current_user: User = Depends(get_current_user), db: Session = Depends(get_db)
):
    form = await request.form()
    experience_list = form.getlist("target_experience_tags")
    lifestyle_list = form.getlist("target_lifestyle_tags")
    target_education_min = _parse_optional_int(form.get("target_education_min"))
    target_education_max = _parse_optional_int(form.get("target_education_max"))
    availability_slots = (form.get("availability_slots") or "").strip()
    if availability_slots:
        try:
            parsed_slots = json.loads(availability_slots)
            availability_slots = json.dumps(parsed_slots if isinstance(parsed_slots, list) else [])
        except Exception:
            availability_slots = None
    else:
        availability_slots = None
    incentive_clean = _clean_target(incentive_type) or "cash"
    is_no_pay = incentive_clean in ("raffle", "volunteer")
    session_count = max(_parse_optional_int(form.get("session_count")) or 1, 1)
    sessions_per_week = _parse_optional_int(form.get("sessions_per_week"))
    try:
        per_person_value = float(per_person_gross) if str(per_person_gross or "").strip() else 0.0
    except (TypeError, ValueError):
        per_person_value = 0.0
    reward = 0.0 if is_no_pay else per_person_value

    survey = Survey(
        publisher_id=current_user.id, title=title, description=description,
        form_url=scheduling_link or "", task_type="in_person", category=category,
        estimated_time=estimated_time, reward_amount=reward, per_person_gross=reward,
        total_budget=round(reward * target_responses * session_count, 2), commission_rate=0.0, payment_status="paid",
        target_responses=target_responses, urgency_level=_clean_target(urgency_level),
        incentive_type=incentive_clean,
        raffle_prize_type=_clean_target(form.get("raffle_prize_type")) if incentive_clean == "raffle" else None,
        target_age_range=_clean_target(form.get("target_age_range")),
        target_education_min=target_education_min,
        target_education_max=target_education_max,
        target_field=_clean_target(form.get("target_field")),
        target_status=_clean_target(form.get("target_status")),
        target_state=_clean_target(form.get("target_state")),
        target_language=_clean_target(form.get("target_language")),
        target_ethnicity=_clean_target(form.get("target_ethnicity")),
        target_sexual_orientation=_clean_target(form.get("target_sexual_orientation")),
        target_mental_health_diagnosis=_clean_target(form.get("target_mental_health_diagnosis")),
        target_physical_health_diagnosis=_clean_target(form.get("target_physical_health_diagnosis")),
        target_sport_type=_clean_target(form.get("target_sport_type")),
        target_sport_frequency=_clean_target(form.get("target_sport_frequency")),
        target_smoking=_clean_target(form.get("target_smoking")),
        target_cannabis_use=_clean_target(form.get("target_cannabis_use")),
        target_student_status=_clean_target(form.get("target_student_status")),
        target_year_in_school=None,
        target_international_domestic=_clean_target(form.get("target_international_domestic")),
        target_experience_tags=",".join(experience_list) if experience_list else None,
        target_participation_format=_clean_target(form.get("target_participation_format")),
        target_device=_clean_target(form.get("target_device")),
        target_income_level=_clean_target(form.get("target_income_level")),
        target_lifestyle_tags=",".join(lifestyle_list) if lifestyle_list else None,
        target_niche_requirements=_clean_target(form.get("target_niche_requirements")),
        availability_slots=availability_slots,
        interview_location=_clean_target(interview_location),
        session_count=session_count,
        sessions_per_week=sessions_per_week,
        status="published", published_at=datetime.utcnow(), closed_at=None,
    )
    _ensure_survey_share_slug(db, survey)
    db.add(survey); db.commit(); db.refresh(survey)
    _record_user_event(
        db,
        request,
        "study_created",
        user=current_user,
        target_type="survey",
        target_id=survey.id,
        metadata={"study_title": survey.title, "source": "Interview publish", "user_role": "researcher"},
    )
    _record_user_event(
        db,
        request,
        "study_published",
        user=current_user,
        target_type="survey",
        target_id=survey.id,
        metadata={"study_title": survey.title, "source": "Interview publish", "user_role": "researcher"},
    )
    db.commit()
    return RedirectResponse("/publisher", status_code=303)


# ---------------------------
# Payment success
# ---------------------------

@app.get("/payment/success", response_class=HTMLResponse)
def payment_success(
    request: Request,
    survey_id: int = Query(None),
    user_id: str = Cookie(None),
    db: Session = Depends(get_db)
):
    current_user = None
    if user_id:
        try:
            current_user = db.query(User).filter(User.id == int(user_id)).first()
        except:
            pass
    survey = db.query(Survey).filter(Survey.id == survey_id).first() if survey_id else None

    # Fallback: mark paid and publish when landing on payment success page
    if survey and survey.payment_status != "paid":
        was_published = survey.status == "published"
        survey.payment_status = "paid"
        survey.status = "published"
        survey.published_at = datetime.utcnow()
        if not was_published:
            _record_user_event(
                db,
                request,
                "study_published",
                user=current_user,
                target_type="survey",
                target_id=survey.id,
                metadata={"study_title": survey.title, "source": "Payment success", "user_role": "researcher"},
            )
        db.commit()

    return templates.TemplateResponse("payment_success.html", {
        "request": request, "survey": survey, "current_user": current_user,
    })

# ---------------------------
# Stripe Webhook
# ---------------------------

@app.post("/webhook/stripe")
async def stripe_webhook(request: Request, db: Session = Depends(get_db)):
    payload = await request.body()
    sig_header = request.headers.get("stripe-signature")
    try:
        event = stripe.Webhook.construct_event(payload, sig_header, STRIPE_WEBHOOK_SECRET)
    except Exception as e:
        raise HTTPException(400, str(e))

    if event["type"] == "checkout.session.completed":
        survey_id = event["data"]["object"].get("metadata", {}).get("survey_id")
        if survey_id:
            survey = db.query(Survey).filter(Survey.id == int(survey_id)).first()
            if survey:
                was_published = survey.status == "published"
                survey.payment_status = "paid"
                survey.status = "published"
                survey.published_at = datetime.utcnow()
                if not was_published:
                    publisher = db.query(User).filter(User.id == survey.publisher_id).first()
                    _record_user_event(
                        db,
                        request,
                        "study_published",
                        user=publisher,
                        target_type="survey",
                        target_id=survey.id,
                        metadata={"study_title": survey.title, "source": "Stripe webhook", "user_role": "researcher"},
                    )
                db.commit()
    elif event["type"] == "account.updated":
        account = event["data"]["object"]
        if account.get("charges_enabled"):
            user = db.query(User).filter(User.stripe_account_id == account.get("id")).first()
            if user:
                user.stripe_onboarding_complete = "true"
                db.commit()

    return JSONResponse({"status": "ok"})


# ---------------------------
# Stripe Connect
# ---------------------------

@app.get("/connect/onboard")
def connect_onboard(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not getattr(current_user, 'stripe_account_id', None):
        account = stripe.Account.create(
            type="express",
            email=current_user.email,
            business_type="individual",
            individual={"email": current_user.email},
            capabilities={"transfers": {"requested": True}}
        )
        current_user.stripe_account_id = account.id; db.commit()
    account_link = stripe.AccountLink.create(
        account=current_user.stripe_account_id,
        refresh_url="https://insightaco.org/connect/onboard",
        return_url="https://insightaco.org/connect/complete",
        type="account_onboarding",
    )
    return RedirectResponse(account_link.url)
@app.get("/logout")
def logout(request: Request):
    response = RedirectResponse("/login", status_code=303)
    policy = _cookie_policy(request)
    response.delete_cookie("user_id", samesite=policy["samesite"], secure=policy["secure"])
    return response

@app.get("/connect/complete", response_class=HTMLResponse)
def connect_complete(request: Request, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if getattr(current_user, 'stripe_account_id', None):
        account = stripe.Account.retrieve(current_user.stripe_account_id)
        if account.charges_enabled:
            current_user.stripe_onboarding_complete = "true"
            db.commit()
    return RedirectResponse("/dashboard")


# ---------------------------
# Participant withdrawal
# ---------------------------

@app.post("/api/withdraw")
async def withdraw(request: Request, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if getattr(current_user, 'stripe_onboarding_complete', 'false') != 'true':
        raise HTTPException(400, "Please complete Stripe onboarding first")
    pending = getattr(current_user, 'pending_earnings', 0.0) or 0.0
    if pending < 1.0: raise HTTPException(400, "Minimum withdrawal is $1.00")
    try:
        transfer = stripe.Transfer.create(amount=int(pending * 100), currency="usd", destination=current_user.stripe_account_id, description=f"Insighta payout for user {current_user.id}")
        for r in db.query(Response).filter(
            Response.participant_id == current_user.id,
            Response.payout_status.in_([APPROVED, LEGACY_RELEASED]),
        ).all():
            r.payout_status = "paid"; r.stripe_transfer_id = transfer.id
        current_user.total_withdrawn = (getattr(current_user, 'total_withdrawn', 0.0) or 0.0) + pending
        current_user.pending_earnings = 0.0; db.commit()
        return JSONResponse({"success": True, "amount": pending, "transfer_id": transfer.id})
    except stripe.error.StripeError as e:
        raise HTTPException(400, str(e))


# ---------------------------
# Survey status management
# ---------------------------

@app.post("/surveys/{survey_id}/publish")
def publish_existing_survey(
    survey_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    s = db.query(Survey).filter(Survey.id == survey_id, Survey.publisher_id == current_user.id).first()
    if not s: raise HTTPException(404, "Survey not found")
    if getattr(s, 'payment_status', 'unpaid') != 'paid': raise HTTPException(400, "Survey must be paid before publishing")
    if s.status != "published":
        s.status = "published"
        s.published_at = datetime.utcnow()
        s.closed_at = None
        _record_user_event(
            db,
            request,
            "study_published",
            user=current_user,
            target_type="survey",
            target_id=s.id,
            metadata={"study_title": s.title, "source": "Publisher", "user_role": "researcher"},
        )
    db.commit()
    return RedirectResponse("/publisher", status_code=303)

@app.post("/surveys/{survey_id}/close")
def close_existing_survey(survey_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    s = db.query(Survey).filter(Survey.id == survey_id, Survey.publisher_id == current_user.id).first()
    if not s: raise HTTPException(404, "Survey not found")
    if s.status != "closed": s.status = "closed"; s.closed_at = datetime.utcnow()
    db.commit()
    return RedirectResponse("/publisher", status_code=303)

@app.post("/surveys/{survey_id}/reopen")
def reopen_closed_survey(survey_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    s = db.query(Survey).filter(Survey.id == survey_id, Survey.publisher_id == current_user.id).first()
    if not s: raise HTTPException(404, "Survey not found")
    if s.status == "closed": s.status = "published"; s.closed_at = None
    db.commit()
    return RedirectResponse("/publisher", status_code=303)


# ---------------------------
# Edit survey
# ---------------------------

@app.get("/publisher/edit/{survey_id}")
def edit_survey_get(
    request: Request,
    survey_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    survey = db.query(Survey).filter(
        Survey.id == survey_id,
        Survey.publisher_id == current_user.id,
    ).first()
    if not survey:
        raise HTTPException(404, "Survey not found")
    current_responses = db.query(Response).filter(Response.survey_id == survey_id, Response.status == "completed").count()
    survey.current_responses = current_responses
    return templates.TemplateResponse("edit_publish.html", {
        "request": request,
        "survey": survey,
        "current_user": current_user,
    })

@app.post("/publisher/edit/{survey_id}")
async def edit_survey_post(
    request: Request, survey_id: int,
    title: str = Form(...), description: str = Form(...), form_url: str = Form(...),
    task_type: str = Form("survey"), category: str = Form(...),
    estimated_time: int = Form(...), reward_amount: float = Form(...), additional_needed: int = Form(...),
    urgency_level: str = Form(None), incentive_type: str = Form(None),
    target_age_range: str = Form(None), target_field: str = Form(None),
    target_status: str = Form(None), target_state: str = Form(None),
    target_language: str = Form(None), target_ethnicity: str = Form(None),
    target_sexual_orientation: str = Form(None), target_mental_health_diagnosis: str = Form(None),
    target_physical_health_diagnosis: str = Form(None), target_sport_type: str = Form(None),
    target_sport_frequency: str = Form(None), target_smoking: str = Form(None),
    target_cannabis_use: str = Form(None), target_student_status: str = Form(None),
    target_year_in_school: str = Form(None), target_international_domestic: str = Form(None),
    target_participation_format: str = Form(None), target_device: str = Form(None),
    target_income_level: str = Form(None), raffle_prize_type: str = Form(None),
    cover_image: UploadFile = File(None),
    current_user: User = Depends(get_current_user), db: Session = Depends(get_db)
):
    form = await request.form()
    target_education_min = _parse_optional_int(form.get("target_education_min"))
    target_education_max = _parse_optional_int(form.get("target_education_max"))
    experience_list = form.getlist("target_experience_tags")
    target_experience_tags = ",".join(experience_list) if experience_list else None
    lifestyle_list = form.getlist("target_lifestyle_tags")
    target_lifestyle_tags = ",".join(lifestyle_list) if lifestyle_list else None

    survey = db.query(Survey).filter(Survey.id == survey_id, Survey.publisher_id == current_user.id).first()
    if not survey:
        raise HTTPException(404, "Survey not found")
    if survey.status == "published":
        reward_amount = survey.reward_amount

    current_responses = db.query(Response).filter(Response.survey_id == survey_id, Response.status == "completed").count()

    # Update all survey fields
    survey.title = title; survey.description = description; survey.form_url = form_url
    survey.task_type = task_type; survey.category = category
    survey.estimated_time = estimated_time; survey.reward_amount = reward_amount
    survey.target_responses = current_responses + additional_needed
    survey.urgency_level = _clean_target(urgency_level); survey.incentive_type = _clean_target(incentive_type)
    survey.raffle_prize_type = _clean_target(raffle_prize_type) if _clean_target(incentive_type) == "raffle" else None
    survey.target_age_range = _clean_target(target_age_range)
    survey.target_education_min = target_education_min; survey.target_education_max = target_education_max
    survey.target_field = _clean_target(target_field); survey.target_status = _clean_target(target_status)
    survey.target_state = _clean_target(target_state); survey.target_language = _clean_target(target_language)
    survey.target_ethnicity = _clean_target(target_ethnicity)
    survey.target_sexual_orientation = _clean_target(target_sexual_orientation)
    survey.target_mental_health_diagnosis = _clean_target(target_mental_health_diagnosis)
    survey.target_physical_health_diagnosis = _clean_target(target_physical_health_diagnosis)
    survey.target_sport_type = _clean_target(target_sport_type)
    survey.target_sport_frequency = _clean_target(target_sport_frequency)
    survey.target_smoking = _clean_target(target_smoking); survey.target_cannabis_use = _clean_target(target_cannabis_use)
    survey.target_student_status = _clean_target(target_student_status)
    survey.target_year_in_school = _clean_target(target_year_in_school)
    survey.target_international_domestic = _clean_target(target_international_domestic)
    survey.target_experience_tags = target_experience_tags
    survey.target_participation_format = _clean_target(target_participation_format)
    survey.target_device = _clean_target(target_device)
    survey.target_income_level = _clean_target(target_income_level)
    survey.target_lifestyle_tags = target_lifestyle_tags
    survey.target_niche_requirements = _clean_target(form.get("target_niche_requirements"))
    _apply_survey_auto_filter_settings(survey, form)

    if cover_image and cover_image.filename:
        uploads_dir = Path("app/static/uploads"); uploads_dir.mkdir(exist_ok=True)
        unique_filename = f"{uuid.uuid4()}{Path(cover_image.filename).suffix}"
        file_path = uploads_dir / unique_filename
        with file_path.open("wb") as buffer: shutil.copyfileobj(cover_image.file, buffer)
        survey.image_url = f"/static/uploads/{unique_filename}"

    db.commit()

    # Published survey with additional responses that require payment
    is_no_pay = _clean_target(incentive_type) in ("raffle", "volunteer")
    if survey.status == "published" and additional_needed > 0 and not is_no_pay:
        total = round(reward_amount * additional_needed, 2)
        session = stripe.checkout.Session.create(
            payment_method_types=["card"],
            line_items=[{
                "price_data": {
                    "currency": "usd",
                    "product_data": {
                        "name": f"Additional responses: {survey.title}",
                        "description": f"{additional_needed} more responses × ${reward_amount:.2f} per person"
                    },
                    "unit_amount": int(round(total * 100))
                },
                "quantity": 1
            }],
            mode="payment",
            success_url=f"https://insightaco.org/payment/success?survey_id={survey.id}",
            cancel_url=f"https://insightaco.org/publisher",
            metadata={"survey_id": str(survey.id), "publisher_id": str(current_user.id)}
        )
        return RedirectResponse(session.url, status_code=303)

    return RedirectResponse("/publisher", status_code=303)
# ---------------------------
# Profile
# ---------------------------

@app.get("/profile", response_class=HTMLResponse)
def profile_get(
    request: Request,
    participant_app: Optional[str] = Cookie(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    referral_code = _ensure_user_referral_code(db, current_user, commit=True)
    invite_url = _referral_invite_url(
        referral_code,
        role="participant" if _should_use_participant_app(request, participant_app) else "researcher",
    )
    if _should_use_participant_app(request, participant_app):
        return templates.TemplateResponse("participant_profile.html", {
            "request": request,
            "current_user": current_user,
            "pending_earnings": getattr(current_user, 'pending_earnings', 0.0) or 0.0,
            "referral_code": referral_code,
            "invite_url": invite_url,
        })
    prev_url = request.headers.get("referer", "/publisher")
    return templates.TemplateResponse("profile.html", {
        "request": request,
        "user": current_user,
        "prev_url": prev_url,
        "referral_code": referral_code,
        "invite_url": invite_url,
    })


@app.get("/profile/edit", response_class=HTMLResponse)
def profile_edit_get(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    prev_url = request.headers.get("referer", "/profile")
    referral_code = _ensure_user_referral_code(db, current_user, commit=True)
    return templates.TemplateResponse("profile.html", {
        "request": request,
        "user": current_user,
        "prev_url": prev_url,
        "referral_code": referral_code,
        "invite_url": _referral_invite_url(referral_code, role="researcher"),
    })


@app.post("/profile")
async def profile_post(
    request: Request,
    participant_app: Optional[str] = Cookie(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    form = await request.form()
    if not _should_use_participant_app(request, participant_app):
        current_user.username = (form.get("username") or "").strip() or None
        if "first_name" in form:
            current_user.first_name = (form.get("first_name") or "").strip() or None
        if "last_name" in form:
            current_user.last_name = (form.get("last_name") or "").strip() or None
        current_user.email = (form.get("email") or "").strip()
        if "phone_number" in form:
            current_user.phone_number = (form.get("phone_number") or "").strip() or None
        current_user.field = _value_with_other(form.get("field"), form.get("field_other"))
        current_user.status = _value_with_other(form.get("status"), form.get("status_other"))
        current_user.profile_description = (form.get("profile_description") or "").strip() or None
        db.commit()
        return_to = form.get("return_to")
        if return_to and is_safe_internal_next(return_to):
            return RedirectResponse(return_to, status_code=303)
        return RedirectResponse("/publisher", status_code=303)

    language_list = _list_with_other(form.getlist("language"), form.get("language_other"))
    experience_list = _list_with_other(form.getlist("experience_tags"), form.get("experience_tags_other"))
    lifestyle_list = _list_with_other(form.getlist("lifestyle_tags"), form.get("lifestyle_tags_other"))

    current_user.username = form.get("username")
    if "first_name" in form:
        current_user.first_name = (form.get("first_name") or "").strip() or None
    if "last_name" in form:
        current_user.last_name = (form.get("last_name") or "").strip() or None
    current_user.email = form.get("email") or ""
    if "phone_number" in form:
        current_user.phone_number = (form.get("phone_number") or "").strip() or None

    birth_year = form.get("birth_year")
    birth_month = form.get("birth_month")
    derived_age_range = _age_range_from_birth_date(birth_year, birth_month)
    current_user.birth_year = birth_year or None
    current_user.birth_month = birth_month or None
    current_user.age_range = derived_age_range or form.get("age_range")
    current_user.profile_description = form.get("profile_description")

    current_user.education_level = form.get("education_level")
    current_user.field = _value_with_other(form.get("field"), form.get("field_other"))
    current_user.status = _value_with_other(form.get("status"), form.get("status_other"))

    current_user.current_country = form.get("current_country") or "United States"
    current_user.current_province = form.get("current_province")
    current_user.current_city = form.get("current_city")
    current_user.origin_country = form.get("origin_country") or ("United States" if form.get("origin_province") else None)
    current_user.origin_province = form.get("origin_province")
    current_user.origin_city = form.get("origin_city")
    current_user.state = form.get("state") or form.get("current_province")

    current_user.race = _value_with_other(form.get("race"), form.get("race_other"))
    current_user.ethnicity = form.get("ethnicity") or current_user.race
    current_user.mental_health_diagnosis = _value_with_other(
        form.get("mental_health_diagnosis"),
        form.get("mental_health_diagnosis_other"),
    )
    current_user.physical_health_diagnosis = _value_with_other(
        form.get("physical_health_diagnosis"),
        form.get("physical_health_diagnosis_other"),
    )
    current_user.sexual_orientation = form.get("sexual_orientation")
    current_user.sport_type = form.get("sport_type"); current_user.sport_frequency = form.get("sport_frequency")
    current_user.smoking = form.get("smoking"); current_user.cannabis_use = form.get("cannabis_use")
    current_user.language = ",".join(language_list) if language_list else None
    current_user.student_status = _value_with_other(form.get("student_status"), form.get("student_status_other"))
    current_user.year_in_school = _value_with_other(form.get("year_in_school"), form.get("year_in_school_other"))
    current_user.international_domestic = form.get("international_domestic")
    current_user.experience_tags = ",".join(experience_list) if experience_list else None
    current_user.income_level = form.get("income_level")
    current_user.lifestyle_tags = ",".join(lifestyle_list) if lifestyle_list else None
    current_user.participation_format = _value_with_other(form.get("participation_format"), form.get("participation_format_other"))
    current_user.device_type = _value_with_other(form.get("device_type"), form.get("device_type_other"))
    _record_user_event(
        db,
        request,
        "profile_update",
        user=current_user,
        metadata={"source": "Profile", "user_role": "participant"},
    )
    db.commit()
    return_to = form.get("return_to")
    if return_to and is_safe_internal_next(return_to):
        return RedirectResponse(return_to, status_code=303)
    if _should_use_participant_app(request, participant_app):
        return RedirectResponse("/profile", status_code=303)
    return RedirectResponse("/publisher", status_code=303)


# ---------------------------
# AI Fill
# ---------------------------

def _fallback_ai_fill(prompt: str, warning: str | None = None) -> dict:
    clean = re.sub(r"\s+", " ", (prompt or "").strip())
    title_words = re.findall(r"[A-Za-z0-9]+", clean)[:8]
    title = " ".join(title_words).strip() or "Research Study"
    if len(title) > 64:
        title = title[:61].rstrip() + "..."
    lower = clean.lower()
    category = "research"
    if any(word in lower for word in ["class", "student", "school", "academic", "college", "university"]):
        category = "academic"
    elif any(word in lower for word in ["food", "habit", "shopping", "consumer", "market", "brand"]):
        category = "market"
    elif any(word in lower for word in ["club", "organization", "community"]):
        category = "clubs"
    elif any(word in lower for word in ["lifestyle", "sleep", "wellness", "fitness", "health"]):
        category = "life"
    numbers = [int(n) for n in re.findall(r"\b\d+\b", lower)]
    target_responses = next((n for n in numbers if 1 <= n <= 10000), 50)
    estimated_time = 10
    time_match = re.search(r"(\d+)\s*(?:min|minute|minutes)", lower)
    if time_match:
        estimated_time = max(1, min(120, int(time_match.group(1))))
    per_person = 5.0
    amount_match = re.search(r"\$\s*(\d+(?:\.\d+)?)", lower)
    if not amount_match:
        amount_match = re.search(r"(?:pay|reward|incentive|compensation)\s*(?:is|of|:)?\s*(\d+(?:\.\d+)?)", lower)
    if amount_match:
        per_person = max(0.0, min(500.0, float(amount_match.group(1))))
    result = {
        "title": title,
        "description": clean or "Help us validate this research study by completing a short survey. Your responses will be used for product and recruitment testing.",
        "category": category,
        "estimated_time": estimated_time,
        "per_person_gross": per_person,
        "target_responses": target_responses,
    }
    if warning:
        result["warning"] = warning
    return result


def _parse_ai_fill_json(text: str) -> dict:
    import json as _json
    match = re.search(r'\{.*\}', text or "", re.DOTALL)
    if not match:
        raise ValueError("AI response parsing failed")
    data = _json.loads(match.group())
    return {
        "title": str(data.get("title") or "").strip()[:80] or "Research Study",
        "description": str(data.get("description") or "").strip() or "Help us validate this research study by completing a short survey.",
        "category": data.get("category") if data.get("category") in {"research", "life", "clubs", "market", "academic", "other"} else "research",
        "estimated_time": int(float(data.get("estimated_time") or 10)),
        "per_person_gross": float(data.get("per_person_gross") or 5.0),
        "target_responses": int(float(data.get("target_responses") or 50)),
    }


async def _anthropic_ai_fill(prompt: str) -> dict:
    api_key = (os.environ.get("ANTHROPIC_API_KEY") or "").strip()
    if not api_key:
        raise RuntimeError("ANTHROPIC_API_KEY is not configured")
    import anthropic
    configured_model = (os.environ.get("ANTHROPIC_AI_FILL_MODEL") or "").strip()
    allow_expensive_models = (os.environ.get("ALLOW_EXPENSIVE_ANTHROPIC_MODELS") or "").strip().lower() == "true"
    if configured_model and any(name in configured_model.lower() for name in ("opus", "fable")) and not allow_expensive_models:
        raise RuntimeError("ANTHROPIC_AI_FILL_MODEL points to an expensive model. Use Haiku or set ALLOW_EXPENSIVE_ANTHROPIC_MODELS=true explicitly.")
    model_candidates = [
        configured_model,
        "claude-haiku-4-5-20251001",
        "claude-haiku-4-5",
        "claude-3-5-haiku-20241022",
    ]
    model_candidates = [m for i, m in enumerate(model_candidates) if m and m not in model_candidates[:i]]
    client = anthropic.Anthropic(api_key=api_key)
    prompt_text = f"""You are helping a researcher fill out a survey publishing form.
Based on this description: "{prompt}"
Return ONLY a valid JSON object with these exact fields, no extra text:
{{"title": "clear survey title under 10 words", "description": "2-3 sentence description", "category": "one of: research, life, clubs, market, academic, other", "estimated_time": 5, "per_person_gross": 5.00, "target_responses": 100}}"""
    errors = []
    for model in model_candidates:
        try:
            message = client.messages.create(
                model=model, max_tokens=450,
                messages=[{"role": "user", "content": prompt_text}]
            )
            result = _parse_ai_fill_json(message.content[0].text)
            result["provider"] = f"anthropic:{model}"
            return result
        except Exception as exc:
            errors.append(f"{model}: {exc}")
    raise RuntimeError(" | ".join(errors))


async def _ai_fill_from_prompt(prompt: str) -> JSONResponse:
    if not prompt:
        raise HTTPException(400, "Prompt is required")
    try:
        return JSONResponse(await _anthropic_ai_fill(prompt))
    except Exception as exc:
        return JSONResponse(_fallback_ai_fill(prompt, f"Anthropic failed; used local fallback. {exc}"))


@app.post("/api/ai-fill")
async def ai_fill(request: Request, current_user: User = Depends(get_current_user)):
    try:
        body = await request.json()
        return await _ai_fill_from_prompt(body.get("prompt", ""))
    except Exception as e:
        import traceback; print(traceback.format_exc())
        raise HTTPException(500, str(e))


# ---------------------------
# AI Generate Questions
# ---------------------------

@app.post("/api/ai-generate-questions")
async def ai_generate_questions(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    try:
        body = await request.json()
        prompt = body.get("prompt", "")
        if not prompt:
            raise HTTPException(400, "Prompt is required")

        import anthropic, json as _json
        client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))
        message = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=2000,
            messages=[{
                "role": "user",
                "content": f"""You are helping a researcher design a survey.
Based on this research goal: "{prompt}"

Generate 6-8 survey questions. Return ONLY a valid JSON array, no extra text:
[
  {{
    "question_text": "What is your current year in school?",
    "question_type": "single",
    "options": ["Freshman", "Sophomore", "Junior", "Senior", "Graduate"],
    "is_required": true,
    "order_index": 1
  }},
  {{
    "question_text": "How many hours do you sleep on average per night?",
    "question_type": "scale",
    "options": null,
    "is_required": true,
    "order_index": 2
  }},
  {{
    "question_text": "Describe any sleep difficulties you experience.",
    "question_type": "text",
    "options": null,
    "is_required": false,
    "order_index": 3
  }}
]

Rules:
- Mix question types: single/multiple for closed, scale for ratings, text for open-ended
- Keep questions neutral, avoid leading language
- scale and text questions must have options: null
- options field is only for single, multiple, dropdown types
"""
            }]
        )

        text = message.content[0].text
        match = re.search(r'\[.*\]', text, re.DOTALL)
        if not match:
            raise HTTPException(500, "AI response parsing failed")

        questions = _json.loads(match.group())
        return JSONResponse({"questions": questions})

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        raise HTTPException(500, str(e))


# ---------------------------
# Feedback
# ---------------------------

@app.get("/feedback", response_class=HTMLResponse)
def feedback_page(request: Request, current_user: User = Depends(get_current_user)):
    return templates.TemplateResponse("feedback.html", {"request": request, "current_user": current_user})

@app.post("/feedback")
async def submit_feedback(request: Request, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    form = await request.form()
    db.add(Feedback(user_id=current_user.id, category=form.get("category", "general"), title=form.get("title", ""), content=form.get("content", ""), status="pending"))
    db.commit()
    return RedirectResponse("/feedback?success=1", status_code=303)


# ---------------------------
# Support chat
# ---------------------------

SUPPORT_AVAILABLE_HOURS = "10am-5pm"

def _support_thread_payload(thread: SupportThread, db: Session):
    user = db.query(User).filter(User.id == thread.user_id).first()
    last_message = db.query(SupportMessage).filter(
        SupportMessage.thread_id == thread.id
    ).order_by(SupportMessage.created_at.desc()).first()
    unread_user_messages = db.query(SupportMessage).filter(
        SupportMessage.thread_id == thread.id,
        SupportMessage.sender_type == "user",
        SupportMessage.read_at.is_(None),
    ).count()
    return {
        "id": thread.id,
        "status": thread.status,
        "user_id": thread.user_id,
        "user_email": user.email if user else "unknown",
        "user_name": user.username if user else None,
        "last_message": last_message.body if last_message else "",
        "last_message_at": thread.last_message_at.isoformat() if thread.last_message_at else None,
        "created_at": thread.created_at.isoformat() if thread.created_at else None,
        "unread_user_messages": unread_user_messages,
    }

def _support_message_payload(message: SupportMessage):
    return {
        "id": message.id,
        "thread_id": message.thread_id,
        "sender_type": message.sender_type,
        "sender_id": message.sender_id,
        "body": message.body,
        "created_at": message.created_at.isoformat() if message.created_at else None,
    }

def _get_or_create_support_thread(db: Session, user_id: int):
    thread = db.query(SupportThread).filter(
        SupportThread.user_id == user_id,
        SupportThread.status == "open",
    ).order_by(SupportThread.updated_at.desc()).first()
    if thread:
        return thread
    thread = SupportThread(user_id=user_id, status="open", last_message_at=datetime.utcnow())
    db.add(thread)
    db.flush()
    db.add(SupportMessage(
        thread_id=thread.id,
        sender_type="system",
        body=f"Support hours are {SUPPORT_AVAILABLE_HOURS}. Leave a message and our team will reply here.",
    ))
    db.commit()
    db.refresh(thread)
    return thread

@app.get("/api/support/availability")
def support_availability():
    return JSONResponse({
        "available_hours": SUPPORT_AVAILABLE_HOURS,
        "label": f"Human support available {SUPPORT_AVAILABLE_HOURS}",
    })

@app.get("/api/support/thread")
def get_support_thread(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    thread = _get_or_create_support_thread(db, current_user.id)
    messages = db.query(SupportMessage).filter(
        SupportMessage.thread_id == thread.id
    ).order_by(SupportMessage.created_at.asc()).all()
    return JSONResponse({
        "thread": _support_thread_payload(thread, db),
        "messages": [_support_message_payload(m) for m in messages],
        "available_hours": SUPPORT_AVAILABLE_HOURS,
    })

@app.get("/api/support/messages")
def get_support_messages(
    thread_id: int = Query(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    thread = db.query(SupportThread).filter(
        SupportThread.id == thread_id,
        SupportThread.user_id == current_user.id,
    ).first()
    if not thread:
        raise HTTPException(404, "Support thread not found")
    messages = db.query(SupportMessage).filter(
        SupportMessage.thread_id == thread.id
    ).order_by(SupportMessage.created_at.asc()).all()
    return JSONResponse([_support_message_payload(m) for m in messages])

@app.post("/api/support/messages")
async def send_support_message(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    body = await request.json()
    text_body = (body.get("body") or "").strip()
    if not text_body:
        raise HTTPException(400, "Message cannot be empty")
    if len(text_body) > 2000:
        raise HTTPException(400, "Message is too long")

    thread_id = body.get("thread_id")
    thread = None
    if thread_id:
        thread = db.query(SupportThread).filter(
            SupportThread.id == int(thread_id),
            SupportThread.user_id == current_user.id,
        ).first()
    if not thread:
        thread = _get_or_create_support_thread(db, current_user.id)

    msg = SupportMessage(
        thread_id=thread.id,
        sender_type="user",
        sender_id=current_user.id,
        body=text_body,
    )
    thread.status = "open"
    thread.last_message_at = datetime.utcnow()
    db.add(msg)
    db.commit()
    db.refresh(msg)
    send_support_alert_email(current_user, thread, msg)
    return JSONResponse({
        "thread": _support_thread_payload(thread, db),
        "message": _support_message_payload(msg),
    })


# ---------------------------
# Admin
# ---------------------------

def _admin_key_matches(value: str | None) -> bool:
    expected = (os.environ.get("ADMIN_KEY") or "").strip()
    provided = (value or "").strip()
    return bool(expected and provided) and hmac.compare_digest(provided, expected)


def _get_admin_publisher_user(db: Session) -> User:
    email = (os.environ.get("ADMIN_PUBLISHER_EMAIL") or "vfsa@bu.edu").strip().lower()
    user = db.query(User).filter(func.lower(User.email) == email).first()
    if user:
        _ensure_user_referral_code(db, user, commit=True)
        return user
    user = User(
        email=email,
        password="admin-managed-publisher",
        username="Insighta Admin",
        status="Researcher",
        referral_code=_generate_referral_code(db),
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


@app.get("/admin", response_class=HTMLResponse)
def admin_page(request: Request):
    return templates.TemplateResponse("admin.html", {"request": request})

@app.get("/admin/discovery", response_class=HTMLResponse)
def admin_discovery_page(request: Request):
    return RedirectResponse("/admin#discovery", status_code=303)

@app.post("/admin/verify")
async def admin_verify(request: Request):
    body = await request.json()
    if not _admin_key_matches(body.get("admin_key")):
        raise HTTPException(403, "Unauthorized")
    return JSONResponse({"success": True})


@app.post("/admin/users/reward-reminder-email")
async def admin_reward_reminder_email(request: Request, db: Session = Depends(get_db)):
    body = await request.json()
    if not _admin_key_matches(body.get("admin_key")):
        raise HTTPException(403, "Unauthorized")
    dry_run = bool(body.get("dry_run", True))
    users = db.query(User).filter(User.email.isnot(None)).order_by(User.created_at.desc()).all()
    recipients: list[User] = []
    seen_emails = set()
    for user in users:
        normalized = _normalize_email(user.email or "")
        if not normalized or "@" not in normalized or normalized in seen_emails:
            continue
        seen_emails.add(normalized)
        recipients.append(user)
    if dry_run:
        return JSONResponse({
            "dry_run": True,
            "recipient_count": len(recipients),
        })

    sent_count = 0
    failures = []
    for user in recipients:
        sent, error = _send_reward_setup_reminder_email(user)
        if sent:
            sent_count += 1
        else:
            failures.append({
                "email": user.email,
                "error": error or "Unknown email error",
            })
    return JSONResponse({
        "dry_run": False,
        "recipient_count": len(recipients),
        "sent_count": sent_count,
        "failed_count": len(failures),
        "failures": failures[:20],
    })


@app.post("/admin/api/ai-fill")
async def admin_ai_fill(request: Request):
    try:
        body = await request.json()
        if not _admin_key_matches(body.get("admin_key")):
            raise HTTPException(403, "Unauthorized")
        return await _ai_fill_from_prompt(body.get("prompt", ""))
    except HTTPException:
        raise
    except Exception as e:
        import traceback; print(traceback.format_exc())
        raise HTTPException(500, str(e))


@app.get("/admin/publish", response_class=HTMLResponse)
def admin_publish_page(request: Request, admin_key: str = Query(None), db: Session = Depends(get_db)):
    if not _admin_key_matches(admin_key):
        raise HTTPException(403, "Unauthorized")
    admin_user = _get_admin_publisher_user(db)
    return templates.TemplateResponse("publish_external.html", {
        "request": request,
        "stripe_publishable_key": STRIPE_PUBLISHABLE_KEY,
        "builtin": False,
        "existing_survey_id": 0,
        "prefill_title": "",
        "prefill_desc": "",
        "existing_survey": None,
        "current_user": admin_user,
        "is_admin_publish": True,
        "admin_key": admin_key,
        "admin_created": request.query_params.get("created") == "1",
    })


@app.post("/admin/publish")
async def admin_publish_survey(request: Request, db: Session = Depends(get_db)):
    form = await request.form()
    if not _admin_key_matches(form.get("admin_key")):
        raise HTTPException(403, "Unauthorized")
    admin_user = _get_admin_publisher_user(db)
    display_amount_raw = (form.get("admin_display_reward_amount") or "").strip()
    display_amount = float(display_amount_raw) if display_amount_raw else None
    return await publish_survey(
        request=request,
        title=form.get("title"),
        description=form.get("description"),
        form_url=form.get("form_url"),
        task_type=_normalize_task_type(form.get("task_type")),
        category=form.get("category"),
        estimated_time=int(form.get("estimated_time") or 0),
        per_person_gross=float(form.get("per_person_gross") or 0),
        total_budget=float(form.get("total_budget") or 0),
        target_responses=int(form.get("target_responses") or 1),
        urgency_level=form.get("urgency_level"),
        incentive_type=form.get("incentive_type"),
        target_age_range=form.get("target_age_range"),
        target_field=form.get("target_field"),
        target_status=form.get("target_status"),
        target_state=form.get("target_state"),
        target_language=form.get("target_language"),
        target_ethnicity=form.get("target_ethnicity"),
        target_sexual_orientation=form.get("target_sexual_orientation"),
        target_mental_health_diagnosis=form.get("target_mental_health_diagnosis"),
        target_physical_health_diagnosis=form.get("target_physical_health_diagnosis"),
        target_sport_type=form.get("target_sport_type"),
        target_sport_frequency=form.get("target_sport_frequency"),
        target_smoking=form.get("target_smoking"),
        target_cannabis_use=form.get("target_cannabis_use"),
        target_student_status=form.get("target_student_status"),
        target_year_in_school=form.get("target_year_in_school"),
        target_international_domestic=form.get("target_international_domestic"),
        target_participation_format=form.get("target_participation_format"),
        target_device=form.get("target_device"),
        target_income_level=form.get("target_income_level"),
        raffle_prize_type=form.get("raffle_prize_type"),
        cover_image=form.get("cover_image"),
        admin_display_reward_amount=display_amount,
        admin_publish=True,
        admin_redirect_key=form.get("admin_key"),
        current_user=admin_user,
        db=db,
    )


@app.get("/admin/listings/{survey_id}/edit", response_class=HTMLResponse)
def admin_edit_listing_page(
    request: Request,
    survey_id: int,
    admin_key: str = Query(None),
    db: Session = Depends(get_db),
):
    if not _admin_key_matches(admin_key):
        raise HTTPException(403, "Unauthorized")
    survey = db.query(Survey).filter(Survey.id == survey_id).first()
    if not survey:
        raise HTTPException(404, "Survey not found")
    survey.current_responses = db.query(Response).filter(
        Response.survey_id == survey_id,
        Response.status == "completed",
    ).count()
    admin_user = _get_admin_publisher_user(db)
    return templates.TemplateResponse("edit_publish.html", {
        "request": request,
        "survey": survey,
        "current_user": admin_user,
        "is_admin_edit": True,
        "admin_key": admin_key,
    })


@app.post("/admin/listings/{survey_id}/edit")
async def admin_edit_listing_post(survey_id: int, request: Request, db: Session = Depends(get_db)):
    form = await request.form()
    if not _admin_key_matches(form.get("admin_key")):
        raise HTTPException(403, "Unauthorized")

    survey = db.query(Survey).filter(Survey.id == survey_id).first()
    if not survey:
        raise HTTPException(404, "Survey not found")

    current_responses = db.query(Response).filter(
        Response.survey_id == survey_id,
        Response.status == "completed",
    ).count()
    additional_needed = max(0, _parse_optional_int(form.get("additional_needed")) or 0)
    incentive_clean = _clean_target(form.get("incentive_type")) or "cash"
    display_amount_raw = (form.get("admin_display_reward_amount") or "").strip()

    survey.title = (form.get("title") or survey.title or "").strip()
    survey.description = (form.get("description") or "").strip()
    survey.form_url = (form.get("form_url") or "").strip()
    survey.task_type = _clean_target(form.get("task_type")) or "survey"
    survey.category = _clean_target(form.get("category")) or survey.category
    survey.estimated_time = _parse_optional_int(form.get("estimated_time")) or survey.estimated_time
    if survey.status != "published":
        try:
            survey.reward_amount = float(form.get("reward_amount") or 0)
        except (TypeError, ValueError):
            survey.reward_amount = 0.0
    survey.target_responses = current_responses + additional_needed
    survey.urgency_level = _clean_target(form.get("urgency_level"))
    survey.incentive_type = incentive_clean
    survey.raffle_prize_type = _clean_target(form.get("raffle_prize_type")) if incentive_clean == "raffle" else None
    survey.target_age_range = _clean_target(form.get("target_age_range"))
    survey.target_education_min = _parse_optional_int(form.get("target_education_min"))
    survey.target_education_max = _parse_optional_int(form.get("target_education_max"))
    survey.target_field = _clean_target(form.get("target_field"))
    survey.target_status = _clean_target(form.get("target_status"))
    survey.target_state = _clean_target(form.get("target_state"))
    survey.target_language = _clean_target(form.get("target_language"))
    survey.target_ethnicity = _clean_target(form.get("target_ethnicity"))
    survey.target_sexual_orientation = _clean_target(form.get("target_sexual_orientation"))
    survey.target_mental_health_diagnosis = _clean_target(form.get("target_mental_health_diagnosis"))
    survey.target_physical_health_diagnosis = _clean_target(form.get("target_physical_health_diagnosis"))
    survey.target_sport_type = _clean_target(form.get("target_sport_type"))
    survey.target_sport_frequency = _clean_target(form.get("target_sport_frequency"))
    survey.target_smoking = _clean_target(form.get("target_smoking"))
    survey.target_cannabis_use = _clean_target(form.get("target_cannabis_use"))
    survey.target_student_status = _clean_target(form.get("target_student_status"))
    survey.target_year_in_school = _clean_target(form.get("target_year_in_school"))
    survey.target_international_domestic = _clean_target(form.get("target_international_domestic"))
    survey.target_experience_tags = ",".join(form.getlist("target_experience_tags")) or None
    survey.target_participation_format = _clean_target(form.get("target_participation_format"))
    survey.target_device = _clean_target(form.get("target_device"))
    survey.target_income_level = _clean_target(form.get("target_income_level"))
    survey.target_lifestyle_tags = ",".join(form.getlist("target_lifestyle_tags")) or None
    survey.target_niche_requirements = _clean_target(form.get("target_niche_requirements"))
    if display_amount_raw:
        try:
            survey.admin_display_reward_amount = max(0.0, float(display_amount_raw))
        except ValueError:
            raise HTTPException(400, "Invalid display reward amount")
    else:
        survey.admin_display_reward_amount = None
    _apply_survey_auto_filter_settings(survey, form)

    cover_image = form.get("cover_image")
    if cover_image is not None and getattr(cover_image, "filename", None):
        uploads_dir = Path("app/static/uploads")
        uploads_dir.mkdir(exist_ok=True)
        unique_filename = f"{uuid.uuid4()}{Path(cover_image.filename).suffix}"
        file_path = uploads_dir / unique_filename
        with file_path.open("wb") as buffer:
            shutil.copyfileobj(cover_image.file, buffer)
        survey.image_url = f"/static/uploads/{unique_filename}"

    db.commit()
    return RedirectResponse("/admin#listings", status_code=303)


@app.post("/admin/feedback/{feedback_id}/credit")
async def grant_credit(feedback_id: int, request: Request, db: Session = Depends(get_db)):
    body = await request.json()
    if not _admin_key_matches(body.get("admin_key")): raise HTTPException(403, "Unauthorized")
    amount = float(body.get("amount", 0))
    if amount <= 0: raise HTTPException(400, "Amount must be > 0")
    feedback = db.query(Feedback).filter(Feedback.id == feedback_id).first()
    if not feedback: raise HTTPException(404, "Feedback not found")
    user = db.query(User).filter(User.id == feedback.user_id).first()
    if not user: raise HTTPException(404, "User not found")
    user.pending_earnings = (getattr(user, 'pending_earnings', 0.0) or 0.0) + amount
    feedback.status = "credited"; feedback.credit_amount = amount; feedback.reviewed_at = datetime.utcnow()
    db.commit()
    return JSONResponse({"success": True, "user_email": user.email, "amount": amount, "new_balance": user.pending_earnings})

@app.post("/admin/feedback/{feedback_id}/reject")
async def reject_feedback(feedback_id: int, request: Request, db: Session = Depends(get_db)):
    body = await request.json()
    if not _admin_key_matches(body.get("admin_key")): raise HTTPException(403, "Unauthorized")
    feedback = db.query(Feedback).filter(Feedback.id == feedback_id).first()
    if not feedback: raise HTTPException(404, "Feedback not found")
    feedback.status = "rejected"; feedback.reviewed_at = datetime.utcnow(); db.commit()
    return JSONResponse({"success": True})

@app.get("/admin/listings")
async def admin_list_listings(
    request: Request,
    admin_key: str = Query(None),
    q: str = Query(""),
    db: Session = Depends(get_db),
):
    if not _admin_key_matches(admin_key): raise HTTPException(403, "Unauthorized")
    query = db.query(Survey).order_by(Survey.created_at.desc())
    search = (q or "").strip()
    if search:
        query = query.filter(Survey.title.ilike(f"%{search}%"))
    surveys = query.limit(200).all()
    slugs_changed = False
    for s in surveys:
        if not getattr(s, "share_slug", None):
            _ensure_survey_share_slug(db, s)
            slugs_changed = True
    if slugs_changed:
        db.commit()
    publisher_ids = {s.publisher_id for s in surveys if s.publisher_id}
    publishers = {}
    if publisher_ids:
        publishers = {u.id: u for u in db.query(User).filter(User.id.in_(publisher_ids)).all()}
    result = []
    for s in surveys:
        completed_count = db.query(Response).filter(
            Response.survey_id == s.id,
            Response.status == "completed",
        ).count()
        publisher = publishers.get(s.publisher_id)
        result.append({
            "id": s.id,
            "title": s.title,
            "description": s.description,
            "status": s.status,
            "payment_status": s.payment_status,
            "task_type": _normalize_task_type(getattr(s, "task_type", None)),
            "task_type_label": _task_type_label(getattr(s, "task_type", None)),
            "category": s.category,
            "reward_amount": s.reward_amount,
            "admin_display_reward_amount": getattr(s, "admin_display_reward_amount", None),
            "target_responses": s.target_responses,
            "completed_count": completed_count,
            "publisher_email": publisher.email if publisher else "unknown",
            "created_at": s.created_at.isoformat() if s.created_at else None,
            "published_at": s.published_at.isoformat() if s.published_at else None,
            "share_url": _survey_share_url(request, db, s),
        })
    return JSONResponse(result)


@app.get("/admin/analytics")
async def admin_analytics(
    request: Request,
    admin_key: str = Query(None),
    days: int = Query(30),
    db: Session = Depends(get_db),
):
    if not _admin_key_matches(admin_key):
        raise HTTPException(403, "Unauthorized")
    days = max(1, min(days, 180))
    since = datetime.utcnow() - timedelta(days=days)

    events = db.query(UserEvent).filter(UserEvent.created_at >= since).order_by(UserEvent.created_at.desc()).limit(20000).all()
    event_counts: dict[str, int] = {}
    unique_people = set()
    for event in events:
        event_counts[event.event_name] = event_counts.get(event.event_name, 0) + 1
        if event.user_id:
            unique_people.add(f"u:{event.user_id}")
        elif event.anonymous_id:
            unique_people.add(f"a:{event.anonymous_id}")

    def in_analytics_window(value) -> bool:
        if not value:
            return False
        if getattr(value, "tzinfo", None):
            value = value.replace(tzinfo=None)
        return value >= since

    def response_counts_in_window(response: Response) -> tuple[bool, bool]:
        started_in_window = in_analytics_window(response.started_at)
        # Older rows may not have timestamps even though their status is useful.
        if not response.started_at:
            started_in_window = response.status in {"started", "completed"}
        completed_in_window = in_analytics_window(response.completed_at)
        if not response.completed_at and response.status == "completed":
            completed_in_window = started_in_window
        return started_in_window, completed_in_window

    all_response_rows = db.query(Response).all()
    response_started_total = 0
    completed_total = 0
    for response in all_response_rows:
        started_in_window, completed_in_window = response_counts_in_window(response)
        if started_in_window:
            response_started_total += 1
        if completed_in_window:
            completed_total += 1
    starts_total = max(event_counts.get("study_start", 0) + event_counts.get("survey_started", 0), response_started_total)
    start_to_complete = round((completed_total / starts_total) * 100, 1) if starts_total else 0

    surveys = db.query(Survey).order_by(Survey.created_at.desc()).limit(200).all()
    survey_ids = [s.id for s in surveys]
    response_rows = []
    if survey_ids:
        response_rows = db.query(Response).filter(Response.survey_id.in_(survey_ids)).all()

    event_by_survey: dict[str, dict[str, Any]] = {}
    for event in events:
        if event.target_type != "survey" or not event.target_id:
            continue
        bucket = event_by_survey.setdefault(event.target_id, {
            "views": 0,
            "unique_viewers": set(),
            "starts": 0,
            "unique_starters": set(),
            "completed": 0,
        })
        actor = f"u:{event.user_id}" if event.user_id else f"a:{event.anonymous_id or event.client_ip or event.id}"
        if event.event_name in VIEW_EVENT_NAMES:
            bucket["views"] += 1
            bucket["unique_viewers"].add(actor)
        elif event.event_name in START_EVENT_NAMES:
            bucket["starts"] += 1
            bucket["unique_starters"].add(actor)
        elif event.event_name in COMPLETE_EVENT_NAMES:
            bucket["completed"] += 1

    responses_by_survey: dict[int, dict[str, int]] = {}

    for response in response_rows:
        bucket = responses_by_survey.setdefault(response.survey_id, {"responses": 0, "completed": 0})
        started_in_window, completed_in_window = response_counts_in_window(response)
        if started_in_window:
            bucket["responses"] += 1
        if completed_in_window:
            bucket["completed"] += 1

    listing_funnel = []
    for survey in surveys:
        event_bucket = event_by_survey.get(str(survey.id), {})
        response_bucket = responses_by_survey.get(survey.id, {"responses": 0, "completed": 0})
        tracked_views = int(event_bucket.get("views", 0) or 0)
        starts = max(int(event_bucket.get("starts", 0) or 0), int(response_bucket.get("responses", 0) or 0))
        completed = max(int(event_bucket.get("completed", 0) or 0), int(response_bucket.get("completed", 0) or 0))
        views = max(tracked_views, starts)
        unique_viewers = max(len(event_bucket.get("unique_viewers", set())), len(event_bucket.get("unique_starters", set())))
        listing_funnel.append({
            "id": survey.id,
            "title": survey.title,
            "status": survey.status,
            "views": views,
            "tracked_views": tracked_views,
            "unique_viewers": unique_viewers,
            "starts": starts,
            "unique_starters": len(event_bucket.get("unique_starters", set())),
            "completed": completed,
            "start_rate": round((starts / views) * 100, 1) if views else None,
            "completion_rate": round((completed / starts) * 100, 1) if starts else None,
            "created_at": survey.created_at.isoformat() if survey.created_at else None,
        })
    listing_funnel.sort(key=lambda item: (item["views"], item["starts"], item["completed"], item["id"]), reverse=True)

    recent = events[:80]
    user_ids = {event.user_id for event in recent if event.user_id}
    users = {u.id: u for u in db.query(User).filter(User.id.in_(user_ids)).all()} if user_ids else {}
    recent_survey_ids = []
    for event in recent:
        if event.target_type == "survey" and event.target_id and str(event.target_id).isdigit():
            recent_survey_ids.append(int(event.target_id))
    survey_map = {s.id: s for s in db.query(Survey).filter(Survey.id.in_(set(recent_survey_ids))).all()} if recent_survey_ids else {}
    recent_activity = []
    for event in recent:
        user = users.get(event.user_id)
        survey = survey_map.get(int(event.target_id)) if event.target_type == "survey" and event.target_id and str(event.target_id).isdigit() else None
        guest_suffix = (event.anonymous_id or "")[-4:].upper()
        actor_label = user.email if user else (f"Guest visitor #{guest_suffix}" if guest_suffix else "Guest visitor")
        recent_activity.append({
            "id": event.id,
            "event_name": event.event_name,
            "user": actor_label,
            "actor_type": "user" if user else "guest",
            "anonymous_id": event.anonymous_id,
            "target_type": event.target_type,
            "target_id": event.target_id,
            "target_label": survey.title if survey else None,
            "page_path": event.page_path,
            "metadata": event.metadata_json or {},
            "created_at": event.created_at.isoformat() if event.created_at else None,
        })

    return JSONResponse({
        "overview": {
            "days": days,
            "total_events": len(events),
            "unique_people": len(unique_people),
            "listing_views": (
                event_counts.get("listing_view", 0)
                + event_counts.get("listing_viewed", 0)
                + event_counts.get("study_impression", 0)
                + event_counts.get("study_card_viewed", 0)
            ),
            "start_clicks": starts_total,
            "completed_surveys": completed_total,
            "start_to_complete_rate": start_to_complete,
        },
        "event_counts": event_counts,
        "listing_funnel": listing_funnel[:100],
        "recent_activity": recent_activity,
    })


def _admin_event_actor(event: UserEvent, users: dict[int, User], anonymous_user_ids: Optional[dict[str, int]] = None) -> dict[str, str]:
    inferred_user_id = None
    if not event.user_id and event.anonymous_id and anonymous_user_ids:
        inferred_user_id = anonymous_user_ids.get(event.anonymous_id)
    user = users.get(event.user_id or inferred_user_id)
    if user:
        label = user.email or user.username or f"User #{user.id}"
        return {"label": label, "type": "user", "email": user.email or "", "username": user.username or ""}
    guest_suffix = (event.anonymous_id or event.client_ip or str(event.id or ""))[-4:].upper()
    label = f"Guest visitor #{guest_suffix}" if guest_suffix else "Guest visitor"
    return {"label": label, "type": "guest", "email": "", "username": ""}


def _admin_event_payload(event: UserEvent, users: dict[int, User], survey_map: dict[int, Survey], anonymous_user_ids: Optional[dict[str, int]] = None) -> dict[str, Any]:
    survey = survey_map.get(int(event.target_id)) if event.target_type == "survey" and event.target_id and str(event.target_id).isdigit() else None
    inferred_user_id = anonymous_user_ids.get(event.anonymous_id) if anonymous_user_ids and event.anonymous_id else None
    actor = _admin_event_actor(event, users, anonymous_user_ids)
    metadata = event.metadata_json or {}
    user = users.get(event.user_id or inferred_user_id)
    return {
        "id": event.id,
        "event_name": event.event_name,
        "user": actor["label"],
        "role": _admin_event_role(event, user, metadata),
        "actor_type": actor["type"],
        "user_email": actor["email"],
        "username": actor["username"],
        "user_id": event.user_id or inferred_user_id,
        "anonymous_id": event.anonymous_id,
        "target_type": event.target_type,
        "target_id": event.target_id,
        "target_label": survey.title if survey else (metadata.get("study_title") or None),
        "page_path": event.page_path,
        "source": _admin_event_source(metadata),
        "details": _admin_event_details(metadata),
        "metadata": metadata,
        "created_at": event.created_at.isoformat() if event.created_at else None,
    }


def _anonymous_user_map(db: Session, anonymous_ids: set[str]) -> dict[str, int]:
    clean_ids = {value for value in anonymous_ids if value}
    if not clean_ids:
        return {}
    rows = db.query(UserEvent.anonymous_id, UserEvent.user_id).filter(
        UserEvent.anonymous_id.in_(clean_ids),
        UserEvent.user_id.isnot(None),
    ).order_by(UserEvent.created_at.desc()).limit(5000).all()
    result: dict[str, int] = {}
    for anonymous_id, user_id in rows:
        if anonymous_id and user_id and anonymous_id not in result:
            result[anonymous_id] = user_id
    return result


@app.get("/admin/activity")
async def admin_activity(
    request: Request,
    admin_key: str = Query(None),
    survey_id: Optional[int] = Query(None),
    event_name: str = Query(""),
    q: str = Query(""),
    limit: int = Query(100),
    offset: int = Query(0),
    db: Session = Depends(get_db),
):
    if not _admin_key_matches(admin_key):
        raise HTTPException(403, "Unauthorized")

    limit = max(20, min(limit, 500))
    offset = max(0, offset)
    event_name = (event_name or "").strip()
    q = (q or "").strip()

    surveys = db.query(Survey).order_by(Survey.created_at.desc()).limit(1000).all()
    survey_options = [{
        "id": survey.id,
        "title": survey.title,
        "status": survey.status,
        "created_at": survey.created_at.isoformat() if survey.created_at else None,
    } for survey in surveys]

    query = db.query(UserEvent)
    if survey_id:
        query = query.filter(UserEvent.target_type == "survey", UserEvent.target_id == str(survey_id))
    if event_name:
        query = query.filter(UserEvent.event_name == event_name)
    if q:
        like = f"%{q}%"
        matching_users = db.query(User.id).filter(or_(User.email.ilike(like), User.username.ilike(like))).limit(500).all()
        matching_user_ids = [row[0] for row in matching_users]
        conditions = [
            UserEvent.anonymous_id.ilike(like),
            UserEvent.client_ip.ilike(like),
            UserEvent.page_path.ilike(like),
        ]
        if matching_user_ids:
            conditions.append(UserEvent.user_id.in_(matching_user_ids))
        query = query.filter(or_(*conditions))

    total = query.count()
    events = query.order_by(UserEvent.created_at.desc()).offset(offset).limit(limit).all()

    anonymous_user_ids = _anonymous_user_map(db, {event.anonymous_id for event in events if event.anonymous_id})
    user_ids = {event.user_id for event in events if event.user_id}
    user_ids.update(anonymous_user_ids.values())
    survey_ids = {
        int(event.target_id)
        for event in events
        if event.target_type == "survey" and event.target_id and str(event.target_id).isdigit()
    }
    users = {u.id: u for u in db.query(User).filter(User.id.in_(user_ids)).all()} if user_ids else {}
    survey_map = {s.id: s for s in db.query(Survey).filter(Survey.id.in_(survey_ids)).all()} if survey_ids else {}
    activity = [_admin_event_payload(event, users, survey_map, anonymous_user_ids) for event in events]

    people = []
    if survey_id:
        survey_events = db.query(UserEvent).filter(
            UserEvent.target_type == "survey",
            UserEvent.target_id == str(survey_id),
        ).order_by(UserEvent.created_at.asc()).limit(20000).all()
        survey_anonymous_user_ids = _anonymous_user_map(db, {event.anonymous_id for event in survey_events if event.anonymous_id})
        response_rows = db.query(Response).filter(Response.survey_id == survey_id).all()
        survey_user_ids = {event.user_id for event in survey_events if event.user_id}
        survey_user_ids.update(survey_anonymous_user_ids.values())
        survey_user_ids.update({response.participant_id for response in response_rows if response.participant_id})
        survey_users = {u.id: u for u in db.query(User).filter(User.id.in_(survey_user_ids)).all()} if survey_user_ids else {}
        grouped: dict[str, dict[str, Any]] = {}
        view_events = VIEW_EVENT_NAMES
        start_events = START_EVENT_NAMES
        complete_events = COMPLETE_EVENT_NAMES
        for event in survey_events:
            inferred_user_id = survey_anonymous_user_ids.get(event.anonymous_id) if event.anonymous_id else None
            actor_key = f"u:{event.user_id or inferred_user_id}" if (event.user_id or inferred_user_id) else f"a:{event.anonymous_id or event.client_ip or event.id}"
            actor = _admin_event_actor(event, survey_users, survey_anonymous_user_ids)
            bucket = grouped.setdefault(actor_key, {
                "user": actor["label"],
                "actor_type": actor["type"],
                "user_email": actor["email"],
                "username": actor["username"],
                "user_id": event.user_id or inferred_user_id,
                "anonymous_id": event.anonymous_id,
                "first_opened_at": None,
                "first_started_at": None,
                "completed_at": None,
                "last_event_at": None,
                "event_count": 0,
                "last_event_name": None,
            })
            bucket["event_count"] += 1
            bucket["last_event_at"] = event.created_at.isoformat() if event.created_at else None
            bucket["last_event_name"] = event.event_name
            if event.event_name in view_events and not bucket["first_opened_at"]:
                bucket["first_opened_at"] = event.created_at.isoformat() if event.created_at else None
            if event.event_name in start_events and not bucket["first_started_at"]:
                bucket["first_started_at"] = event.created_at.isoformat() if event.created_at else None
            if event.event_name in complete_events and not bucket["completed_at"]:
                bucket["completed_at"] = event.created_at.isoformat() if event.created_at else None

        for response in response_rows:
            user = survey_users.get(response.participant_id)
            actor_key = f"u:{response.participant_id}" if response.participant_id else f"response:{response.id}"
            bucket = grouped.setdefault(actor_key, {
                "user": (user.email or user.username or f"User #{user.id}") if user else f"User #{response.participant_id}",
                "actor_type": "user",
                "user_email": user.email if user else "",
                "username": user.username if user else "",
                "user_id": response.participant_id,
                "anonymous_id": None,
                "first_opened_at": None,
                "first_started_at": None,
                "completed_at": None,
                "last_event_at": None,
                "event_count": 0,
                "last_event_name": None,
            })
            started_at = response.started_at.isoformat() if response.started_at else None
            completed_at = response.completed_at.isoformat() if response.completed_at else None
            if started_at and not bucket["first_started_at"]:
                bucket["first_started_at"] = started_at
            if completed_at and not bucket["completed_at"]:
                bucket["completed_at"] = completed_at
            latest_at = completed_at or started_at
            if latest_at and (not bucket["last_event_at"] or latest_at > bucket["last_event_at"]):
                bucket["last_event_at"] = latest_at
                bucket["last_event_name"] = "study_complete" if completed_at else "study_start"
            if response.id and bucket["event_count"] == 0:
                bucket["event_count"] = 1
        people = sorted(
            grouped.values(),
            key=lambda item: item.get("last_event_at") or "",
            reverse=True,
        )

    return JSONResponse({
        "surveys": survey_options,
        "events": activity,
        "people": people,
        "total": total,
        "limit": limit,
        "offset": offset,
    })


@app.post("/admin/listings/{survey_id}/delete")
async def admin_delete_listing(survey_id: int, request: Request, db: Session = Depends(get_db)):
    body = await request.json()
    if not _admin_key_matches(body.get("admin_key")): raise HTTPException(403, "Unauthorized")
    survey = db.query(Survey).filter(Survey.id == survey_id).first()
    if not survey:
        raise HTTPException(404, "Survey not found")
    deleted = {"id": survey.id, "title": survey.title}
    _delete_survey_tree(db, survey)
    return JSONResponse({"success": True, "deleted": deleted})


@app.post("/admin/listings/delete-by-title")
async def admin_delete_listings_by_title(request: Request, db: Session = Depends(get_db)):
    body = await request.json()
    if not _admin_key_matches(body.get("admin_key")): raise HTTPException(403, "Unauthorized")
    titles = [str(t).strip() for t in (body.get("titles") or []) if str(t).strip()]
    if not titles:
        raise HTTPException(400, "No titles provided")
    surveys = db.query(Survey).filter(Survey.title.in_(titles)).all()
    deleted = []
    for survey in surveys:
        deleted.append({"id": survey.id, "title": survey.title})
        _delete_survey_tree(db, survey)
    return JSONResponse({"success": True, "deleted": deleted})


@app.post("/admin/listings/reset-research-participation-demo")
async def admin_reset_research_participation_demo(request: Request, db: Session = Depends(get_db)):
    body = await request.json()
    if not _admin_key_matches(body.get("admin_key")):
        raise HTTPException(403, "Unauthorized")

    title = RESEARCH_PARTICIPATION_DEMO_TITLE
    calendly_url = RESEARCH_PARTICIPATION_DEMO_CALENDLY_URL
    title_compact = _compact_text(title)
    candidates = db.query(Survey).filter(
        Survey.title.ilike("%Motivations%"),
        Survey.title.ilike("%Barriers%"),
        Survey.title.ilike("%Research%"),
    ).order_by(Survey.created_at.desc()).all()
    existing = []
    for survey in candidates:
        survey_title_compact = _compact_text(survey.title)
        if survey_title_compact == title_compact or (
            "motivation" in survey_title_compact
            and "barrier" in survey_title_compact
            and "onlinesurveys" in survey_title_compact
            and "researchstudies" in survey_title_compact
        ):
            existing.append(survey)
    source = existing[0] if existing else None
    admin_user = _get_admin_publisher_user(db)

    source_data = {
        "description": (
            getattr(source, "description", None)
            or "Help us understand what motivates people to participate in online surveys and research studies, what barriers get in the way, and what would make research participation feel clearer, safer, and more worthwhile."
        ),
        "category": getattr(source, "category", None) or "research",
        "estimated_time": getattr(source, "estimated_time", None) or 30,
        "reward_amount": getattr(source, "reward_amount", None) or 0.0,
        "admin_display_reward_amount": getattr(source, "admin_display_reward_amount", None),
        "target_responses": getattr(source, "target_responses", None) or 50,
        "target_age_range": getattr(source, "target_age_range", None),
        "target_education_min": getattr(source, "target_education_min", None),
        "target_education_max": getattr(source, "target_education_max", None),
        "target_field": getattr(source, "target_field", None),
        "target_status": getattr(source, "target_status", None),
        "target_state": getattr(source, "target_state", None),
        "target_language": getattr(source, "target_language", None),
        "target_ethnicity": getattr(source, "target_ethnicity", None),
        "target_sexual_orientation": getattr(source, "target_sexual_orientation", None),
        "target_mental_health_diagnosis": getattr(source, "target_mental_health_diagnosis", None),
        "target_physical_health_diagnosis": getattr(source, "target_physical_health_diagnosis", None),
        "target_sport_type": getattr(source, "target_sport_type", None),
        "target_sport_frequency": getattr(source, "target_sport_frequency", None),
        "target_smoking": getattr(source, "target_smoking", None),
        "target_cannabis_use": getattr(source, "target_cannabis_use", None),
        "target_student_status": getattr(source, "target_student_status", None),
        "target_international_domestic": getattr(source, "target_international_domestic", None),
        "target_experience_tags": getattr(source, "target_experience_tags", None),
        "target_device": getattr(source, "target_device", None),
        "target_income_level": getattr(source, "target_income_level", None),
        "target_lifestyle_tags": getattr(source, "target_lifestyle_tags", None),
        "target_niche_requirements": getattr(source, "target_niche_requirements", None),
        "image_url": getattr(source, "image_url", None),
        "share_slug": getattr(source, "share_slug", None),
        "incentive_type": getattr(source, "incentive_type", None) or "cash",
        "raffle_prize_type": getattr(source, "raffle_prize_type", None),
        "session_count": None,
        "sessions_per_week": None,
    }
    if source_data["admin_display_reward_amount"] is None:
        source_data["admin_display_reward_amount"] = source_data["reward_amount"]

    deleted = []
    for survey in existing:
        deleted.append({"id": survey.id, "title": survey.title})
        _delete_survey_tree(db, survey)

    survey = Survey(
        publisher_id=admin_user.id,
        title=title,
        description=source_data["description"],
        form_url=calendly_url,
        task_type="interview",
        category=source_data["category"],
        estimated_time=source_data["estimated_time"],
        reward_amount=source_data["reward_amount"],
        admin_display_reward_amount=source_data["admin_display_reward_amount"],
        per_person_gross=0.0,
        total_budget=0.0,
        commission_rate=0.0,
        payment_status="admin_demo",
        target_responses=source_data["target_responses"],
        status="published",
        published_at=datetime.utcnow(),
        target_age_range=source_data["target_age_range"],
        target_education_min=source_data["target_education_min"],
        target_education_max=source_data["target_education_max"],
        target_field=source_data["target_field"],
        target_status=source_data["target_status"],
        target_state=source_data["target_state"],
        target_language=source_data["target_language"],
        target_ethnicity=source_data["target_ethnicity"],
        target_sexual_orientation=source_data["target_sexual_orientation"],
        target_mental_health_diagnosis=source_data["target_mental_health_diagnosis"],
        target_physical_health_diagnosis=source_data["target_physical_health_diagnosis"],
        target_sport_type=source_data["target_sport_type"],
        target_sport_frequency=source_data["target_sport_frequency"],
        target_smoking=source_data["target_smoking"],
        target_cannabis_use=source_data["target_cannabis_use"],
        target_student_status=source_data["target_student_status"],
        target_year_in_school=None,
        target_international_domestic=source_data["target_international_domestic"],
        target_experience_tags=source_data["target_experience_tags"],
        target_participation_format="Video interview",
        target_device=source_data["target_device"],
        target_income_level=source_data["target_income_level"],
        target_lifestyle_tags=source_data["target_lifestyle_tags"],
        target_niche_requirements=source_data["target_niche_requirements"],
        incentive_type=source_data["incentive_type"],
        raffle_prize_type=source_data["raffle_prize_type"],
        image_url=source_data["image_url"],
        share_slug=source_data["share_slug"],
        session_count=source_data["session_count"],
        sessions_per_week=source_data["sessions_per_week"],
        availability_slots=None,
    )
    db.add(survey)
    if not survey.share_slug:
        _ensure_survey_share_slug(db, survey)
    db.commit()
    db.refresh(survey)
    return JSONResponse({
        "success": True,
        "deleted": deleted,
        "created": {
            "id": survey.id,
            "title": survey.title,
            "share_url": _survey_share_url(request, db, survey),
            "start_url": survey.form_url,
            "type_label": _participant_study_type_label(survey),
        },
    })


@app.get("/admin/feedbacks")
async def list_feedbacks(request: Request, admin_key: str = Query(None), db: Session = Depends(get_db)):
    if not _admin_key_matches(admin_key): raise HTTPException(403, "Unauthorized")
    feedbacks = db.query(Feedback).order_by(Feedback.created_at.desc()).all()
    result = []
    for f in feedbacks:
        user = db.query(User).filter(User.id == f.user_id).first()
        result.append({"id": f.id, "user_email": user.email if user else "unknown", "category": f.category, "title": f.title, "content": f.content, "status": f.status, "credit_amount": f.credit_amount, "created_at": str(f.created_at)})
    return JSONResponse(result)

@app.post("/admin/discovery/find")
async def admin_find_discovery_channels(request: Request):
    body = await request.json()
    if not _admin_key_matches(body.get("admin_key")): raise HTTPException(403, "Unauthorized")
    try:
        criteria = DiscoveryCriteria(**(body.get("criteria") or {}))
    except Exception as exc:
        raise HTTPException(422, f"Invalid discovery criteria: {exc}")
    result = discover_channels(criteria)
    result.channels = rank_discovery_channels(result.channels, criteria)
    return result

@app.get("/admin/support/threads")
async def admin_support_threads(admin_key: str = Query(None), db: Session = Depends(get_db)):
    if not _admin_key_matches(admin_key): raise HTTPException(403, "Unauthorized")
    threads = db.query(SupportThread).order_by(SupportThread.last_message_at.desc()).all()
    return JSONResponse([_support_thread_payload(thread, db) for thread in threads])

@app.get("/admin/support/threads/{thread_id}/messages")
async def admin_support_messages(thread_id: int, admin_key: str = Query(None), db: Session = Depends(get_db)):
    if not _admin_key_matches(admin_key): raise HTTPException(403, "Unauthorized")
    thread = db.query(SupportThread).filter(SupportThread.id == thread_id).first()
    if not thread:
        raise HTTPException(404, "Support thread not found")
    messages = db.query(SupportMessage).filter(
        SupportMessage.thread_id == thread.id
    ).order_by(SupportMessage.created_at.asc()).all()
    now = datetime.utcnow()
    for msg in messages:
        if msg.sender_type == "user" and msg.read_at is None:
            msg.read_at = now
    db.commit()
    return JSONResponse({
        "thread": _support_thread_payload(thread, db),
        "messages": [_support_message_payload(m) for m in messages],
        "available_hours": SUPPORT_AVAILABLE_HOURS,
    })

@app.post("/admin/support/threads/{thread_id}/messages")
async def admin_send_support_message(thread_id: int, request: Request, db: Session = Depends(get_db)):
    body = await request.json()
    if not _admin_key_matches(body.get("admin_key")): raise HTTPException(403, "Unauthorized")
    text_body = (body.get("body") or "").strip()
    if not text_body:
        raise HTTPException(400, "Message cannot be empty")
    if len(text_body) > 2000:
        raise HTTPException(400, "Message is too long")
    thread = db.query(SupportThread).filter(SupportThread.id == thread_id).first()
    if not thread:
        raise HTTPException(404, "Support thread not found")
    msg = SupportMessage(thread_id=thread.id, sender_type="admin", body=text_body)
    thread.status = "open"
    thread.last_message_at = datetime.utcnow()
    db.add(msg)
    db.commit()
    db.refresh(msg)
    return JSONResponse({"message": _support_message_payload(msg), "thread": _support_thread_payload(thread, db)})

@app.post("/admin/support/threads/{thread_id}/status")
async def admin_update_support_status(thread_id: int, request: Request, db: Session = Depends(get_db)):
    body = await request.json()
    if not _admin_key_matches(body.get("admin_key")): raise HTTPException(403, "Unauthorized")
    status = (body.get("status") or "").strip().lower()
    if status not in {"open", "closed"}:
        raise HTTPException(400, "Unsupported status")
    thread = db.query(SupportThread).filter(SupportThread.id == thread_id).first()
    if not thread:
        raise HTTPException(404, "Support thread not found")
    thread.status = status
    thread.updated_at = datetime.utcnow()
    db.commit()
    return JSONResponse({"thread": _support_thread_payload(thread, db)})

# ---------------------------
# Quick create built-in survey
# ---------------------------

@app.post("/surveys/create-builtin")
def create_builtin_survey(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    survey = Survey(
        publisher_id=current_user.id,
        title="Untitled Survey",
        description="",
        form_url="__builtin__",
        task_type="survey",
        category="research",
        estimated_time=10,
        reward_amount=0.0,
        per_person_gross=0.0,
        total_budget=0.0,
        commission_rate=0.15,
        payment_status="unpaid",
        target_responses=50,
        status="draft",
    )
    _ensure_survey_share_slug(db, survey)
    db.add(survey)
    db.commit()
    db.refresh(survey)
    _record_user_event(
        db,
        request,
        "study_created",
        user=current_user,
        target_type="survey",
        target_id=survey.id,
        metadata={"study_title": survey.title, "source": "Builtin builder", "user_role": "researcher"},
    )
    db.commit()
    return RedirectResponse(f"/surveys/{survey.id}/builder", status_code=303)

@app.post("/publish")
async def publish_survey(
    request: Request,
    title: str = Form(...), description: str = Form(...), form_url: str = Form(...),
    task_type: str = Form("survey"), category: str = Form(...), estimated_time: int = Form(...),
    per_person_gross: Optional[str] = Form(None), total_budget: Optional[str] = Form(None),
    target_responses: int = Form(...), urgency_level: str = Form(None), incentive_type: str = Form(None),
    target_age_range: str = Form(None), target_field: str = Form(None), target_status: str = Form(None),
    target_state: str = Form(None), target_language: str = Form(None), target_ethnicity: str = Form(None),
    target_sexual_orientation: str = Form(None), target_mental_health_diagnosis: str = Form(None),
    target_physical_health_diagnosis: str = Form(None), target_sport_type: str = Form(None),
    target_sport_frequency: str = Form(None), target_smoking: str = Form(None),
    target_cannabis_use: str = Form(None), target_student_status: str = Form(None),
    target_year_in_school: str = Form(None), target_international_domestic: str = Form(None),
    target_participation_format: str = Form(None), target_device: str = Form(None),
    target_income_level: str = Form(None), raffle_prize_type: str = Form(None),
    cover_image: UploadFile = File(None),
    admin_display_reward_amount: Optional[float] = Form(None),
    admin_publish: bool = False,
    admin_redirect_key: Optional[str] = None,
    current_user: User = Depends(get_current_user), db: Session = Depends(get_db)
):
    form = await request.form()
    target_education_min = _parse_optional_int(form.get("target_education_min"))
    target_education_max = _parse_optional_int(form.get("target_education_max"))
    experience_list = form.getlist("target_experience_tags")
    target_experience_tags = ",".join(experience_list) if experience_list else None
    lifestyle_list = form.getlist("target_lifestyle_tags")
    target_lifestyle_tags = ",".join(lifestyle_list) if lifestyle_list else None
    target_niche_requirements = _clean_target(form.get("target_niche_requirements"))
    participant_benefits = _join_form_list_with_other(
        form.getlist("participant_benefits"),
        form.get("participant_benefits_other")
    )
    existing_survey_id = int(form.get("existing_survey_id") or 0)
    raw_task_type = (task_type or form.get("task_type") or "survey").strip()
    is_online_interview = raw_task_type in {"online_interview", "remote_interview"}
    is_in_person_study = raw_task_type in {"in_person_study", "in_person"}
    is_interview_listing = raw_task_type == "interview" or is_online_interview or is_in_person_study
    normalized_task_type = "interview" if is_interview_listing else raw_task_type
    if is_online_interview:
        target_participation_format = "Video interview"
    elif is_in_person_study:
        target_participation_format = "In-person study"
    session_count = max(_parse_optional_int(form.get("session_count")) or 1, 1)
    sessions_per_week = _parse_optional_int(form.get("sessions_per_week"))
    availability_slots = (form.get("availability_slots") or "").strip()
    if availability_slots:
        try:
            parsed_slots = json.loads(availability_slots)
            availability_slots = json.dumps(parsed_slots if isinstance(parsed_slots, list) else [])
        except Exception:
            availability_slots = None
    else:
        availability_slots = None
    scheduling_link = (form.get("scheduling_link") or "").strip()
    interview_location = _clean_target(form.get("interview_location"))
    if is_interview_listing:
        form_url = scheduling_link or form_url or ""

    is_builtin = (form_url == "__builtin__")
    incentive_clean = _clean_target(incentive_type) or "cash"
    raffle_clean = _clean_target(raffle_prize_type)
    is_raffle = incentive_clean == "raffle"
    is_volunteer = incentive_clean == "volunteer"
    is_no_pay = is_raffle

    if is_raffle:
        ppg = 0.0; rate = 0.0; reward = 0.0; total = 0.0
    elif is_volunteer:
        ppg = 0.0; rate = 0.0; reward = 0.0
        total = volunteer_platform_fee(target_responses)
    else:
        per_person_gross_value = _parse_optional_float(per_person_gross)
        total_budget_value = _parse_optional_float(total_budget)
        if admin_publish and not per_person_gross_value and not total_budget_value:
            ppg = 0.0; rate = 0.0; reward = 0.0; total = 0.0
        elif per_person_gross_value is not None and per_person_gross_value > 0:
            ppg = per_person_gross_value
        elif total_budget_value is not None and total_budget_value > 0:
            ppg = total_budget_value / (int(target_responses) * session_count)
        else: ppg = 5.0
        if not (admin_publish and ppg == 0.0):
            rate, reward = calculate_commission(ppg)
            total = round(ppg * int(target_responses) * session_count, 2)
    total = round(total * timeline_multiplier(urgency_level), 2)

    image_url = None
    if cover_image and cover_image.filename:
        uploads_dir = Path("app/static/uploads"); uploads_dir.mkdir(exist_ok=True)
        unique_filename = f"{uuid.uuid4()}{Path(cover_image.filename).suffix}"
        file_path = uploads_dir / unique_filename
        with file_path.open("wb") as buffer: shutil.copyfileobj(cover_image.file, buffer)
        image_url = f"/static/uploads/{unique_filename}"

    if existing_survey_id:
        survey = db.query(Survey).filter(
            Survey.id == existing_survey_id,
            Survey.publisher_id == current_user.id
        ).first()
        if not survey:
            raise HTTPException(404, "Survey not found")
        survey.title = title
        survey.description = description
        survey.form_url = form_url
        survey.task_type = normalized_task_type
        survey.category = category
        survey.estimated_time = estimated_time
        survey.reward_amount = reward
        survey.per_person_gross = ppg
        survey.total_budget = total
        survey.commission_rate = rate
        survey.payment_status = "unpaid" if not is_no_pay else "paid"
        survey.target_responses = target_responses
        survey.urgency_level = _clean_target(urgency_level)
        survey.incentive_type = incentive_clean
        survey.raffle_prize_type = raffle_clean if is_raffle else None
        survey.target_age_range = _clean_target(target_age_range)
        survey.target_education_min = target_education_min
        survey.target_education_max = target_education_max
        survey.target_field = _clean_target(target_field)
        survey.target_status = _clean_target(target_status)
        survey.target_state = _clean_target(target_state)
        survey.target_language = _clean_target(target_language)
        survey.target_ethnicity = _clean_target(target_ethnicity)
        survey.target_sexual_orientation = _clean_target(target_sexual_orientation)
        survey.target_mental_health_diagnosis = _clean_target(target_mental_health_diagnosis)
        survey.target_physical_health_diagnosis = _clean_target(target_physical_health_diagnosis)
        survey.target_sport_type = _clean_target(target_sport_type)
        survey.target_sport_frequency = _clean_target(target_sport_frequency)
        survey.target_smoking = _clean_target(target_smoking)
        survey.target_cannabis_use = _clean_target(target_cannabis_use)
        survey.target_student_status = _clean_target(target_student_status)
        survey.target_year_in_school = None
        survey.target_international_domestic = _clean_target(target_international_domestic)
        survey.target_experience_tags = target_experience_tags
        survey.target_participation_format = _clean_target(target_participation_format)
        survey.target_device = _clean_target(target_device)
        survey.target_income_level = _clean_target(target_income_level)
        survey.target_lifestyle_tags = target_lifestyle_tags
        survey.target_niche_requirements = target_niche_requirements
        survey.participant_benefits = participant_benefits
        survey.availability_slots = availability_slots if is_interview_listing else None
        survey.interview_location = interview_location if is_interview_listing else None
        survey.session_count = session_count if is_interview_listing else None
        survey.sessions_per_week = sessions_per_week if is_interview_listing else None
        survey.admin_display_reward_amount = admin_display_reward_amount if admin_publish else None
        _ensure_survey_share_slug(db, survey)
        _apply_survey_auto_filter_settings(survey, form)
        db.commit()
        db.refresh(survey)
    else:
        survey = Survey(
            publisher_id=current_user.id, title=title, description=description, form_url=form_url,
            task_type=normalized_task_type, category=category, estimated_time=estimated_time,
            reward_amount=reward, per_person_gross=ppg, total_budget=total, commission_rate=rate,
            payment_status="unpaid" if not is_no_pay else "paid",
            target_responses=target_responses, urgency_level=_clean_target(urgency_level),
            incentive_type=incentive_clean, raffle_prize_type=raffle_clean if is_raffle else None,
            target_age_range=_clean_target(target_age_range),
            target_education_min=target_education_min, target_education_max=target_education_max,
            target_field=_clean_target(target_field), target_status=_clean_target(target_status),
            target_state=_clean_target(target_state), target_language=_clean_target(target_language),
            target_ethnicity=_clean_target(target_ethnicity),
            target_sexual_orientation=_clean_target(target_sexual_orientation),
            target_mental_health_diagnosis=_clean_target(target_mental_health_diagnosis),
            target_physical_health_diagnosis=_clean_target(target_physical_health_diagnosis),
            target_sport_type=_clean_target(target_sport_type),
            target_sport_frequency=_clean_target(target_sport_frequency),
            target_smoking=_clean_target(target_smoking), target_cannabis_use=_clean_target(target_cannabis_use),
            target_student_status=_clean_target(target_student_status),
            target_year_in_school=None,
            target_international_domestic=_clean_target(target_international_domestic),
            target_experience_tags=target_experience_tags,
            target_participation_format=_clean_target(target_participation_format),
            target_device=_clean_target(target_device),
            target_income_level=_clean_target(target_income_level),
            target_lifestyle_tags=target_lifestyle_tags,
            target_niche_requirements=target_niche_requirements,
            participant_benefits=participant_benefits,
            availability_slots=availability_slots if is_interview_listing else None,
            interview_location=interview_location if is_interview_listing else None,
            session_count=session_count if is_interview_listing else None,
            sessions_per_week=sessions_per_week if is_interview_listing else None,
            admin_display_reward_amount=admin_display_reward_amount if admin_publish else None,
            image_url=image_url, status="draft", published_at=None, closed_at=None,
        )
        _ensure_survey_share_slug(db, survey)
        _apply_survey_auto_filter_settings(survey, form)
        db.add(survey); db.commit(); db.refresh(survey)
        _record_user_event(
            db,
            request,
            "study_created",
            user=current_user,
            target_type="survey",
            target_id=survey.id,
            metadata={"study_title": survey.title, "source": "Publish flow", "user_role": "researcher"},
        )

    if admin_publish:
        survey.status = "published"
        survey.payment_status = "admin_demo"
        survey.published_at = datetime.utcnow()
        _record_user_event(
            db,
            request,
            "study_published",
            user=current_user,
            target_type="survey",
            target_id=survey.id,
            metadata={"study_title": survey.title, "source": "Admin publish", "user_role": "researcher"},
        )
        db.commit()
        redirect_qs = urlencode({"admin_key": admin_redirect_key or "", "created": "1"})
        return RedirectResponse(f"/admin/publish?{redirect_qs}", status_code=303)

    # Handle no-pay incentives, missing Stripe key, and Stripe checkout
    if is_no_pay or not stripe.api_key:
        survey.status = "published"
        survey.published_at = datetime.utcnow()
        if not stripe.api_key and not is_no_pay:
            survey.payment_status = "paid"
        _record_user_event(
            db,
            request,
            "study_published",
            user=current_user,
            target_type="survey",
            target_id=survey.id,
            metadata={"study_title": survey.title, "source": "Publish flow", "user_role": "researcher"},
        )
        db.commit()
        return RedirectResponse("/publisher", status_code=303)

    success_url = f"https://insightaco.org/payment/success?survey_id={survey.id}"
    if is_volunteer:
        stripe_description = f"Volunteer recruitment fee: ${volunteer_platform_fee(target_responses):.2f} per 10 participants"
    else:
        stripe_description = f"{survey.target_responses} responses x ${reward:.2f} per person"
    session = stripe.checkout.Session.create(
        payment_method_types=["card"],
        line_items=[{"price_data": {"currency": "usd", "product_data": {"name": f"Survey: {survey.title}", "description": stripe_description}, "unit_amount": int(round(total * 100))}, "quantity": 1}],
        mode="payment",
        success_url=success_url,
        cancel_url="https://insightaco.org/publisher",
        metadata={"survey_id": str(survey.id), "publisher_id": str(current_user.id)}
    )
    survey.stripe_payment_intent_id = session.id; db.commit()
    return RedirectResponse(session.url, status_code=303)

# ---------------------------
# Update survey info from builder
# ---------------------------

@app.post("/surveys/{survey_id}/update-info")
async def update_survey_info(
    survey_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    survey = db.query(Survey).filter(
        Survey.id == survey_id,
        Survey.publisher_id == current_user.id
    ).first()
    if not survey:
        raise HTTPException(404, "Survey not found")

    body = await request.json()
    survey.title = body.get("title", survey.title)
    survey.description = body.get("description", survey.description)
    survey.category = body.get("category", survey.category)
    survey.estimated_time = body.get("estimated_time", survey.estimated_time)
    survey.per_person_gross = body.get("per_person_gross", survey.per_person_gross)
    survey.reward_amount = body.get("reward_amount", survey.reward_amount)
    survey.total_budget = body.get("total_budget", survey.total_budget)
    survey.commission_rate = body.get("commission_rate", survey.commission_rate)
    survey.target_responses = body.get("target_responses", survey.target_responses)
    if "quality_auto_filter_enabled" in body:
        survey.quality_auto_filter_enabled = bool(body.get("quality_auto_filter_enabled"))
    if "quality_auto_filter_min_score" in body:
        survey.quality_auto_filter_min_score = _parse_auto_approve_min_score(
            body.get("quality_auto_filter_min_score", 80)
        )
    db.commit()
    return JSONResponse({"message": "updated"})
# ---------------------------
# Questions API
# ---------------------------

@app.post("/surveys/{survey_id}/questions")
async def add_question(
    survey_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    survey = db.query(Survey).filter(
        Survey.id == survey_id,
        Survey.publisher_id == current_user.id
    ).first()
    if not survey:
        raise HTTPException(404, "Survey not found")

    request_data = await request.json()

    max_order = db.query(func.max(Question.order_index)).filter(
        Question.survey_id == survey_id
    ).scalar() or 0

    q = Question(
        survey_id=survey_id,
        question_text=request_data.get("question_text", ""),
        question_type=request_data.get("question_type", "single"),
        options=request_data.get("options"),
        is_required=request_data.get("is_required", True),
        order_index=request_data.get("order_index", max_order + 1)
    )
    db.add(q)
    db.commit()
    db.refresh(q)
    return JSONResponse({
        "id": q.id,
        "survey_id": q.survey_id,
        "question_text": q.question_text,
        "question_type": q.question_type,
        "options": q.options,
        "is_required": q.is_required,
        "order_index": q.order_index
    })


@app.get("/surveys/{survey_id}/questions")
def get_questions(
    survey_id: int,
    db: Session = Depends(get_db)
):
    questions = db.query(Question).filter(
        Question.survey_id == survey_id
    ).order_by(Question.order_index).all()
    return JSONResponse([{
        "id": q.id,
        "question_text": q.question_text,
        "question_type": q.question_type,
        "options": q.options,
        "is_required": q.is_required,
        "order_index": q.order_index
    } for q in questions])


@app.put("/surveys/{survey_id}/questions/{question_id}")
async def update_question(
    survey_id: int,
    question_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    survey = db.query(Survey).filter(
        Survey.id == survey_id,
        Survey.publisher_id == current_user.id
    ).first()
    if not survey:
        raise HTTPException(404, "Survey not found")

    request_data = await request.json()

    q = db.query(Question).filter(
        Question.id == question_id,
        Question.survey_id == survey_id
    ).first()
    if not q:
        raise HTTPException(404, "Question not found")

    if "question_text" in request_data:
        q.question_text = request_data["question_text"]
    if "question_type" in request_data:
        q.question_type = request_data["question_type"]
    if "options" in request_data:
        q.options = request_data["options"]
    if "is_required" in request_data:
        q.is_required = request_data["is_required"]

    db.commit()
    return JSONResponse({"message": "updated"})


@app.delete("/surveys/{survey_id}/questions/{question_id}")
def delete_question(
    survey_id: int,
    question_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    survey = db.query(Survey).filter(
        Survey.id == survey_id,
        Survey.publisher_id == current_user.id
    ).first()
    if not survey:
        raise HTTPException(404, "Survey not found")

    q = db.query(Question).filter(
        Question.id == question_id,
        Question.survey_id == survey_id
    ).first()
    if not q:
        raise HTTPException(404, "Question not found")

    db.delete(q)
    db.commit()
    return JSONResponse({"message": "deleted"})


@app.post("/surveys/{survey_id}/questions/reorder")
async def reorder_questions(
    survey_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    survey = db.query(Survey).filter(
        Survey.id == survey_id,
        Survey.publisher_id == current_user.id
    ).first()
    if not survey:
        raise HTTPException(404, "Survey not found")

    body = await request.json()
    for item in body:
        q = db.query(Question).filter(
            Question.id == item["id"],
            Question.survey_id == survey_id
        ).first()
        if q:
            q.order_index = item["order_index"]

    db.commit()
    return JSONResponse({"message": "reordered"})


# ---------------------------
# Survey Builder page
# ---------------------------

@app.get("/surveys/{survey_id}/builder", response_class=HTMLResponse)
def survey_builder(
    survey_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    survey = db.query(Survey).filter(
        Survey.id == survey_id,
        Survey.publisher_id == current_user.id
    ).first()
    if not survey:
        raise HTTPException(404, "Survey not found")
    return templates.TemplateResponse("survey_builder.html", {
        "request": request,
        "survey": survey,
        "survey_id": survey_id,
        "current_user": current_user
    })


# ---------------------------
# Survey Take page
# ---------------------------

@app.get("/surveys/{survey_id}/take", response_class=HTMLResponse)
def survey_take(
    survey_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    survey = db.query(Survey).filter(
        Survey.id == survey_id,
        Survey.status == "published"
    ).first()
    if not survey:
        raise HTTPException(404, "Survey not found")
    return templates.TemplateResponse("survey_take.html", {
        "request": request,
        "survey": survey,
        "current_user": current_user
    })


# ---------------------------
# Survey Results page (HTML)
# ---------------------------

@app.get("/surveys/{survey_id}/results", response_class=HTMLResponse)
def survey_results_page(
    survey_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    survey = db.query(Survey).filter(
        Survey.id == survey_id,
        Survey.publisher_id == current_user.id
    ).first()
    if not survey:
        raise HTTPException(404, "Survey not found")
    return templates.TemplateResponse("survey_results.html", {
        "request": request,
        "survey": survey,
        "survey_id": survey_id,
        "current_user": current_user
    })


# ---------------------------
# Survey Results data (JSON) — used by survey_results.html
# ---------------------------

@app.get("/api/surveys/{survey_id}/results")
def get_survey_results(
    survey_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    survey = db.query(Survey).filter(
        Survey.id == survey_id,
        Survey.publisher_id == current_user.id
    ).first()
    if not survey:
        raise HTTPException(404, "Survey not found")

    questions = db.query(Question).filter(
        Question.survey_id == survey_id
    ).order_by(Question.order_index).all()

    total_responses = db.query(Response).filter(
        Response.survey_id == survey_id,
        Response.status == "completed"
    ).count()

    result = []
    for q in questions:
        answers = db.query(Answer).join(Response).filter(
            Response.survey_id == survey_id,
            Response.status == "completed",
            Answer.question_id == q.id
        ).all()

        if q.question_type in ("single", "multiple", "dropdown"):
            distribution = {}
            for a in answers:
                val = a.answer_value
                if isinstance(val, list):
                    for v in val:
                        distribution[v] = distribution.get(v, 0) + 1
                else:
                    distribution[str(val)] = distribution.get(str(val), 0) + 1
            result.append({
                "question_id": q.id,
                "question_text": q.question_text,
                "question_type": q.question_type,
                "total_answers": len(answers),
                "distribution": distribution
            })

        elif q.question_type == "scale":
            values = [a.answer_value for a in answers if a.answer_value is not None]
            avg = round(sum(values) / len(values), 2) if values else 0
            distribution = {str(i): values.count(i) for i in range(1, 6)}
            result.append({
                "question_id": q.id,
                "question_text": q.question_text,
                "question_type": q.question_type,
                "total_answers": len(answers),
                "average": avg,
                "distribution": distribution
            })

        elif q.question_type == "text":
            result.append({
                "question_id": q.id,
                "question_text": q.question_text,
                "question_type": q.question_type,
                "total_answers": len(answers),
                "responses": [a.answer_value for a in answers]
            })

    return JSONResponse({
        "survey_id": survey_id,
        "survey_title": survey.title,
        "total_responses": total_responses,
        "questions": result
    })


# ---------------------------
# AI Analyze survey
# ---------------------------

@app.post("/surveys/{survey_id}/analyze")
async def analyze_survey(
    survey_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    survey = db.query(Survey).filter(
        Survey.id == survey_id,
        Survey.publisher_id == current_user.id
    ).first()
    if not survey:
        raise HTTPException(404, "Survey not found")

    questions = db.query(Question).filter(
        Question.survey_id == survey_id
    ).order_by(Question.order_index).all()

    answers_data = []
    for q in questions:
        answers = db.query(Answer).join(Response).filter(
            Response.survey_id == survey_id,
            Response.status == "completed",
            Answer.question_id == q.id
        ).all()
        answers_data.append({
            "question": q.question_text,
            "type": q.question_type,
            "answers": [a.answer_value for a in answers]
        })

    total_responses = db.query(Response).filter(
        Response.survey_id == survey_id,
        Response.status == "completed"
    ).count()

    import anthropic, json as _json
    api_key = (os.environ.get("ANTHROPIC_API_KEY") or "").strip()
    if not api_key:
        raise HTTPException(503, "AI analysis is not available right now.")
    client = anthropic.Anthropic(api_key=api_key)
    message = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=1500,
        messages=[{
            "role": "user",
            "content": f"""You are analyzing survey results for a researcher.

Survey title: {survey.title}
Survey description: {survey.description}
Total responses: {total_responses}

Data:
{_json.dumps(answers_data, ensure_ascii=False, indent=2)}

Please provide:
1. Key findings (3-5 bullet points with actual numbers)
2. Notable patterns or trends
3. Text response themes (if any open-ended questions)
4. One recommendation for the researcher

Be concise and specific. Use the actual numbers from the data.
Write in the same language as the survey questions.
"""
        }]
    )

    return JSONResponse({
        "analysis": message.content[0].text,
        "generated_at": datetime.utcnow().isoformat()
    })


# ---------------------------
# Answers submit API
# ---------------------------

@app.post("/surveys/{survey_id}/submit")
async def submit_answers(
    survey_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    survey = db.query(Survey).filter(Survey.id == survey_id).first()
    if not survey:
        raise HTTPException(404, "Survey not found")
    if survey.status != "published":
        raise HTTPException(400, "Survey not published")

    body = await request.json()
    meta = _extract_client_meta(request, current_user)

    r = db.query(Response).filter(
        Response.survey_id == survey_id,
        Response.participant_id == current_user.id
    ).first()
    if not r:
        r = Response(
            survey_id=survey_id,
            participant_id=current_user.id,
            status="started",
            client_ip=meta["client_ip"],
            user_agent=meta["user_agent"],
            device_fingerprint=meta["device_fingerprint"],
        )
        db.add(r)
        db.flush()
    else:
        r.client_ip = meta["client_ip"]
        r.user_agent = meta["user_agent"]
        r.device_fingerprint = meta["device_fingerprint"]

    for item in body:
        existing = db.query(Answer).filter(
            Answer.response_id == r.id,
            Answer.question_id == item["question_id"]
        ).first()
        if existing:
            existing.answer_value = item["answer_value"]
        else:
            db.add(Answer(
                response_id=r.id,
                question_id=item["question_id"],
                answer_value=item["answer_value"]
            ))

    if r.status != "completed":
        r.status = "completed"
        r.completed_at = datetime.now(timezone.utc)
        r.payout_amount = survey.reward_amount
        mark_response_under_review(r)

        publisher = db.query(User).filter(User.id == survey.publisher_id).first()
        if publisher and publisher.email:
            send_email(
                to=publisher.email,
                subject=f"[Insighta] New response ready for review: {survey.title}",
                body=f"""
                <div style="font-family: sans-serif; max-width: 560px; margin: 0 auto; padding: 32px 24px; color: #1a1a18;">
                  <h2 style="font-size: 22px; margin-bottom: 8px;">📋 Response Ready for Review</h2>
                  <p style="color: #8a8a82; margin-bottom: 24px;">A participant has completed your survey.</p>
                  <div style="background: #f3f1ea; border-radius: 10px; padding: 20px 24px; margin-bottom: 24px;">
                    <div style="font-size: 13px; color: #8a8a82; margin-bottom: 4px;">Survey</div>
                    <div style="font-size: 17px; font-weight: 600; margin-bottom: 12px;">{survey.title}</div>
                    <div style="font-size: 13px; color: #8a8a82; margin-bottom: 4px;">Participant</div>
                    <div style="font-size: 15px;">{current_user.email}</div>
                  </div>
                  <a href="https://insightaco.org/publisher" style="display: inline-block; padding: 12px 24px; background: #2d6a4f; color: white; border-radius: 8px; text-decoration: none; font-weight: 600; font-size: 14px;">
                    Review & Approve →
                  </a>
                </div>
                """
            )

    quality_payload = None
    try:
        quality_result = evaluate_builtin_response(db, survey_id=survey_id, response_id=r.id)
        upsert_builtin_quality_check(db, survey_id=survey_id, response_id=r.id, result=quality_result)
        quality_payload = {
            "score": quality_result.quality_score,
            "label": quality_result.quality_label,
            "fraud_risk": quality_result.fraud_risk,
            "reasons": quality_result.reasons,
        }
    except Exception as exc:
        print(f"Quality check failed for response {r.id}: {exc}")

    db.commit()
    mark_latest_jump_completed_for_response(db, r)
    body = {"message": "submitted successfully"}
    if quality_payload:
        body["quality"] = quality_payload
    return JSONResponse(body)


@app.get("/api/surveys/{survey_id}/quality-results")
def get_survey_quality_results(
    survey_id: int,
    scope: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    survey = db.query(Survey).filter(
        Survey.id == survey_id,
        Survey.publisher_id == current_user.id,
    ).first()
    if not survey:
        raise HTTPException(404, "Survey not found")

    normalized_scope = (scope or "").strip().lower()
    if normalized_scope == "builtin":
        ensure_builtin_quality_checks(db, survey_id)

    rows = db.query(ResponseQualityCheck).filter(
        ResponseQualityCheck.survey_id == survey_id
    ).order_by(ResponseQualityCheck.created_at.desc()).all()

    if normalized_scope == "builtin":
        rows = [row for row in rows if row.source_type == "builtin"]
    elif normalized_scope == "excel":
        rows = _latest_excel_quality_rows([row for row in rows if row.source_type == "excel"])
    else:
        rows = _latest_excel_quality_rows(rows)

    payload = []
    for index, row in enumerate(rows):
        participant_email = None
        if row.response_id:
            resp = db.query(Response).filter(Response.id == row.response_id).first()
            if resp:
                participant = db.query(User).filter(User.id == resp.participant_id).first()
                participant_email = participant.email if participant else None
        row_label = participant_email or _quality_row_label(index)
        payload.append({
            "id": row.id,
            "source_type": row.source_type,
            "source_ref": row.source_ref,
            "row_label": row_label,
            "response_id": row.response_id,
            "participant_email": participant_email,
            "quality_score": row.quality_score,
            "quality_label": row.quality_label,
            "fraud_risk": row.fraud_risk,
            "rule_penalty": row.rule_penalty,
            "anomaly_score": row.anomaly_score,
            "semantic_risk": row.semantic_risk,
            "triggered_rules": row.triggered_rules or [],
            "reasons": row.reasons or [],
            "review_status": row.review_status,
            "reviewer_label": row.reviewer_label,
            "llm_result_json": row.llm_result_json,
            "notes": row.notes,
            "created_at": row.created_at.isoformat() if row.created_at else None,
        })
    return JSONResponse(payload)


@app.get("/api/quality-checks/{check_id}")
def get_quality_check_detail(
    check_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    row, survey = _get_quality_check_for_publisher(db, check_id, current_user.id)

    participant_email = None
    participant_name = None
    response_meta = None
    answers_payload = []
    submission_history = []

    if row.response_id:
        resp = db.query(Response).filter(Response.id == row.response_id).first()
        if resp:
            participant = db.query(User).filter(User.id == resp.participant_id).first()
            participant_email = participant.email if participant else None
            participant_name = (participant.username or participant.email) if participant else None
            response_meta = {
                "client_ip": resp.client_ip,
                "user_agent": resp.user_agent,
                "device_fingerprint": resp.device_fingerprint,
                "started_at": resp.started_at.isoformat() if resp.started_at else None,
                "completed_at": resp.completed_at.isoformat() if resp.completed_at else None,
            }
            answers = db.query(Answer).filter(Answer.response_id == resp.id).all()
            q_map = {
                q.id: q for q in db.query(Question).filter(Question.survey_id == survey.id).all()
            }
            for ans in answers:
                q = q_map.get(ans.question_id)
                answers_payload.append({
                    "question_id": ans.question_id,
                    "question_text": q.question_text if q else "",
                    "question_type": q.question_type if q else "",
                    "answer_value": ans.answer_value,
                })
            if resp.participant_id:
                history_rows = db.query(Response).filter(
                    Response.participant_id == resp.participant_id
                ).order_by(Response.completed_at.desc()).limit(8).all()
                for h in history_rows:
                    h_survey = db.query(Survey).filter(Survey.id == h.survey_id).first()
                    submission_history.append({
                        "response_id": h.id,
                        "survey_id": h.survey_id,
                        "survey_title": h_survey.title if h_survey else "",
                        "status": h.status,
                        "completed_at": h.completed_at.isoformat() if h.completed_at else None,
                        "client_ip": h.client_ip,
                    })

    display_rows = db.query(ResponseQualityCheck).filter(
        ResponseQualityCheck.survey_id == row.survey_id,
    ).order_by(ResponseQualityCheck.created_at.desc()).all()
    if row.source_type == "excel":
        display_rows = _latest_excel_quality_rows(
            [item for item in display_rows if item.source_type == "excel"]
        )
    else:
        display_rows = [item for item in display_rows if item.source_type == row.source_type]
    row_index = next((i for i, item in enumerate(display_rows) if item.id == row.id), 0)

    return JSONResponse({
        "id": row.id,
        "survey_id": row.survey_id,
        "survey_title": survey.title,
        "source_type": row.source_type,
        "source_ref": row.source_ref,
        "row_label": _quality_row_label(row_index),
        "response_id": row.response_id,
        "participant_email": participant_email,
        "participant_name": participant_name,
        "quality_score": row.quality_score,
        "quality_label": row.quality_label,
        "fraud_risk": row.fraud_risk,
        "rule_penalty": row.rule_penalty,
        "anomaly_score": row.anomaly_score,
        "semantic_risk": row.semantic_risk,
        "triggered_rules": row.triggered_rules or [],
        "reasons": row.reasons or [],
        "llm_result_json": row.llm_result_json,
        "review_status": row.review_status,
        "reviewer_label": row.reviewer_label,
        "notes": row.notes,
        "response_meta": response_meta,
        "answers": answers_payload,
        "raw_response_json": row.raw_response_json,
        "submission_history": submission_history,
        "created_at": row.created_at.isoformat() if row.created_at else None,
    })


@app.post("/api/quality-checks/{check_id}/review")
async def review_quality_check(
    check_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    row, survey = _get_quality_check_for_publisher(db, check_id, current_user.id)
    body = await request.json()
    action = (body.get("action") or "").strip().lower()
    notes = body.get("notes")
    reviewer_label = body.get("reviewer_label")

    response = db.query(Response).filter(Response.id == row.response_id).first() if row.response_id else None
    participant = db.query(User).filter(User.id == response.participant_id).first() if response and response.participant_id else None

    if action == "approve":
        row.review_status = "approved"
        if response:
            release_response_payout(db, response)
        if survey:
            _record_user_event(
                db,
                request,
                "participant_approved",
                user=current_user,
                target_type="survey",
                target_id=survey.id,
                metadata={
                    "study_title": survey.title,
                    "participant_id": response.participant_id if response else None,
                    "participant_email": participant.email if participant else None,
                    "source": "Quality review",
                    "user_role": "researcher",
                },
            )
    elif action == "reject":
        row.review_status = "rejected"
        if response:
            reject_response_payout(db, response)
        if survey:
            _record_user_event(
                db,
                request,
                "participant_rejected",
                user=current_user,
                target_type="survey",
                target_id=survey.id,
                metadata={
                    "study_title": survey.title,
                    "participant_id": response.participant_id if response else None,
                    "participant_email": participant.email if participant else None,
                    "source": "Quality review",
                    "user_role": "researcher",
                },
            )
    elif action == "pending":
        row.review_status = "pending"
        if response:
            return_response_to_review(db, response)
    elif action == "needs_review":
        row.review_status = "needs_review"
        if response:
            return_response_to_review(db, response)
    elif action == "mark_fraud":
        row.review_status = "rejected"
        row.fraud_risk = True
        row.quality_label = "fraud_risk"
        row.reviewer_label = reviewer_label or "fraud"
        if response:
            reject_response_payout(db, response)
        if survey:
            _record_user_event(
                db,
                request,
                "participant_rejected",
                user=current_user,
                target_type="survey",
                target_id=survey.id,
                metadata={
                    "study_title": survey.title,
                    "participant_id": response.participant_id if response else None,
                    "participant_email": participant.email if participant else None,
                    "source": "Quality review",
                    "user_role": "researcher",
                },
            )
    elif action == "mark_low_quality":
        row.review_status = "rejected"
        row.reviewer_label = reviewer_label or "low_quality"
        if row.quality_label == "high":
            row.quality_label = "low"
        if response:
            reject_response_payout(db, response)
        if survey:
            _record_user_event(
                db,
                request,
                "participant_rejected",
                user=current_user,
                target_type="survey",
                target_id=survey.id,
                metadata={
                    "study_title": survey.title,
                    "participant_id": response.participant_id if response else None,
                    "participant_email": participant.email if participant else None,
                    "source": "Quality review",
                    "user_role": "researcher",
                },
            )
    else:
        raise HTTPException(400, "Invalid action")

    if reviewer_label and action not in {"mark_fraud", "mark_low_quality"}:
        row.reviewer_label = reviewer_label
    if notes is not None:
        row.notes = notes

    db.commit()
    return JSONResponse({
        "message": "review updated",
        "id": row.id,
        "review_status": row.review_status,
        "reviewer_label": row.reviewer_label,
        "fraud_risk": row.fraud_risk,
        "quality_label": row.quality_label,
    })


@app.post("/api/quality-checks/{check_id}/blacklist")
async def blacklist_from_quality_check(
    check_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    row, _survey = _get_quality_check_for_publisher(db, check_id, current_user.id)
    body = await request.json()
    block_type = (body.get("block_type") or "").strip().lower()
    reason = body.get("reason") or "Marked from quality review"

    if block_type not in {"ip", "user", "device"}:
        raise HTTPException(400, "block_type must be ip, user, or device")

    block_value = None
    if row.response_id:
        resp = db.query(Response).filter(Response.id == row.response_id).first()
        if resp:
            if block_type == "ip":
                block_value = resp.client_ip
            elif block_type == "user":
                block_value = str(resp.participant_id)
            elif block_type == "device":
                block_value = resp.device_fingerprint

    if not block_value:
        raise HTTPException(400, "No value available to blacklist for this record")

    existing = db.query(QualityBlacklist).filter(
        QualityBlacklist.block_type == block_type,
        QualityBlacklist.block_value == block_value,
    ).first()
    if not existing:
        db.add(QualityBlacklist(
            block_type=block_type,
            block_value=block_value,
            reason=reason,
            created_by=current_user.id,
        ))

    row.fraud_risk = True
    row.review_status = "rejected"
    row.reviewer_label = row.reviewer_label or "blacklisted"
    db.commit()
    return JSONResponse({
        "message": "blacklist updated",
        "block_type": block_type,
        "block_value": block_value,
    })


def _parse_survey_auto_filter_from_form(form) -> tuple:
    """Read quality auto-filter defaults from publish/edit form."""
    enabled_raw = form.get("quality_auto_filter_enabled")
    if isinstance(enabled_raw, list):
        enabled_raw = enabled_raw[-1] if enabled_raw else "0"
    enabled = str(enabled_raw or "").strip().lower() in {"1", "on", "true", "yes"}
    min_score = _parse_auto_approve_min_score(form.get("quality_auto_filter_min_score", 80))
    return enabled, min_score


def _apply_survey_auto_filter_settings(survey: Survey, form) -> None:
    enabled, min_score = _parse_survey_auto_filter_from_form(form)
    survey.quality_auto_filter_enabled = enabled
    survey.quality_auto_filter_min_score = min_score


def _parse_auto_approve_min_score(value: Any) -> float:
    try:
        score = float(value)
    except (TypeError, ValueError):
        raise HTTPException(400, "Auto-approve score must be a number between 0 and 100.")
    return min(100.0, max(0.0, score))


@app.post("/api/surveys/{survey_id}/quality/auto-approve")
async def auto_approve_quality_by_score(
    survey_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    survey = db.query(Survey).filter(
        Survey.id == survey_id,
        Survey.publisher_id == current_user.id,
    ).first()
    if not survey:
        raise HTTPException(404, "Survey not found")

    body = await request.json()
    min_score = _parse_auto_approve_min_score(body.get("min_score", 80))
    scope = (body.get("scope") or "excel").strip().lower()

    rows = db.query(ResponseQualityCheck).filter(
        ResponseQualityCheck.survey_id == survey_id
    ).all()
    if scope == "builtin":
        rows = [row for row in rows if row.source_type == "builtin"]
    elif scope == "excel":
        rows = [row for row in rows if row.source_type == "excel"]

    approved, rejected = apply_auto_approve_checks(rows, min_score)
    response_ids = [row.response_id for row in rows if row.response_id]
    responses_by_id = {}
    if response_ids:
        responses_by_id = {
            response.id: response
            for response in db.query(Response).filter(Response.id.in_(response_ids)).all()
        }
    for row in rows:
        response = responses_by_id.get(row.response_id)
        if not response:
            continue
        if row.review_status == "approved":
            release_response_payout(db, response)
        elif row.review_status == "rejected":
            reject_response_payout(db, response)
    db.commit()
    return JSONResponse({
        "message": "auto filter applied",
        "approved_count": approved,
        "rejected_count": rejected,
        "min_score": min_score,
        "scope": scope,
    })


@app.post("/api/surveys/{survey_id}/quality/import-excel")
def import_excel_for_quality_check(
    survey_id: int,
    file: UploadFile = File(...),
    auto_approve: Optional[str] = Form(None),
    auto_approve_min_score: Optional[float] = Form(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    survey = db.query(Survey).filter(
        Survey.id == survey_id,
        Survey.publisher_id == current_user.id,
    ).first()
    if not survey:
        raise HTTPException(404, "Survey not found")

    filename = file.filename or "upload.xlsx"
    lower_name = filename.lower()
    if not (lower_name.endswith(".xlsx") or lower_name.endswith(".xlsm")):
        raise HTTPException(400, "Only .xlsx/.xlsm files are supported for now")

    try:
        from openpyxl import load_workbook
    except Exception:
        raise HTTPException(500, "openpyxl is required. Please install dependency first.")

    content = file.file.read()
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"Invalid Excel file: {exc}")

    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, "Excel should include a header row and at least one data row")
    if len(rows) - 1 > 500:
        raise HTTPException(400, "Excel has too many rows. Please upload 500 responses or fewer per file.")

    raw_headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    headers = [_normalize_excel_header(h) for h in raw_headers]
    if not any(headers):
        raise HTTPException(400, "Header row is empty")

    replaced_rows = db.query(ResponseQualityCheck).filter(
        ResponseQualityCheck.survey_id == survey_id,
        ResponseQualityCheck.source_type == "excel",
    ).delete(synchronize_session=False)

    questions = db.query(Question).filter(Question.survey_id == survey_id).all()
    question_map = {q.id: q for q in questions}

    header_to_qid = {}
    for idx, header in enumerate(headers):
        if not header:
            continue
        for q in questions:
            candidates = {
                str(q.id).strip().lower(),
                f"q_{q.id}".lower(),
                _normalize_excel_header(q.question_text),
            }
            if header in candidates:
                header_to_qid[idx] = q.id
                break

    historical = []
    completed = db.query(Response).filter(
        Response.survey_id == survey_id,
        Response.status == "completed",
        Response.started_at.isnot(None),
        Response.completed_at.isnot(None),
    ).all()
    for item in completed:
        sec = _duration_seconds_between(item.started_at, item.completed_at)
        if sec:
            historical.append(sec)

    duration_columns = {"duration", "duration_seconds", "time_spent", "completion_seconds"}
    parsed_rows = []

    for row_index, row in enumerate(rows[1:], start=2):
        if row is None:
            continue
        values = list(row)
        if not any(v is not None and str(v).strip() != "" for v in values):
            continue

        answers_by_qid = {}
        row_dict = {}
        for idx, value in enumerate(values):
            header_name = raw_headers[idx] if idx < len(raw_headers) else f"col_{idx+1}"
            safe_value = _json_safe_value(value)
            row_dict[header_name] = safe_value
            if idx in header_to_qid:
                answers_by_qid[header_to_qid[idx]] = safe_value

        duration_seconds = None
        for idx, header in enumerate(headers):
            if header in duration_columns and idx < len(values):
                duration_seconds = _value_as_float(values[idx])
                if duration_seconds is not None:
                    break

        row_question_map, row_answers, row_features = resolve_excel_row_context(
            row_dict=row_dict,
            mapped_question_map=question_map,
            mapped_answers=answers_by_qid,
            duration_seconds=duration_seconds,
        )
        parsed_rows.append({
            "row_index": row_index,
            "row_dict": row_dict,
            "question_map": row_question_map,
            "answers_by_qid": row_answers,
            "duration_seconds": duration_seconds,
            "features": row_features,
        })

    if not parsed_rows:
        raise HTTPException(400, "No data rows found in Excel.")

    all_features = [item["features"] for item in parsed_rows]
    batch_anomalies = batch_anomaly_scores_for_features(all_features)
    imported = 0
    by_label = {"high": 0, "medium": 0, "low": 0, "fraud_risk": 0}
    llm_rows_reviewed = 0
    llm_rows_failed = 0
    created_checks: List[ResponseQualityCheck] = []
    use_auto_approve = str(auto_approve or "").strip().lower() in {"1", "true", "yes", "on"}
    auto_min_score = _parse_auto_approve_min_score(auto_approve_min_score if use_auto_approve else 100)

    for index, item in enumerate(parsed_rows):
        result = compute_excel_row_quality(
            row_dict=item["row_dict"],
            mapped_question_map=item["question_map"],
            mapped_answers=item["answers_by_qid"],
            duration_seconds=item["duration_seconds"],
            historical_durations=historical,
            survey_title=survey.title,
            survey_description=survey.description,
            survey_reward=float(survey.reward_amount or 0.0),
            run_llm=True,
            precomputed_anomaly=batch_anomalies[index] if index < len(batch_anomalies) else None,
        )
        created_checks.append(create_excel_quality_check(
            db,
            survey_id=survey_id,
            source_ref=f"{filename}:row_{item['row_index']}",
            raw_response_json=item["row_dict"],
            result=result,
        ))
        if result.llm_result_json:
            if result.llm_result_json.get("error"):
                llm_rows_failed += 1
            elif not result.llm_result_json.get("skipped"):
                llm_rows_reviewed += 1
        by_label[result.quality_label] = by_label.get(result.quality_label, 0) + 1
        imported += 1

    if imported == 0:
        raise HTTPException(400, "No data rows found in Excel.")

    auto_approved_count = 0
    auto_rejected_count = 0
    if use_auto_approve:
        auto_approved_count, auto_rejected_count = apply_auto_approve_checks(created_checks, auto_min_score)

    db.commit()
    return JSONResponse({
        "message": "excel imported and quality checks generated",
        "survey_id": survey_id,
        "imported_rows": imported,
        "replaced_rows": replaced_rows,
        "label_distribution": by_label,
        "used_generic_columns": len(question_map) == 0,
        "auto_approve": {
            "enabled": use_auto_approve,
            "min_score": auto_min_score if use_auto_approve else None,
            "approved_count": auto_approved_count,
            "rejected_count": auto_rejected_count,
        },
        "llm_summary": {
            "rows_reviewed": llm_rows_reviewed,
            "rows_failed": llm_rows_failed,
            "api_key_configured": anthropic_api_key_configured(),
        },
    })




app.include_router(router)
